import os
import logging
import time
import json
from datetime import datetime, timedelta
import threading
from telebot import TeleBot, types, util
import requests
import urllib.parse
import qrcode
import sys
import io
from io import BytesIO
import webbrowser
import random
import zipfile
import shutil
from yookassa_integration import create_yookassa_payment, check_payment_status, create_payment_with_methods_menu
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Конфигурация бота
TOKEN = 'ваш токен'
ADMIN_ID = айди в тг ваш
bot = TeleBot(TOKEN, threaded=True, num_threads=4)
# Настройки канала
CHANNEL_USERNAME = ""  # Оставить пустым чтобы отключить проверку подписки
SUBSCRIPTION_WARNING_DAYS = 3  # За сколько дней предупреждать об окончании
VIDEO_INSTRUCTION_URL = "https://t.me/karachay_aj"

def reload_all_data():
    """Перезагружает все данные из файлов и синхронизирует ключи"""
    global users_db, payments_db, servers_db, payment_methods, PROMO_CODES
    
    try:
        # Загрузка данных пользователей
        if os.path.exists('users_db.json'):
            with open('users_db.json', 'r', encoding='utf-8') as f:
                users_db = json.load(f)
        else:
            users_db = {}
        
        # Загрузка данных платежей
        if os.path.exists('payments_db.json'):
            with open('payments_db.json', 'r', encoding='utf-8') as f:
                payments_db = json.load(f)
        else:
            payments_db = {}
        
        # Загрузка данных серверов
        if os.path.exists('servers_db.json'):
            with open('servers_db.json', 'r', encoding='utf-8') as f:
                servers_db = json.load(f)
        else:
            servers_db = {
                'server1': {
                    'name': '🇩🇪 Germany', 
                    'location': 'Germany, Frankfurt', 
                    'load': 'Low', 
                    'protocol': 'VLESS/V2Ray',
                    'ip': 'vpn-de1.example.com',
                    'available_keys': [],
                    'used_keys': {}
                },
                'server2': {
                    'name': '🇳🇱 Netherlands', 
                    'location': 'Netherlands, Amsterdam', 
                    'load': 'Low', 
                    'protocol': 'VLESS/V2Ray',
                    'ip': 'vpn-nl1.example.com',
                    'available_keys': [],
                    'used_keys': {}
                }
            }
        
        # Загрузка методов оплаты
        if os.path.exists('payment_methods.json'):
            with open('payment_methods.json', 'r', encoding='utf-8') as f:
                payment_methods = json.load(f)
        else:
            payment_methods = {}
        
        # Загрузка промокодов
        if os.path.exists('promo_codes.json'):
            with open('promo_codes.json', 'r', encoding='utf-8') as f:
                PROMO_CODES = json.load(f)
        else:
            PROMO_CODES = {}
        
        # ============ СИНХРОНИЗАЦИЯ КЛЮЧЕЙ ============
        # Получаем все ключи из файла keys_storage.json
        all_keys = []
        if os.path.exists('keys_storage.json'):
            with open('keys_storage.json', 'r', encoding='utf-8') as f:
                keys_storage = json.load(f)
                all_keys = keys_storage.get('keys', [])
        
        logger.info(f"=== СИНХРОНИЗАЦИЯ КЛЮЧЕЙ ===")
        logger.info(f"Найдено ключей в хранилище: {len(all_keys)}")
        
        # Группируем ключи по серверам
        germany_keys = []
        netherlands_keys = []
        unknown_keys = []
        
        for key_data in all_keys:
            key_string = key_data.get('key', '').lower()
            server_tag = key_data.get('server', '').lower()
            
            is_germany = False
            is_netherlands = False
            
            # Проверяем по тегу сервера или по содержимому ключа
            germany_patterns = ['grm', 'germany', 'de', 'ger', 'frankfurt', '🇩🇪']
            netherlands_patterns = ['ndr', 'netherlands', 'nl', 'ned', 'amsterdam', 'holland', '🇳🇱']
            
            for pattern in germany_patterns:
                if pattern in server_tag or pattern in key_string:
                    is_germany = True
                    break
            
            for pattern in netherlands_patterns:
                if pattern in server_tag or pattern in key_string:
                    is_netherlands = True
                    break
            
            if is_germany and not is_netherlands:
                germany_keys.append(key_data)
                logger.info(f"Ключ -> Германия")
            elif is_netherlands and not is_germany:
                netherlands_keys.append(key_data)
                logger.info(f"Ключ -> Нидерланды")
            else:
                unknown_keys.append(key_data)
                logger.warning(f"Ключ -> НЕ ОПРЕДЕЛЕН")
        
        logger.info(f"Германия: {len(germany_keys)} ключей")
        logger.info(f"Нидерланды: {len(netherlands_keys)} ключей")
        logger.info(f"Не определено: {len(unknown_keys)} ключей")
        
        # Обновляем данные для серверов
        for server_key, server_data in servers_db.items():
            server_name = server_data['name']
            
            if 'Germany' in server_name or '🇩🇪' in server_name:
                # Получаем используемые ключи
                used_keys_list = list(server_data.get('used_keys', {}).values())
                
                # Доступные ключи - те что есть в хранилище и не используются
                available_keys = [k for k in germany_keys if k not in used_keys_list]
                server_data['available_keys'] = available_keys
                logger.info(f"Сервер {server_name}: доступно {len(available_keys)}, используется {len(used_keys_list)}")
                
            elif 'Netherlands' in server_name or '🇳🇱' in server_name:
                used_keys_list = list(server_data.get('used_keys', {}).values())
                available_keys = [k for k in netherlands_keys if k not in used_keys_list]
                server_data['available_keys'] = available_keys
                logger.info(f"Сервер {server_name}: доступно {len(available_keys)}, используется {len(used_keys_list)}")
        
        # Сохраняем обновленные данные
        save_data_to_file()
        
        # Выводим итоговую статистику
        total_available = sum(len(s.get('available_keys', [])) for s in servers_db.values())
        total_used = sum(len(s.get('used_keys', {})) for s in servers_db.values())
        
        logger.info(f"=== ИТОГО: {total_available} доступных, {total_used} используемых ключей ===")
        
        return True
        
    except Exception as e:
        logger.error(f"Ошибка перезагрузки данных: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return False

# Промокоды
try:
    with open('promo_codes.json', 'r') as f:
        PROMO_CODES = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    PROMO_CODES = {}

# Принудительно устанавливаем stdout в UTF-8
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Настройка логгирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('vpn_bot.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Загрузка данных
try:
    with open('users_db.json', 'r') as f:
        users_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    users_db = {}

try:
    with open('payments_db.json', 'r') as f:
        payments_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    payments_db = {}

try:
    with open('servers_db.json', 'r') as f:
        servers_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
# Добавляем реальные серверы в базу
    servers_db = {
    'server1': {
        'name': '🇩🇪 Germany', 
        'location': 'Germany, Frankfurt', 
        'load': 'Low', 
        'protocol': 'WireGuard',
        'ip': 'vpn-de1.example.com',
        'available_configs': [],
        'used_configs': {}
    },
    'server2': {
        'name': '🇳🇱 Netherlands', 
        'location': 'Netherlands, Amsterdam', 
        'load': 'Low', 
        'protocol': 'WireGuard',
        'ip': 'vpn-nl1.example.com',
        'available_configs': [],
        'used_configs': {}
    }
}

# Функция для загрузки конфигов сервера
def load_server_configs(server_name, config_files):
    server_key = next(k for k, v in servers_db.items() if v['name'] == server_name)
    servers_db[server_key]['available_configs'] = config_files
    save_data_to_file()

def get_random_config(server_name, user_id):
    """Получает случайный ключ для сервера"""
    try:
        # Находим сервер по имени
        server_key = None
        for key, server_data in servers_db.items():
            clean_server_name = server_name.replace('🇩🇪', '').replace('🇳🇱', '').strip()
            clean_db_name = server_data['name'].replace('🇩🇪', '').replace('🇳🇱', '').strip()
            
            if clean_db_name.lower() == clean_server_name.lower():
                server_key = key
                break
            
            if ('germany' in clean_db_name.lower() and 'germany' in clean_server_name.lower()) or \
               ('netherlands' in clean_db_name.lower() and 'netherlands' in clean_server_name.lower()):
                server_key = key
                break
        
        if not server_key:
            logger.error(f"Сервер {server_name} не найден в базе")
            return None
        
        # Проверяем есть ли доступные ключи
        if not servers_db[server_key].get('available_keys', []):
            logger.error(f"Нет доступных ключей для сервера {server_name}")
            return None
            
        # Выбираем случайный ключ
        available_keys = servers_db[server_key]['available_keys']
        key_data = random.choice(available_keys)
        
        # Переносим ключ из доступных в используемые
        servers_db[server_key]['available_keys'].remove(key_data)
        if 'used_keys' not in servers_db[server_key]:
            servers_db[server_key]['used_keys'] = {}
        servers_db[server_key]['used_keys'][str(user_id)] = key_data
        save_data_to_file()
        
        logger.info(f"Выдан ключ для пользователя {user_id} на сервере {server_name}")
        return key_data
        
    except Exception as e:
        logger.error(f"Ошибка в get_random_config: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return None
    
try:
    with open('payment_methods.json', 'r') as f:
        payment_methods = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    payment_methods = {}
    
# Цены и периоды подписки
SUBSCRIPTION_PLANS = {
    '1 месяц': {'price': '100₽', 'days': 30},
    '3 месяца': {'price': '300₽', 'days': 90},
    '6 месяцев': {'price': '600₽', 'days': 180}
}

def debug_servers_status():
    """Отладочная функция для проверки статуса серверов"""
    logger.info("=== DEBUG: СТАТУС СЕРВЕРОВ ===")
    for server_key, server_data in servers_db.items():
        logger.info(f"Сервер: {server_data['name']} (ключ: {server_key})")
        logger.info(f"  Доступные ключи: {len(server_data.get('available_keys', []))}")
        logger.info(f"  Используемые ключи: {len(server_data.get('used_keys', {}))}")
    logger.info("=== КОНЕЦ DEBUG ===")
    
def save_referral_data():
    try:
        with open('referral_db.json', 'w', encoding='utf-8') as f:
            json.dump(referral_db, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error(f"Ошибка сохранения referral_db: {e}")
# Модифицируем функцию save_data_to_file для использования UTF-8 encoding
def save_data_to_file():
    """Сохраняет данные в JSON файлы"""
    try:
        # Создаем backup перед сохранением
        backup_files = ['users_db.json', 'payments_db.json', 'servers_db.json', 'payment_methods.json', 'promo_codes.json']
        for file in backup_files:
            if os.path.exists(file):
                shutil.copy2(file, f"{file}.backup")
        
        with open('users_db.json', 'w', encoding='utf-8') as f:
            json.dump(users_db, f, ensure_ascii=False, indent=2)
        with open('payments_db.json', 'w', encoding='utf-8') as f:
            json.dump(payments_db, f, ensure_ascii=False, indent=2)
        with open('servers_db.json', 'w', encoding='utf-8') as f:
            json.dump(servers_db, f, ensure_ascii=False, indent=2)
        with open('payment_methods.json', 'w', encoding='utf-8') as f:
            json.dump(payment_methods, f, ensure_ascii=False, indent=2)
        with open('promo_codes.json', 'w', encoding='utf-8') as f:
            json.dump(PROMO_CODES, f, ensure_ascii=False, indent=2)
        
        logger.info("Данные успешно сохранены")
    except Exception as e:
        logger.error(f"Ошибка сохранения данных: {e}")

# Модифицируем загрузку данных в начале файла для использования UTF-8
# Загрузка данных
try:
    with open('users_db.json', 'r', encoding='utf-8') as f:
        users_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    users_db = {}

try:
    with open('payments_db.json', 'r', encoding='utf-8') as f:
        payments_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    payments_db = {}

try:
    with open('servers_db.json', 'r', encoding='utf-8') as f:
        servers_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    # Добавляем реальные серверы в базу
    servers_db = {
        'server1': {
            'name': '🇩🇪 Germany', 
            'location': 'Germany, Frankfurt', 
            'load': 'Low', 
            'protocol': 'VLESS/V2Ray',
            'ip': 'vpn-de1.example.com',
            'available_keys': [],
            'used_keys': {}
        },
        'server2': {
            'name': '🇳🇱 Netherlands', 
            'location': 'Netherlands, Amsterdam', 
            'load': 'Low', 
            'protocol': 'VLESS/V2Ray',
            'ip': 'vpn-nl1.example.com',
            'available_keys': [],
            'used_keys': {}
        }
    }

# Сразу после загрузки выполняем синхронизацию
reload_all_data()

try:
    with open('payment_methods.json', 'r', encoding='utf-8') as f:
        payment_methods = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    payment_methods = {}

try:
    with open('promo_codes.json', 'r', encoding='utf-8') as f:
        PROMO_CODES = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    PROMO_CODES = {}
    
# Реферальная система
try:
    with open('referral_db.json', 'r', encoding='utf-8') as f:
        referral_db = json.load(f)
except (FileNotFoundError, json.JSONDecodeError):
    referral_db = {
        'allowed_users': {}, # {user_id: {'balance': 0, 'ref_code': 'abc123', 'added_by': admin_id}}
        'referrals': {},     # {ref_code: referrer_id}
        'withdraw_requests': []
    }

def save_user_data(user_id, data):
    users_db[str(user_id)] = data
    save_data_to_file()

def get_user_data(user_id):
    return users_db.get(str(user_id), {})

def save_payment(payment_id, data):
    payments_db[payment_id] = data
    save_data_to_file()

def get_payment(payment_id):
    return payments_db.get(payment_id)

def generate_payment_id():
    return str(random.randint(100, 999))  # 3 цифры

def is_admin(user_id):
    return str(user_id) == str(ADMIN_ID)

def subscription_monitor():
    """Фоновая задача для мониторинга подписок"""
    while True:
        try:
            # Проверяем подписки каждые 6 часов
            time.sleep(6 * 60 * 60)  # 6 часов в секундах
            
            # Проверяем уведомления об окончании
            check_and_send_subscription_warnings()
            
            # Проверяем истекшие подписки
            expired = check_expired_subscriptions()
            if expired:
                logger.info(f"Найдено {len(expired)} просроченных подписок")
                
        except Exception as e:
            logger.error(f"Ошибка в subscription_monitor: {e}")
            time.sleep(300)  # Ждем 5 минут при ошибке


def create_config_file(server_name, user_id=None, config_content=None):
    config_filename = f"{server_name}_{user_id}.conf" if user_id else f"{server_name}.conf"
    try:
        with open(config_filename, 'w') as f:
            if config_content:
                f.write(config_content)
            else:
                private_key = f"user_{user_id}_private_key" if user_id else "your_private_key"
                public_key = f"server_{server_name}_public_key"
                f.write(f"[Interface]\nPrivateKey = {private_key}\nAddress = 10.0.0.1/24\n\n[Peer]\nPublicKey = {public_key}\nAllowedIPs = 0.0.0.0/0\nEndpoint = {server_name}:51820")
        return config_filename
    except Exception as e:
        logger.error(f"Ошибка создания конфига: {e}")
        return None

def delete_previous_message(chat_id, message_id=None):
    try:
        if message_id and isinstance(message_id, int):
            bot.delete_message(chat_id, message_id)
    except Exception as e:
        error_msg = str(e).lower()
        # Игнорируем распространенные ошибки удаления
        if any(phrase in error_msg for phrase in [
            "message to delete not found",
            "message can't be deleted",
            "bad request: message can't be deleted"
        ]):
            logger.debug(f"Не удалось удалить сообщение (нормально): {e}")
        else:
            logger.error(f"Ошибка удаления сообщения: {e}")

def check_expired_subscriptions():
    """Проверяет истекшие подписки и отправляет уведомления пользователям и админу"""
    expired_users = []
    current_time = datetime.now()
    
    for user_id, user_data in users_db.items():
        if 'subscriptions' in user_data:
            for sub in user_data['subscriptions']:
                expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                
                # Проверяем истекла ли подписка
                if expiry_date < current_time:
                    # Проверяем не отправляли ли уже уведомление об окончании
                    if not sub.get('expiry_notification_sent', False):
                        expired_user_info = {
                            'user_id': user_id,
                            'username': user_data.get('username', 'N/A'),
                            'config_file': sub.get('config_file', 'N/A'),
                            'server': sub.get('server', 'N/A'),
                            'expiry_date': sub['expiry_date']
                        }
                        expired_users.append(expired_user_info)
                        
                        # Отправляем уведомление пользователю
                        send_subscription_expired_notification(user_id, sub)
                        
                        # Отправляем уведомление админу
                        admin_notification = f"""⚠️ ПРОСРОЧЕНА ПОДПИСКА ⚠️

👤 Пользователь: @{expired_user_info['username']} (ID: {user_id})
🖥 Сервер: {expired_user_info['server']}
🔑 Конфиг: {expired_user_info['config_file']}
📅 Истекла: {expired_user_info['expiry_date']}

Для проверки перейдите в раздел:
🔍 Просроченные подписки"""

                        try:
                            bot.send_message(ADMIN_ID, admin_notification)
                        except Exception as e:
                            logger.error(f"Ошибка отправки уведомления админу: {e}")
                        
                        # Помечаем что уведомление отправлено
                        sub['expiry_notification_sent'] = True
    
    # Сохраняем изменения
    if expired_users:
        save_data_to_file()
        logger.info(f"Найдено {len(expired_users)} просроченных подписок. Уведомления отправлены.")
    
    return expired_users

# Клавиатуры
def main_menu_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton('🛒 Купить VPN | Продлить')
    )
    keyboard.row(
        types.KeyboardButton('Активировать промокод'),
        types.KeyboardButton('🔑 Мои ключи')
    )
    keyboard.row(
        types.KeyboardButton('🛟 Поддержка'),
        types.KeyboardButton('👥 Рефералы'),
        types.KeyboardButton('🎬 Видео инструкция')
    )
    keyboard.add(
        types.KeyboardButton('Инструкция установки')
    )
    return keyboard

# Добавляем кнопку в админ-меню
def admin_menu_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton('📊 Статистика'),
        types.KeyboardButton('🔁 Статистика продлений'),
        types.KeyboardButton('📝 Список серверов'),
        types.KeyboardButton('📢 Рассылка'),
        types.KeyboardButton('🗂 Управление ключами'),
        types.KeyboardButton('👥 Список покупателей'),
        types.KeyboardButton('🎁 Управление промокодами'),
        types.KeyboardButton('🎁 Реферальная система'),
        types.KeyboardButton('💾 Создать резервную копию'),
        types.KeyboardButton('📥 Восстановить из копии'),
        types.KeyboardButton('📊 Экспорт в Excel'),
        types.KeyboardButton('🗑 Удалить пользователя'),
        types.KeyboardButton('👤 Список пользователей'),
        types.KeyboardButton('📤 Добавить ключи массово')
    )
    return keyboard
    
def export_to_excel():
    """Экспортирует все данные в Excel файл с красивым форматированием"""
    try:
        # Создаем папку для экспорта если ее нет
        if not os.path.exists('exports'):
            os.makedirs('exports')
        
        # Создаем имя файла с timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"exports/vpn_export_{timestamp}.xlsx"
        
        # Создаем Excel workbook
        wb = Workbook()
        
        # Удаляем дефолтный лист (проверяем что он есть)
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # ============ ЛИСТ 1: ПОЛЬЗОВАТЕЛИ ============
        ws_users = wb.create_sheet("Пользователи")
        
        # Заголовки
        users_headers = ['ID', 'Username', 'Подписок', 'Активных', 'Просрочено', 'Промокоды']
        for col_num, header in enumerate(users_headers, 1):
            cell = ws_users.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данные пользователей
        current_time = datetime.now()
        row_num = 2
        for user_id, user_data in users_db.items():
            total_subs = len(user_data.get('subscriptions', []))
            active_subs = 0
            expired_subs = 0
            
            for sub in user_data.get('subscriptions', []):
                try:
                    expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    if expiry_date > current_time:
                        active_subs += 1
                    else:
                        expired_subs += 1
                except:
                    pass
            
            used_promos = ', '.join(user_data.get('used_promo_codes', [])) or '-'
            
            ws_users.cell(row=row_num, column=1, value=user_id).alignment = cell_alignment
            ws_users.cell(row=row_num, column=2, value=user_data.get('username', 'N/A')).alignment = cell_alignment
            ws_users.cell(row=row_num, column=3, value=total_subs).alignment = cell_alignment
            ws_users.cell(row=row_num, column=4, value=active_subs).alignment = cell_alignment
            ws_users.cell(row=row_num, column=5, value=expired_subs).alignment = cell_alignment
            ws_users.cell(row=row_num, column=6, value=used_promos).alignment = cell_alignment
            row_num += 1
        
        # Ширина колонок
        ws_users.column_dimensions['A'].width = 15
        ws_users.column_dimensions['B'].width = 20
        ws_users.column_dimensions['C'].width = 10
        ws_users.column_dimensions['D'].width = 10
        ws_users.column_dimensions['E'].width = 12
        ws_users.column_dimensions['F'].width = 30
        
        # ============ ЛИСТ 2: ПОДПИСКИ ============
        ws_subs = wb.create_sheet("Подписки")
        
        # Заголовки
        subs_headers = ['User ID', 'Username', 'Сервер', 'Ключ', 'Куплено', 'Действует до', 'Статус', 'Тип']
        for col_num, header in enumerate(subs_headers, 1):
            cell = ws_subs.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данные подписок
        row_num = 2
        for user_id, user_data in users_db.items():
            username = user_data.get('username', 'N/A')
            
            for sub in user_data.get('subscriptions', []):
                try:
                    expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    is_active = expiry_date > current_time
                    status = 'Активна' if is_active else 'Истекла'
                except:
                    status = 'Ошибка'
                    expiry_date_str = sub.get('expiry_date', 'N/A')
                else:
                    expiry_date_str = expiry_date.strftime("%d.%m.%Y %H:%M")
                
                # Получаем ключ
                key_data = sub.get('key_data', {})
                key_value = key_data.get('key', '') if key_data else ''
                if key_value and len(key_value) > 50:
                    key_value = key_value[:47] + "..."
                
                ws_subs.cell(row=row_num, column=1, value=user_id).alignment = cell_alignment
                ws_subs.cell(row=row_num, column=2, value=username).alignment = cell_alignment
                ws_subs.cell(row=row_num, column=3, value=sub.get('server', 'N/A')).alignment = cell_alignment
                
                key_cell = ws_subs.cell(row=row_num, column=4, value=key_value or 'N/A')
                key_cell.alignment = cell_alignment
                
                ws_subs.cell(row=row_num, column=5, value=sub.get('purchase_date', 'N/A')).alignment = cell_alignment
                ws_subs.cell(row=row_num, column=6, value=expiry_date_str).alignment = cell_alignment
                
                status_cell = ws_subs.cell(row=row_num, column=7, value=status)
                status_cell.alignment = cell_alignment
                if status == 'Активна':
                    status_cell.font = Font(color="008000")
                elif status == 'Истекла':
                    status_cell.font = Font(color="FF0000")
                
                ws_subs.cell(row=row_num, column=8, value=sub.get('type', 'платная')).alignment = cell_alignment
                row_num += 1
        
        # Ширина колонок
        ws_subs.column_dimensions['A'].width = 15
        ws_subs.column_dimensions['B'].width = 20
        ws_subs.column_dimensions['C'].width = 18
        ws_subs.column_dimensions['D'].width = 55
        ws_subs.column_dimensions['E'].width = 20
        ws_subs.column_dimensions['F'].width = 18
        ws_subs.column_dimensions['G'].width = 10
        ws_subs.column_dimensions['H'].width = 12
        
        # ============ ЛИСТ 3: ПЛАТЕЖИ ============
        ws_payments = wb.create_sheet("Платежи")
        
        # Заголовки
        pay_headers = ['ID платежа', 'User ID', 'Username', 'Сервер', 'Срок', 'Сумма', 'Статус', 'Создан']
        for col_num, header in enumerate(pay_headers, 1):
            cell = ws_payments.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данные платежей
        row_num = 2
        for payment_id, payment_data in payments_db.items():
            status = payment_data.get('status', 'pending')
            status_text = {'approved': 'Оплачен', 'pending': 'Ожидает', 'rejected': 'Отклонен'}.get(status, status)
            
            ws_payments.cell(row=row_num, column=1, value=payment_id).alignment = cell_alignment
            ws_payments.cell(row=row_num, column=2, value=str(payment_data.get('user_id', 'N/A'))).alignment = cell_alignment
            ws_payments.cell(row=row_num, column=3, value=payment_data.get('username', 'N/A')).alignment = cell_alignment
            ws_payments.cell(row=row_num, column=4, value=payment_data.get('server', 'N/A')).alignment = cell_alignment
            ws_payments.cell(row=row_num, column=5, value=payment_data.get('duration', 'N/A')).alignment = cell_alignment
            ws_payments.cell(row=row_num, column=6, value=payment_data.get('amount', 'N/A')).alignment = cell_alignment
            
            status_cell = ws_payments.cell(row=row_num, column=7, value=status_text)
            status_cell.alignment = cell_alignment
            if status == 'approved':
                status_cell.font = Font(color="008000")
            elif status == 'rejected':
                status_cell.font = Font(color="FF0000")
            elif status == 'pending':
                status_cell.font = Font(color="FFA500")
            
            ws_payments.cell(row=row_num, column=8, value=payment_data.get('timestamp', 'N/A')).alignment = cell_alignment
            row_num += 1
        
        # Ширина колонок
        ws_payments.column_dimensions['A'].width = 12
        ws_payments.column_dimensions['B'].width = 15
        ws_payments.column_dimensions['C'].width = 20
        ws_payments.column_dimensions['D'].width = 18
        ws_payments.column_dimensions['E'].width = 15
        ws_payments.column_dimensions['F'].width = 10
        ws_payments.column_dimensions['G'].width = 12
        ws_payments.column_dimensions['H'].width = 20
        
        # ============ ЛИСТ 4: СЕРВЕРЫ ============
        ws_servers = wb.create_sheet("Серверы")
        
        # Заголовки
        serv_headers = ['Сервер', 'Локация', 'Протокол', 'IP', 'Доступно', 'Используется', 'Всего']
        for col_num, header in enumerate(serv_headers, 1):
            cell = ws_servers.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данные серверов
        row_num = 2
        for server_key, server_data in servers_db.items():
            available = len(server_data.get('available_keys', []))
            used = len(server_data.get('used_keys', {}))
            
            ws_servers.cell(row=row_num, column=1, value=server_data.get('name', 'N/A')).alignment = cell_alignment
            ws_servers.cell(row=row_num, column=2, value=server_data.get('location', 'N/A')).alignment = cell_alignment
            ws_servers.cell(row=row_num, column=3, value=server_data.get('protocol', 'VLESS')).alignment = cell_alignment
            ws_servers.cell(row=row_num, column=4, value=server_data.get('ip', 'N/A')).alignment = cell_alignment
            ws_servers.cell(row=row_num, column=5, value=available).alignment = cell_alignment
            ws_servers.cell(row=row_num, column=6, value=used).alignment = cell_alignment
            ws_servers.cell(row=row_num, column=7, value=available + used).alignment = cell_alignment
            row_num += 1
        
        # Ширина колонок
        ws_servers.column_dimensions['A'].width = 18
        ws_servers.column_dimensions['B'].width = 25
        ws_servers.column_dimensions['C'].width = 15
        ws_servers.column_dimensions['D'].width = 20
        ws_servers.column_dimensions['E'].width = 12
        ws_servers.column_dimensions['F'].width = 14
        ws_servers.column_dimensions['G'].width = 10
        
        # ============ ЛИСТ 5: КЛЮЧИ ============
        ws_keys = wb.create_sheet("Ключи")
        
        # Заголовки
        keys_headers = ['№', 'Сервер', 'Название', 'Ключ', 'Добавлен', 'Статус']
        for col_num, header in enumerate(keys_headers, 1):
            cell = ws_keys.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Загружаем ключи
        all_keys = []
        if os.path.exists('keys_storage.json'):
            try:
                with open('keys_storage.json', 'r', encoding='utf-8') as f:
                    keys_storage = json.load(f)
                    all_keys = keys_storage.get('keys', [])
            except:
                all_keys = []
        
        # Используемые ключи
        used_keys_set = set()
        for server in servers_db.values():
            for key_data in server.get('used_keys', {}).values():
                if key_data and key_data.get('key'):
                    used_keys_set.add(key_data.get('key'))
        
        # Данные ключей
        row_num = 2
        for i, key_data in enumerate(all_keys, 1):
            key_string = key_data.get('key', '')
            status = 'Используется' if key_string in used_keys_set else 'Свободен'
            
            key_preview = key_string[:50] + '...' if len(key_string) > 50 else key_string
            
            ws_keys.cell(row=row_num, column=1, value=i).alignment = cell_alignment
            ws_keys.cell(row=row_num, column=2, value=key_data.get('server', 'N/A')).alignment = cell_alignment
            ws_keys.cell(row=row_num, column=3, value=key_data.get('name', '-')).alignment = cell_alignment
            
            key_cell = ws_keys.cell(row=row_num, column=4, value=key_preview or 'N/A')
            key_cell.alignment = cell_alignment
            
            ws_keys.cell(row=row_num, column=5, value=key_data.get('added_at', 'N/A')).alignment = cell_alignment
            
            status_cell = ws_keys.cell(row=row_num, column=6, value=status)
            status_cell.alignment = cell_alignment
            if status == 'Свободен':
                status_cell.font = Font(color="008000")
            else:
                status_cell.font = Font(color="FFA500")
            
            row_num += 1
        
        # Ширина колонок
        ws_keys.column_dimensions['A'].width = 8
        ws_keys.column_dimensions['B'].width = 18
        ws_keys.column_dimensions['C'].width = 20
        ws_keys.column_dimensions['D'].width = 55
        ws_keys.column_dimensions['E'].width = 20
        ws_keys.column_dimensions['F'].width = 14
        
        # ============ ЛИСТ 6: ПРОМОКОДЫ ============
        ws_promo = wb.create_sheet("Промокоды")
        
        # Заголовки
        promo_headers = ['Код', 'Сервер', 'Дней', 'Создан']
        for col_num, header in enumerate(promo_headers, 1):
            cell = ws_promo.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данные промокодов
        row_num = 2
        for code, promo_info in PROMO_CODES.items():
            ws_promo.cell(row=row_num, column=1, value=code).alignment = cell_alignment
            ws_promo.cell(row=row_num, column=2, value=promo_info.get('server', 'N/A')).alignment = cell_alignment
            ws_promo.cell(row=row_num, column=3, value=promo_info.get('days', 'N/A')).alignment = cell_alignment
            ws_promo.cell(row=row_num, column=4, value=promo_info.get('created_at', 'N/A')).alignment = cell_alignment
            row_num += 1
        
        # Ширина колонок
        ws_promo.column_dimensions['A'].width = 20
        ws_promo.column_dimensions['B'].width = 18
        ws_promo.column_dimensions['C'].width = 10
        ws_promo.column_dimensions['D'].width = 20
        
        # ============ ЛИСТ 7: СТАТИСТИКА ============
        ws_stats = wb.create_sheet("Статистика")
        
        # Заголовки
        ws_stats.cell(row=1, column=1, value="Показатель").font = header_font
        ws_stats.cell(row=1, column=1).fill = header_fill
        ws_stats.cell(row=1, column=1).alignment = header_alignment
        
        ws_stats.cell(row=1, column=2, value="Значение").font = header_font
        ws_stats.cell(row=1, column=2).fill = header_fill
        ws_stats.cell(row=1, column=2).alignment = header_alignment
        
        # Данные статистики
        total_users = len(users_db)
        approved_payments = len([p for p in payments_db.values() if p.get('status') == 'approved'])
        pending_payments = len([p for p in payments_db.values() if p.get('status') == 'pending'])
        
        revenue = 0
        for p in payments_db.values():
            if p.get('status') == 'approved':
                amount_str = p.get('amount', '0')
                try:
                    amount_num = int(''.join(filter(str.isdigit, str(amount_str))))
                    revenue += amount_num
                except:
                    pass
        
        total_available = sum(len(s.get('available_keys', [])) for s in servers_db.values())
        total_used = sum(len(s.get('used_keys', {})) for s in servers_db.values())
        
        stats_data = [
            ('Всего пользователей', total_users),
            ('Всего платежей', len(payments_db)),
            ('Подтверждено платежей', approved_payments),
            ('Ожидает платежей', pending_payments),
            ('Общий доход', f"{revenue}₽"),
            ('Доступно ключей', total_available),
            ('Используется ключей', total_used),
            ('Всего ключей', total_available + total_used),
            ('Активных промокодов', len(PROMO_CODES)),
            ('Дата экспорта', datetime.now().strftime("%d.%m.%Y %H:%M"))
        ]
        
        row_num = 2
        for name, value in stats_data:
            ws_stats.cell(row=row_num, column=1, value=name).alignment = cell_alignment
            ws_stats.cell(row=row_num, column=1).font = Font(bold=True)
            ws_stats.cell(row=row_num, column=2, value=value).alignment = cell_alignment
            row_num += 1
        
        # Ширина колонок
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 20
        
        # Сохраняем файл
        wb.save(excel_filename)
        wb.close()
        
        logger.info(f"Excel отчет успешно создан: {excel_filename}")
        return excel_filename
        
    except Exception as e:
        logger.error(f"Ошибка экспорта в Excel: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return None
    
# Добавляем функцию создания резервной копии
def create_backup():
    """Создает резервную копию всех данных"""
    try:
        # Создаем папку для резервных копий если ее нет
        if not os.path.exists('backups'):
            os.makedirs('backups')
        
        # Создаем имя файла с timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"backups/backup_{timestamp}.zip"
        
        # Файлы для резервного копирования
        files_to_backup = [
            'users_db.json', 
            'payments_db.json', 
            'servers_db.json',
            'payment_methods.json',
            'promo_codes.json'
        ]
        
        # Добавляем конфигурационные файлы
        config_files = [f for f in os.listdir() if f.endswith('.conf')]
        files_to_backup.extend(config_files)
        
        # Создаем zip архив
        with zipfile.ZipFile(backup_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file in files_to_backup:
                if os.path.exists(file):
                    zipf.write(file)
        
        return backup_filename
    except Exception as e:
        logger.error(f"Ошибка создания резервной копии: {e}")
        return None
        
# Добавляем функцию восстановления из резервной копии
def restore_from_backup(backup_file):
    """Восстанавливает данные из резервной копии с merge"""
    try:
        # Временная папка для распаковки
        temp_dir = 'temp_restore'
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        os.makedirs(temp_dir)
        
        # Распаковываем архив
        with zipfile.ZipFile(backup_file, 'r') as zipf:
            zipf.extractall(temp_dir)
        
        # Функция для безопасного merge JSON данных
        def merge_json_data(current_data, new_data, key_field=None):
            """Безопасное слияние JSON данных"""
            if isinstance(current_data, dict) and isinstance(new_data, dict):
                # Для словарей объединяем ключи
                result = current_data.copy()
                for key, value in new_data.items():
                    if key in result:
                        if isinstance(result[key], (dict, list)) and isinstance(value, (dict, list)):
                            result[key] = merge_json_data(result[key], value, key_field)
                        else:
                            result[key] = value
                    else:
                        result[key] = value
                return result
            elif isinstance(current_data, list) and isinstance(new_data, list) and key_field:
                # Для списков объединяем по ключу
                result = current_data.copy()
                existing_keys = {item[key_field]: item for item in result if key_field in item}
                
                for new_item in new_data:
                    if key_field in new_item and new_item[key_field] in existing_keys:
                        # Обновляем существующий элемент
                        index = next(i for i, item in enumerate(result) 
                                   if item.get(key_field) == new_item[key_field])
                        result[index] = merge_json_data(result[index], new_item, key_field)
                    else:
                        # Добавляем новый элемент
                        result.append(new_item)
                return result
            else:
                # Для простых типов или несовместимых структур возвращаем новые данные
                return new_data
        
        # Восстанавливаем JSON файлы с merge
        json_files = {
            'users_db.json': 'user_id',
            'payments_db.json': None,  # Простая замена
            'servers_db.json': 'name',
            'payment_methods.json': 'bank',
            'promo_codes.json': None   # Простая замена
        }
        
        for json_file, merge_key in json_files.items():
            temp_file = os.path.join(temp_dir, json_file)
            if os.path.exists(temp_file):
                with open(temp_file, 'r', encoding='utf-8') as f:
                    new_data = json.load(f)
                
                current_file = json_file
                if os.path.exists(current_file):
                    with open(current_file, 'r', encoding='utf-8') as f:
                        current_data = json.load(f)
                    
                    if merge_key:
                        merged_data = merge_json_data(current_data, new_data, merge_key)
                    else:
                        merged_data = new_data  # Простая замена
                else:
                    merged_data = new_data
                
                # Сохраняем объединенные данные
                with open(current_file, 'w', encoding='utf-8') as f:
                    json.dump(merged_data, f, ensure_ascii=False, indent=2)
        
        # Восстанавливаем конфигурационные файлы
        for file in os.listdir(temp_dir):
            if file.endswith('.conf') and not os.path.exists(file):
                shutil.move(os.path.join(temp_dir, file), file)
        
        # Очищаем временную папку
        shutil.rmtree(temp_dir)
        
        # Перезагружаем данные в память
        reload_all_data()
        
        return True
    except Exception as e:
        logger.error(f"Ошибка восстановления из резервной копии: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return False
        
# Добавляем функцию перезагрузки всех данных


def servers_menu_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    keyboard.add(types.KeyboardButton('🇩🇪 Германия'))
    keyboard.add(types.KeyboardButton('🇳🇱 Нидерланды'))
    keyboard.add(types.KeyboardButton('🔙 Назад'))
    return keyboard

def payment_methods_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    
    # Группируем методы по банкам
    bank_groups = {}
    for method in payment_methods.values():
        if method['bank'] not in bank_groups:
            bank_groups[method['bank']] = []
        bank_groups[method['bank']].append(method)
    
    # Создаем кнопки для каждого банка (без указания карт)
    buttons = [types.KeyboardButton(bank) for bank in bank_groups.keys()]
    keyboard.add(*buttons)
    keyboard.add(types.KeyboardButton('🔙 Назад'))
    
    return keyboard

def payment_verification_keyboard(payment_id):
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton(f'✅ Подтвердить {payment_id}'),
        types.KeyboardButton(f'❌ Отклонить {payment_id}'),
        types.KeyboardButton('🔙 Назад')
    )
    return keyboard

def duration_menu_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton('1 месяц - 100₽'),
        types.KeyboardButton('3 месяца - 300₽'),
        types.KeyboardButton('6 месяцев - 600₽'),
        types.KeyboardButton('🔙 Назад')
    )
    return keyboard

def config_actions_keyboard(config_path):
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton('📲 Установить приложение'),
        types.KeyboardButton(f'⚙️ Импортировать {os.path.basename(config_path)}'),
        types.KeyboardButton('💾 Скачать конфиг'),
        types.KeyboardButton('🔙 Назад')
    )
    return keyboard

def config_management_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton('📤 Загрузить новый конфиг'),
        types.KeyboardButton('🗑 Удалить конфиг'),
        types.KeyboardButton('🔙 Назад')
    )
    return keyboard

def customers_list_keyboard():
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    keyboard.add(
        types.KeyboardButton('🔍 Просроченные подписки'),
        types.KeyboardButton('🔙 Назад')
    )
    return keyboard
    
def check_channel_subscription(user_id):
    try:
        if not CHANNEL_USERNAME:  # Если канал не указан, пропускаем проверку
            return True
            
        member = bot.get_chat_member(f"@{CHANNEL_USERNAME}", user_id)
        return member.status in ['member', 'administrator', 'creator']
    except Exception as e:
        logger.error(f"Ошибка проверки подписки: {e}")
        return False

def check_and_send_subscription_warnings():
    """Проверяет подписки и отправляет уведомления об скором окончании"""
    try:
        current_time = datetime.now()
        warned_users = set()  # Чтобы не дублировать уведомления
        
        for user_id, user_data in users_db.items():
            if 'subscriptions' not in user_data:
                continue
                
            for sub in user_data['subscriptions']:
                expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                days_left = (expiry_date - current_time).days
                
                # Проверяем нужно ли отправлять уведомление
                if 0 < days_left <= SUBSCRIPTION_WARNING_DAYS:
                    # Проверяем не отправляли ли уже уведомление за сегодня
                    warning_key = f"warning_{user_id}_{sub['config_file']}_{days_left}"
                    
                    if warning_key not in sub.get('last_warnings', []):
                        # Отправляем уведомление
                        try:
                            warning_text = f"""⚠️ <b>Внимание!</b>

Ваша подписка на сервере {sub['server']} заканчивается через <b>{days_left}</b> дней.

Дата окончания: {expiry_date.strftime('%d.%m.%Y')}

Чтобы продолжить пользоваться VPN без перерывов, продлите подписку заранее.

🔄 <b>Продлить сейчас:</b> /buy"""
                            
                            bot.send_message(
                                user_id,
                                warning_text,
                                parse_mode='HTML'
                            )
                            
                            # Сохраняем информацию об отправленном уведомлении
                            if 'last_warnings' not in sub:
                                sub['last_warnings'] = []
                            sub['last_warnings'].append(warning_key)
                            warned_users.add(user_id)
                            
                        except Exception as e:
                            logger.error(f"Ошибка отправки уведомления пользователю {user_id}: {e}")
        
        # Сохраняем данные если были отправлены уведомления
        if warned_users:
            save_data_to_file()
            logger.info(f"Отправлены уведомления об окончании подписки для {len(warned_users)} пользователей")
            
    except Exception as e:
        logger.error(f"Ошибка в check_and_send_subscription_warnings: {e}")
        
@bot.message_handler(func=lambda message: message.text == '🔄 Синхронизировать конфиги' and is_admin(message.from_user.id))
def sync_configs_command(message):
    """Кнопка синхронизации конфигов"""
    sync_configs(message)
        
def send_subscription_expired_notification(user_id, subscription):
    """Отправляет уведомление об окончании подписки"""
    try:
        expiry_date = datetime.strptime(subscription['expiry_date'], "%Y-%m-%d %H:%M:%S")
        
        expired_text = f"""❌ <b>Подписка закончилась</b>

Ваша подписка на сервере {subscription['server']} закончилась {expiry_date.strftime('%d.%m.%Y')}.

Для возобновления работы VPN приобретите новую подписку.

🛒 <b>Приобрести новую подписку:</b> /buy"""

        bot.send_message(
            user_id,
            expired_text,
            parse_mode='HTML'
        )
        
    except Exception as e:
        logger.error(f"Ошибка отправки уведомления об окончании подписки пользователю {user_id}: {e}")
        

        
def generate_qr_code(config_content, filename):
    try:
        # Создаем QR-код
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(config_content)
        qr.make(fit=True)
        
        # Создаем изображение QR-кода
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Сохраняем изображение
        img.save(filename)
        return filename
    except Exception as e:
        logger.error(f"Ошибка генерации QR-кода: {e}")
        return None
        
@bot.message_handler(func=lambda message: message.text == '👤 Список пользователей' and is_admin(message.from_user.id))
def users_list_handler(message):
    """Показывает список всех пользователей с подписками"""
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        if not users_db:
            bot.send_message(message.chat.id, "❌ В базе нет пользователей.")
            return
        
        current_time = datetime.now()
        users_with_subscriptions = []
        
        # Собираем информацию о пользователях с подписками
        for user_id, user_data in users_db.items():
            if 'subscriptions' in user_data and user_data['subscriptions']:
                user_info = {
                    'user_id': user_id,
                    'username': user_data.get('username', 'N/A'),
                    'subscriptions': []
                }
                
                for sub in user_data['subscriptions']:
                    try:
                        expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                        days_left = (expiry_date - current_time).days
                        is_expired = expiry_date < current_time
                    except:
                        days_left = 0
                        is_expired = True
                    
                    # Получаем ключ
                    key_data = sub.get('key_data', {})
                    key_value = key_data.get('key', 'N/A') if isinstance(key_data, dict) else 'N/A'
                    key_preview = key_value[:30] + '...' if len(key_value) > 30 else key_value
                    
                    subscription_info = {
                        'server': sub.get('server', 'N/A'),
                        'key_preview': key_preview,
                        'purchase_date': sub.get('purchase_date', 'N/A'),
                        'expiry_date': sub.get('expiry_date', 'N/A'),
                        'days_left': days_left,
                        'is_expired': is_expired,
                        'expired_days': (current_time - expiry_date).days if is_expired else 0
                    }
                    user_info['subscriptions'].append(subscription_info)
                
                users_with_subscriptions.append(user_info)
        
        if not users_with_subscriptions:
            bot.send_message(message.chat.id, "❌ В базе нет пользователей с подписками.")
            return
        
        # Сортируем
        def sort_key(user):
            if not user['subscriptions']:
                return (1, 9999)
            has_expired = any(sub['is_expired'] for sub in user['subscriptions'])
            if has_expired:
                expired_subs = [sub for sub in user['subscriptions'] if sub['is_expired']]
                max_expired_days = max(sub['expired_days'] for sub in expired_subs)
                return (0, -max_expired_days)
            else:
                min_days_left = min(sub['days_left'] for sub in user['subscriptions'])
                return (1, min_days_left)
        
        users_with_subscriptions.sort(key=sort_key)
        
        # Статистика
        total_users = len(users_with_subscriptions)
        users_with_expired = sum(1 for user in users_with_subscriptions if any(sub['is_expired'] for sub in user['subscriptions']))
        users_with_active = total_users - users_with_expired
        
        stats_text = f"👥 <b>Список пользователей с подписками</b>\n\n"
        stats_text += f"📊 Пользователей: {total_users}\n"
        stats_text += f"✅ Активных: {users_with_active}\n"
        stats_text += f"⚠️ Просроченных: {users_with_expired}\n\n"
        stats_text += "────────────────────\n\n"
        
        bot.send_message(message.chat.id, stats_text, parse_mode='HTML')
        time.sleep(1)
        
        # Отправляем пользователей частями
        users_per_message = 2
        total_parts = (len(users_with_subscriptions) - 1) // users_per_message + 1
        
        for part_num in range(total_parts):
            start_idx = part_num * users_per_message
            end_idx = start_idx + users_per_message
            chunk = users_with_subscriptions[start_idx:end_idx]
            
            text = f"📄 <b>Часть {part_num + 1}/{total_parts}</b>\n\n"
            
            for user in chunk:
                text += f"🆔 <b>ID:</b> <code>{user['user_id']}</code>\n"
                text += f"👤 <b>Username:</b> @{user['username']}\n"
                
                for i, sub in enumerate(user['subscriptions'], 1):
                    status_icon = "❌" if sub['is_expired'] else "✅"
                    status_text = f"<b>ПРОСРОЧЕН</b> ({sub['expired_days']} дн.)" if sub['is_expired'] else f"активен ({sub['days_left']} дн.)"
                    
                    text += f"\n<b>Подписка #{i}:</b>\n"
                    text += f"   {status_icon} <b>Сервер:</b> {sub['server']}\n"
                    text += f"   🔑 <b>Ключ:</b> <code>{sub['key_preview']}</code>\n"
                    text += f"   📅 <b>До:</b> {sub['expiry_date']}\n"
                    text += f"   📊 <b>Статус:</b> {status_text}\n"
                
                text += "────────────────────\n\n"
            
            if len(text) > 4000:
                lines = text.split('\n')
                current_part = ""
                for line in lines:
                    if len(current_part + line + '\n') > 4000:
                        bot.send_message(message.chat.id, current_part, parse_mode='HTML')
                        current_part = line + '\n'
                        time.sleep(2)
                    else:
                        current_part += line + '\n'
                if current_part:
                    bot.send_message(message.chat.id, current_part, parse_mode='HTML')
                    time.sleep(2)
            else:
                bot.send_message(message.chat.id, text, parse_mode='HTML')
                if part_num < total_parts - 1:
                    time.sleep(2)
        
        # Инструкция
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('🗑 Удалить пользователя'))
        keyboard.add(types.KeyboardButton('🗑 Удалить конкретную подписку'))
        keyboard.add(types.KeyboardButton('🔍 Просроченные подписки'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(
            message.chat.id,
            "💡 <b>Действия:</b>\n"
            "• Для удаления пользователя используйте ID\n"
            "• Для удаления подписки используйте её данные\n"
            "• ❌ - просроченные, ✅ - активные",
            parse_mode='HTML',
            reply_markup=keyboard
        )
        
    except Exception as e:
        logger.error(f"Ошибка в users_list_handler: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        bot.send_message(message.chat.id, "❌ Ошибка при загрузке списка пользователей")
        
@bot.message_handler(func=lambda message: message.text == '🎬 Видео инструкция')
def video_instruction_handler(message):
    """Отправляет видео инструкцию по настройке VPN"""
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        # URL видео инструкции - замените на свой
        VIDEO_URL = "https://t.me/karachay_aj"
        
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton(
            "📺 Смотреть видеоинструкцию",
            url=VIDEO_URL
        ))
        
        instruction_text = """🎬 <b>Видео инструкция по настройке VPN</b>

Нажмите на кнопку ниже, чтобы посмотреть подробную видеоинструкцию по установке и настройке VPN на вашем устройстве.

В видео показано:
• Как скачать приложение
• Как импортировать ключ
• Как подключиться к VPN

Если останутся вопросы, обращайтесь в поддержку 🛟"""
        
        bot.send_message(
            user_id,
            instruction_text,
            parse_mode='HTML',
            reply_markup=markup
        )
        
        # Возвращаем главное меню
        bot.send_message(
            user_id,
            "Выберите следующее действие:",
            reply_markup=main_menu_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка в video_instruction_handler: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при загрузке видео инструкции")
        
@bot.message_handler(func=lambda message: message.text == 'Инструкция установки')
def installation_instructions(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        instructions_text = """📋 <b>Инструкция по установке VPN:</b>

1️⃣ <b>Скачайте приложение:</b>
• <b>Android:</b> Happ или V2RayTun
• <b>iOS:</b> Happ или V2RayTun

2️⃣ <b>Купите подписку</b> 🛒 или <b>активируйте промокод</b> 🎫 в боте

3️⃣ <b>Скопируйте полученный ключ</b> 🔑

4️⃣ <b>Вставьте ключ в приложение:</b>
• Откройте приложение
• Нажмите «+» или «Добавить»
• Выберите «Импорт из буфера обмена»
• Ключ добавится автоматически

5️⃣ <b>Нажмите «Подключиться»</b> ✅

Готово! 🚀"""

        # Отправляем фото с инструкцией если есть
        try:
            with open('Inst.png', 'rb') as photo:
                bot.send_photo(
                    user_id,
                    photo,
                    caption=instructions_text,
                    parse_mode='HTML'
                )
        except FileNotFoundError:
            bot.send_message(
                user_id,
                instructions_text,
                parse_mode='HTML'
            )
        
        # Кнопки для скачивания
        app_markup = types.InlineKeyboardMarkup(row_width=2)
        app_markup.add(
            types.InlineKeyboardButton("📱 Happ Android", url="https://play.google.com/store/apps/details?id=com.happproxy"),
            types.InlineKeyboardButton("📱 Happ iOS", url="https://apps.apple.com/us/app/happ-proxy-utility/id6504287215?l=ru"),
            types.InlineKeyboardButton("📱 V2raytun Android", url="https://play.google.com/store/apps/details?id=com.v2raytun.android"),
            types.InlineKeyboardButton("📱 V2raytun iOS", url="https://apps.apple.com/us/app/v2raytun/id6476628951?l=ru")
        )
        
        bot.send_message(
            user_id,
            "📥 <b>Скачать приложение:</b>",
            parse_mode='HTML',
            reply_markup=app_markup
        )
        
        # Добавляем кнопку "Назад" (главное меню) чтобы пользователь не терялся
        bot.send_message(
            user_id,
            "После установки приложения вернитесь в главное меню:",
            reply_markup=main_menu_keyboard()
        )
            
    except Exception as e:
        logger.error(f"Ошибка в installation_instructions: {e}")
        back_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        back_markup.add(types.KeyboardButton('🔙 Назад'))
        bot.send_message(
            message.chat.id,
            "❌ Произошла ошибка при загрузке инструкции.\n\n"
            "Нажмите «Назад» чтобы вернуться в главное меню.",
            reply_markup=back_markup
        )
        
@bot.message_handler(func=lambda message: message.text == '🔗 ссылка на приложение')
def download_app_handler(message):
    """Обработчик кнопки скачивания приложения после активации промокода"""
    try:
        user_id = message.from_user.id
        
        download_text = "📥 <b>Скачать приложение Happ:</b>"
        markup = types.InlineKeyboardMarkup()
        markup.row(
            types.InlineKeyboardButton("📱 Happ Android", url="https://play.google.com/store/apps/details?id=com.happproxy"),
            types.InlineKeyboardButton("📱 Happ iOS", url="https://apps.apple.com/us/app/happ-proxy-utility/id6504287215?l=ru"),
            types.InlineKeyboardButton("📱 V2raytun Android", url="https://play.google.com/store/apps/details?id=com.v2raytun.android"),
            types.InlineKeyboardButton("📱 V2raytun iOS", url="https://apps.apple.com/us/app/v2raytun/id6476628951?l=ru")
        )
        
        bot.send_message(
            user_id,
            download_text,
            parse_mode='HTML',
            reply_markup=markup
        )
        

        
    except Exception as e:
        logger.error(f"Ошибка в download_app_handler: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при отправке ссылок на приложение.")
        
@bot.message_handler(func=lambda message: not check_channel_subscription(message.from_user.id))
def require_subscription(message):
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Подписаться", url=f"https://t.me/{CHANNEL_USERNAME}"))
    
    bot.send_message(
        message.chat.id,
        "Для использования бота необходимо подписаться на наш канал:",
        reply_markup=markup
    )
    
# Добавляем обработчик для кнопки экспорта
@bot.message_handler(func=lambda message: message.text == '📊 Экспорт в Excel' and is_admin(message.from_user.id))
def export_excel_handler(message):
    try:
        bot.send_message(message.chat.id, "📊 Создаю Excel отчет... Это может занять несколько секунд.")
        
        excel_file = export_to_excel()
        if excel_file:
            with open(excel_file, 'rb') as f:
                bot.send_document(
                    message.chat.id,
                    f,
                    caption="✅ Excel отчет успешно создан!\n\n"
                           "📋 Содержит листы:\n"
                           "• Пользователи\n"
                           "• Подписки\n"
                           "• Платежи\n"
                           "• Серверы\n"
                           "• Промокоды",
                    visible_file_name=os.path.basename(excel_file)
                )
        else:
            bot.send_message(message.chat.id, "❌ Ошибка создания Excel отчета")
    except Exception as e:
        logger.error(f"Ошибка в export_excel_handler: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка создания Excel отчета")
        
    # Обработчики команд
@bot.message_handler(commands=['start', 'help'])
def start_handler(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        # ==================== ОБРАБОТКА РЕФЕРАЛЬНОЙ ССЫЛКИ ====================
        invited_by = None
        if message.text and len(message.text.split()) > 1:
            ref_code = message.text.split()[1]
            logger.info(f"🔗 РЕФЕРАЛ: Пользователь {user_id} перешел по реф коду: {ref_code}")
            
            # Проверяем существование реферального кода
            if 'referrals' in referral_db and ref_code in referral_db['referrals']:
                referrer_id = referral_db['referrals'][ref_code]
                
                # Проверяем, что пользователь не приглашает сам себя
                if str(user_id) != referrer_id:
                    # Проверяем, что реферер активен в программе
                    if referrer_id in referral_db.get('allowed_users', {}):
                        # Инициализируем структуры если нужно
                        if 'relations' not in referral_db:
                            referral_db['relations'] = {}
                        
                        # Сохраняем связь, только если ее еще нет
                        if str(user_id) not in referral_db['relations']:
                            referral_db['relations'][str(user_id)] = referrer_id
                            save_referral_data()
                            invited_by = referrer_id
                            logger.info(f"🔗 РЕФЕРАЛ: Пользователь {user_id} привязан к рефереру {referrer_id}")
                            
                            # Получаем username реферера для уведомления
                            referrer_data = get_user_data(referrer_id)
                            referrer_username = referrer_data.get('username', 'Пользователь')
                            
                            # Отправляем красивое уведомление приглашенному пользователю
                            welcome_text = (
                                f"🌟 <b>Добро пожаловать в Aj VPN!</b>\n\n"
                                f"🎁 Вас пригласил пользователь @{referrer_username}\n"
                                f"💫 Приятного использования!\n\n"
                                f"🔐 <b>Ваши преимущества:</b>\n"
                                f"• Высокая скорость соединения\n"
                                f"• Стабильная работа 24/7\n"
                                f"• Доступ к любым сайтам и сервисам\n\n"
                                f"👇 Выберите действие в меню ниже:"
                            )
                            bot.send_message(user_id, welcome_text, parse_mode='HTML')
                            
                            # Уведомляем реферера о новом приглашенном
                            try:
                                ref_notification = (
                                    f"🎉 <b>Новый реферал!</b>\n\n"
                                    f"👤 По вашей ссылке присоединился новый пользователь!\n"
                                    f"💡 Когда он совершит первую покупку, вы получите <b>25%</b> на ваш баланс.\n\n"
                                    f"📊 Статистика в разделе «👥 Рефералы»"
                                )
                                bot.send_message(referrer_id, ref_notification, parse_mode='HTML')
                                logger.info(f"🔗 РЕФЕРАЛ: Реферер {referrer_id} уведомлен о новом реферале")
                            except Exception as ref_err:
                                logger.error(f"🔗 РЕФЕРАЛ: Ошибка уведомления реферера: {ref_err}")
                        else:
                            logger.info(f"🔗 РЕФЕРАЛ: Связь для пользователя {user_id} уже существует")
                    else:
                        logger.warning(f"🔗 РЕФЕРАЛ: Реферер {referrer_id} не в allowed_users (код недействителен)")
                else:
                    # Пользователь пытается пригласить сам себя
                    bot.send_message(
                        user_id,
                        "⚠️ <b>Нельзя приглашать самого себя!</b>\n\n"
                        "Поделитесь вашей реферальной ссылкой с друзьями, чтобы получать бонусы.",
                        parse_mode='HTML'
                    )
                    logger.warning(f"🔗 РЕФЕРАЛ: Пользователь {user_id} попытался пригласить сам себя")
            else:
                logger.warning(f"🔗 РЕФЕРАЛ: Недействительный реферальный код: {ref_code}")
        # ==================== КОНЕЦ ОБРАБОТКИ РЕФЕРАЛЬНОЙ ССЫЛКИ ====================
        
        # Проверяем есть ли активные подписки с истекающим сроком
        user_data = get_user_data(user_id)
        if 'subscriptions' in user_data:
            current_time = datetime.now()
            for sub in user_data['subscriptions']:
                expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                days_left = (expiry_date - current_time).days
                
                if 0 < days_left <= 7:  # Показываем уведомление если осталось меньше недели
                    bot.send_message(
                        user_id,
                        f"ℹ️ У вас есть активная подписка на {sub['server']} "
                        f"(осталось {days_left} дней). Для продления нажмите /buy",
                        reply_markup=main_menu_keyboard()
                    )
        
        # Если пользователь был приглашен, пропускаем стандартное приветствие
        if invited_by:
            bot.send_message(user_id, "Выберите действие:", reply_markup=main_menu_keyboard())
            return
            
        # Стандартное приветствие для обычных пользователей
        if is_admin(user_id):
            bot.send_message(user_id, "👋 Добро пожаловать в админ-панель!", reply_markup=admin_menu_keyboard())
        else:
            # Отправляем фото с новым приветствием
            try:
                with open('welcome.png', 'rb') as photo:
                    bot.send_photo(
                        user_id, 
                        photo,
                        caption="""Привет, это Aj VPN — лучший VPN, созданный специально для тебя! 

VPN который можно не выключать""",
                        reply_markup=main_menu_keyboard()
                    )
            except FileNotFoundError:
                bot.send_message(
                    user_id,
                    """Привет, это Aj VPN — лучший VPN, созданный специально для тебя! 

VPN который можно не выключать""",
                    reply_markup=main_menu_keyboard()
                )
    except Exception as e:
        logger.error(f"Ошибка в start_handler: {e}")

@bot.message_handler(func=lambda message: message.text == '🔙 В главное меню' and is_admin(message.from_user.id))
def back_to_main_admin(message):
    start_handler(message)

@bot.message_handler(func=lambda message: message.text == '🔙 Назад')
def back_handler(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        if is_admin(user_id):
            bot.send_message(user_id, "Возвращаемся в админ-меню", reply_markup=admin_menu_keyboard())
        else:
            bot.send_message(user_id, "Возвращаемся в главное меню", reply_markup=main_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в back_handler: {e}")


@bot.message_handler(func=lambda message: message.text == '🇩🇪 Германия')
def select_germany_server(message):
    """Обработчик выбора сервера Германия"""
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        server_name = '🇩🇪 Germany'
        
        # Находим сервер и проверяем наличие ключей
        server_key = None
        for key, server_data in servers_db.items():
            if 'Germany' in server_data['name']:
                server_key = key
                break
        
        if server_key:
            available_keys = servers_db[server_key].get('available_keys', [])
            if not available_keys:
                # Нет доступных ключей
                bot.send_message(
                    user_id,
                    f"❌ <b>На сервере {server_name} нет доступных ключей!</b>\n\n"
                    "Попробуйте выбрать другой сервер или подождите пока администратор добавит новые ключи.",
                    parse_mode='HTML',
                    reply_markup=servers_menu_keyboard()
                )
                
                # Уведомляем админа
                bot.send_message(
                    ADMIN_ID,
                    f"⚠️ <b>ВНИМАНИЕ!</b>\n"
                    f"Пользователь @{message.from_user.username} (ID: {user_id}) "
                    f"пытался выбрать сервер {server_name}, но ключи закончились!",
                    parse_mode='HTML'
                )
                return
        
        # Есть ключи - продолжаем
        user_data = get_user_data(user_id)
        user_data['selected_server'] = server_name
        save_user_data(user_id, user_data)
        
        logger.info(f"Пользователь {user_id} выбрал сервер: {server_name}")
        
        bot.send_message(
            user_id, 
            f"✅ Выбран сервер: {server_name}\n\nВыберите срок подписки:", 
            reply_markup=duration_menu_keyboard()
        )
    except Exception as e:
        logger.error(f"Ошибка в select_germany_server: {e}")
        bot.send_message(message.chat.id, "❌ Произошла ошибка. Попробуйте снова.", reply_markup=main_menu_keyboard())


@bot.message_handler(func=lambda message: message.text == '🇳🇱 Нидерланды')
def select_netherlands_server(message):
    """Обработчик выбора сервера Нидерланды"""
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        server_name = '🇳🇱 Netherlands'
        
        # Находим сервер и проверяем наличие ключей
        server_key = None
        for key, server_data in servers_db.items():
            if 'Netherlands' in server_data['name']:
                server_key = key
                break
        
        if server_key:
            available_keys = servers_db[server_key].get('available_keys', [])
            if not available_keys:
                # Нет доступных ключей
                bot.send_message(
                    user_id,
                    f"❌ <b>На сервере {server_name} нет доступных ключей!</b>\n\n"
                    "Попробуйте выбрать другой сервер или подождите пока администратор добавит новые ключи.",
                    parse_mode='HTML',
                    reply_markup=servers_menu_keyboard()
                )
                
                # Уведомляем админа
                bot.send_message(
                    ADMIN_ID,
                    f"⚠️ <b>ВНИМАНИЕ!</b>\n"
                    f"Пользователь @{message.from_user.username} (ID: {user_id}) "
                    f"пытался выбрать сервер {server_name}, но ключи закончились!",
                    parse_mode='HTML'
                )
                return
        
        # Есть ключи - продолжаем
        user_data = get_user_data(user_id)
        user_data['selected_server'] = server_name
        save_user_data(user_id, user_data)
        
        logger.info(f"Пользователь {user_id} выбрал сервер: {server_name}")
        
        bot.send_message(
            user_id, 
            f"✅ Выбран сервер: {server_name}\n\nВыберите срок подписки:", 
            reply_markup=duration_menu_keyboard()
        )
    except Exception as e:
        logger.error(f"Ошибка в select_netherlands_server: {e}")
        bot.send_message(message.chat.id, "❌ Произошла ошибка. Попробуйте снова.", reply_markup=main_menu_keyboard())
        
@bot.message_handler(func=lambda message: message.text == '🛒 Купить VPN | Продлить')
def buy_or_extend_vpn(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        # Проверяем есть ли доступные ключи на серверах
        available_keys = False
        for server in servers_db.values():
            if server.get('available_keys'):
                available_keys = True
                break
        
        if not available_keys:
            bot.send_message(
                user_id,
                "❌ <b>В настоящее время нет доступных VPN ключей.</b>\n\n"
                "Администратор уже уведомлен и скоро добавит новые ключи.\n"
                "Пожалуйста, попробуйте позже.",
                parse_mode='HTML',
                reply_markup=main_menu_keyboard()
            )
            
            bot.send_message(
                ADMIN_ID,
                "⚠️ <b>ВНИМАНИЕ: ЗАКОНЧИЛИСЬ КЛЮЧИ!</b>\n\n"
                "Пользователь пытался купить VPN, но ключи закончились.\n"
                "Срочно добавьте новые ключи!",
                parse_mode='HTML'
            )
            return
        
        user_data = get_user_data(user_id)
        active_subs = []
        
        # Проверяем активные подписки
        current_time = datetime.now()
        for sub in user_data.get('subscriptions', []):
            try:
                expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                if expiry_date > current_time:
                    active_subs.append(sub)
            except:
                pass
        
        if active_subs:
            # Создаем клавиатуру с кнопками продления
            keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
            for sub in active_subs:
                keyboard.add(types.KeyboardButton(f'🔄 Продлить {sub["server"]}'))
            keyboard.add(types.KeyboardButton('➕ Купить новый ключ'))
            keyboard.add(types.KeyboardButton('🔙 Назад'))
            
            bot.send_message(
                user_id,
                "У вас есть активные подписки. Вы можете продлить существующие или купить новый ключ:",
                reply_markup=keyboard
            )
        else:
            # Нет активных подписок - покупаем новую
            bot.send_message(user_id, "Выберите сервер:", reply_markup=servers_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в buy_or_extend_vpn: {e}")
        bot.send_message(user_id, "❌ Произошла ошибка. Попробуйте позже.")
        
@bot.message_handler(func=lambda message: message.text == '🗂 Управление ключами' and is_admin(message.from_user.id))
def keys_management(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        keyboard.add(
            types.KeyboardButton('📤 Загрузить новый ключ'),
            types.KeyboardButton('🗑 Удалить ключ'),
            types.KeyboardButton('📋 Просмотреть ключи'),
            types.KeyboardButton('🔙 Назад')
        )
        bot.send_message(message.chat.id, "Управление VPN ключами:", reply_markup=keyboard)
    except Exception as e:
        logger.error(f"Ошибка в keys_management: {e}")
        
@bot.message_handler(func=lambda message: message.text == '📋 Просмотреть ключи' and is_admin(message.from_user.id))
def view_keys(message):
    """Просмотр всех ключей в системе"""
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        # Загружаем ключи из хранилища
        all_keys = []
        if os.path.exists('keys_storage.json'):
            with open('keys_storage.json', 'r', encoding='utf-8') as f:
                keys_storage = json.load(f)
                all_keys = keys_storage.get('keys', [])
        
        if not all_keys:
            bot.send_message(message.chat.id, "❌ В хранилище нет ключей.", reply_markup=admin_menu_keyboard())
            return
        
        # Группируем ключи по серверам
        germany_keys = []
        netherlands_keys = []
        other_keys = []
        
        for key_data in all_keys:
            server = key_data.get('server', '').lower()
            if 'germany' in server or 'de' in server or '🇩🇪' in server:
                germany_keys.append(key_data)
            elif 'netherlands' in server or 'nl' in server or '🇳🇱' in server:
                netherlands_keys.append(key_data)
            else:
                other_keys.append(key_data)
        
        # Создаем клавиатуру для фильтрации
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        if germany_keys:
            keyboard.add(types.KeyboardButton(f'🇩🇪 Германия ({len(germany_keys)})'))
        if netherlands_keys:
            keyboard.add(types.KeyboardButton(f'🇳🇱 Нидерланды ({len(netherlands_keys)})'))
        if other_keys:
            keyboard.add(types.KeyboardButton(f'❓ Другие ({len(other_keys)})'))
        keyboard.add(types.KeyboardButton('📋 Все ключи'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        text = f"📋 <b>Ключи в хранилище</b>\n\n"
        text += f"📊 <b>Всего ключей:</b> {len(all_keys)}\n"
        text += f"🇩🇪 Германия: {len(germany_keys)}\n"
        text += f"🇳🇱 Нидерланды: {len(netherlands_keys)}\n"
        text += f"❓ Другие: {len(other_keys)}\n\n"
        text += "Выберите категорию для просмотра:"
        
        msg = bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_view_keys_filter)
        
    except Exception as e:
        logger.error(f"Ошибка в view_keys: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())

def process_view_keys_filter(message):
    """Обрабатывает выбор фильтра для просмотра ключей"""
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Возврат в меню", reply_markup=admin_menu_keyboard())
            return
        
        # Загружаем ключи
        with open('keys_storage.json', 'r', encoding='utf-8') as f:
            keys_storage = json.load(f)
            all_keys = keys_storage.get('keys', [])
        
        # Фильтруем ключи
        filtered_keys = []
        filter_name = ""
        
        if 'Германия' in message.text or '🇩🇪' in message.text:
            for key_data in all_keys:
                server = key_data.get('server', '').lower()
                if 'germany' in server or 'de' in server or '🇩🇪' in server:
                    filtered_keys.append(key_data)
            filter_name = "Германия"
        elif 'Нидерланды' in message.text or '🇳🇱' in message.text:
            for key_data in all_keys:
                server = key_data.get('server', '').lower()
                if 'netherlands' in server or 'nl' in server or '🇳🇱' in server:
                    filtered_keys.append(key_data)
            filter_name = "Нидерланды"
        elif 'Другие' in message.text or '❓' in message.text:
            for key_data in all_keys:
                server = key_data.get('server', '').lower()
                if not ('germany' in server or 'de' in server or '🇩🇪' in server or
                       'netherlands' in server or 'nl' in server or '🇳🇱' in server):
                    filtered_keys.append(key_data)
            filter_name = "Другие"
        else:
            filtered_keys = all_keys
            filter_name = "Все"
        
        if not filtered_keys:
            bot.send_message(message.chat.id, f"❌ Нет ключей в категории '{filter_name}'", 
                           reply_markup=admin_menu_keyboard())
            return
        
        # Отправляем ключи частями
        chunk_size = 5
        total_parts = (len(filtered_keys) - 1) // chunk_size + 1
        
        for part in range(total_parts):
            start = part * chunk_size
            end = start + chunk_size
            chunk = filtered_keys[start:end]
            
            text = f"📋 <b>Ключи - {filter_name} (часть {part + 1}/{total_parts})</b>\n\n"
            
            for i, key_data in enumerate(chunk, start + 1):
                key_string = key_data.get('key', 'N/A')
                key_name = key_data.get('name', 'Без имени')
                added_at = key_data.get('added_at', 'N/A')
                
                # Обрезаем ключ для отображения
                if len(key_string) > 50:
                    key_preview = key_string[:50] + "..."
                else:
                    key_preview = key_string
                
                text += f"<b>#{i}</b> {key_name}\n"
                text += f"🔑 <code>{key_preview}</code>\n"
                text += f"📅 Добавлен: {added_at}\n"
                text += f"────────────────────\n\n"
            
            # Добавляем кнопки для копирования ключа
            markup = types.InlineKeyboardMarkup(row_width=1)
            for idx, key_data in enumerate(chunk, start):
                key_string = key_data.get('key', '')
                if len(key_string) > 64:
                    short_key = key_string[:30] + "..." + key_string[-20:]
                else:
                    short_key = key_string
                markup.add(types.InlineKeyboardButton(
                    f"📋 Копировать ключ #{idx + 1}", 
                    callback_data=f"show_key_{idx}"
                ))
            
            bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=markup)
            time.sleep(1)
        
        # Клавиатура для возврата
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        keyboard.add(types.KeyboardButton('📋 Просмотреть ключи'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(message.chat.id, "Выберите действие:", reply_markup=keyboard)
        
    except Exception as e:
        logger.error(f"Ошибка в process_view_keys_filter: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())

@bot.callback_query_handler(func=lambda call: call.data.startswith('show_key_'))
def show_full_key_callback(call):
    """Показывает полный ключ при нажатии на кнопку"""
    try:
        # Извлекаем индекс ключа
        key_index = int(call.data.replace('show_key_', ''))
        
        # Загружаем ключи
        with open('keys_storage.json', 'r', encoding='utf-8') as f:
            keys_storage = json.load(f)
            all_keys = keys_storage.get('keys', [])
        
        if key_index >= len(all_keys):
            bot.answer_callback_query(call.id, "❌ Ключ не найден")
            return
        
        key_data = all_keys[key_index]
        key_string = key_data.get('key', '')
        key_name = key_data.get('name', 'Без имени')
        
        # Отправляем полный ключ
        key_text = f"🔑 <b>Ключ #{key_index + 1}: {key_name}</b>\n\n"
        key_text += f"<pre><code class=\"language-text\">{key_string}</code></pre>\n\n"
        key_text += "<i>Нажмите на ключ чтобы скопировать</i>"
        
        bot.send_message(call.message.chat.id, key_text, parse_mode='HTML')
        bot.answer_callback_query(call.id, "✅ Ключ отправлен")
        
    except Exception as e:
        logger.error(f"Ошибка в show_full_key_callback: {e}")
        bot.answer_callback_query(call.id, "❌ Ошибка")
        
@bot.message_handler(func=lambda message: message.text == '📤 Загрузить новый ключ' and is_admin(message.from_user.id))
def upload_single_key(message):
    """Загрузка одного нового ключа"""
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        # Выбор сервера
        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        for server in servers_db.values():
            keyboard.add(types.KeyboardButton(f"🔑 {server['name']}"))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "🖥 <b>Выберите сервер для ключа:</b>",
            parse_mode='HTML',
            reply_markup=keyboard
        )
        bot.register_next_step_handler(msg, process_key_server_selection)
        
    except Exception as e:
        logger.error(f"Ошибка в upload_single_key: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())

def process_key_server_selection(message):
    """Обрабатывает выбор сервера для ключа"""
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        server_name = message.text.replace('🔑 ', '')
        
        # Проверяем существование сервера
        server_key = next((k for k, v in servers_db.items() if v['name'] == server_name), None)
        if not server_key:
            bot.send_message(message.chat.id, "❌ Сервер не найден", reply_markup=admin_menu_keyboard())
            return
        
        # Сохраняем выбранный сервер
        user_data = get_user_data(message.from_user.id)
        user_data['temp_key_server'] = server_name
        save_user_data(message.from_user.id, user_data)
        
        # Запрашиваем ключ
        msg = bot.send_message(
            message.chat.id,
            f"🔑 <b>Введите VPN ключ для {server_name}</b>\n\n"
            f"<b>Поддерживаемые форматы:</b>\n"
            f"• <code>vless://uuid@server:port?params#name</code>\n"
            f"• <code>vmess://base64_encoded</code>\n"
            f"• <code>trojan://password@server:port#name</code>\n"
            f"• <code>ss://base64_encoded</code>\n\n"
            f"<i>Отправьте ключ одним сообщением</i>",
            parse_mode='HTML',
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(msg, process_key_input)
        
    except Exception as e:
        logger.error(f"Ошибка в process_key_server_selection: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())

def process_key_input(message):
    """Обрабатывает ввод ключа"""
    try:
        if message.text == '🔙 Назад' or message.text == '/cancel':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        key_string = message.text.strip()
        
        # Проверяем формат ключа
        valid_formats = ['vless://', 'vmess://', 'trojan://', 'ss://', 'ssr://', 'hysteria://', 'tuic://', 'https://', 'http://']
        is_valid = any(key_string.lower().startswith(fmt) for fmt in valid_formats)
        
        if not is_valid:
            msg = bot.send_message(
                message.chat.id,
                "❌ Неверный формат ключа!\n\n"
                "Ключ должен начинаться с:\n"
                "• vless://\n• vmess://\n• trojan://\n• ss://\n• https://\n\n"
                "Попробуйте снова:",
                reply_markup=types.ReplyKeyboardRemove()
            )
            bot.register_next_step_handler(msg, process_key_input)
            return
        
        # Получаем сохраненный сервер
        user_data = get_user_data(message.from_user.id)
        server_name = user_data.get('temp_key_server', 'Unknown')
        
        # Запрашиваем имя ключа
        msg = bot.send_message(
            message.chat.id,
            "📝 <b>Введите название ключа (опционально):</b>\n\n"
            "Например: 'Германия 1' или 'Франкфурт'\n"
            "Или отправьте <code>/skip</code> для авто-названия",
            parse_mode='HTML',
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(msg, process_key_name, key_string, server_name)
        
    except Exception as e:
        logger.error(f"Ошибка в process_key_input: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())

def process_key_name(message, key_string, server_name):
    """Обрабатывает ввод имени ключа и сохраняет его"""
    try:
        if message.text == '/skip':
            key_name = f"VPN-{server_name.replace('🇩🇪', '').replace('🇳🇱', '').strip()}"
        else:
            key_name = message.text.strip()
        
        # Создаем объект ключа
        key_data = {
            'key': key_string,
            'name': key_name,
            'server': server_name,
            'added_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'added_by': str(message.from_user.id)
        }
        
        # Загружаем существующие ключи
        if os.path.exists('keys_storage.json'):
            with open('keys_storage.json', 'r', encoding='utf-8') as f:
                keys_storage = json.load(f)
        else:
            keys_storage = {'keys': []}
        
        # Добавляем ключ
        keys_storage['keys'].append(key_data)
        
        # Сохраняем
        with open('keys_storage.json', 'w', encoding='utf-8') as f:
            json.dump(keys_storage, f, ensure_ascii=False, indent=2)
        
        # Добавляем в доступные ключи сервера
        server_key = next((k for k, v in servers_db.items() if v['name'] == server_name), None)
        if server_key:
            if 'available_keys' not in servers_db[server_key]:
                servers_db[server_key]['available_keys'] = []
            servers_db[server_key]['available_keys'].append(key_data)
            save_data_to_file()
        
        # Очищаем временные данные
        user_data = get_user_data(message.from_user.id)
        if 'temp_key_server' in user_data:
            del user_data['temp_key_server']
        save_user_data(message.from_user.id, user_data)
        
        # Показываем результат
        key_preview = key_string[:50] + '...' if len(key_string) > 50 else key_string
        
        result_text = (
            f"✅ <b>Ключ успешно добавлен!</b>\n\n"
            f"📝 <b>Название:</b> {key_name}\n"
            f"🖥 <b>Сервер:</b> {server_name}\n"
            f"🔑 <b>Ключ:</b> <code>{key_preview}</code>\n"
            f"📅 <b>Добавлен:</b> {key_data['added_at']}"
        )
        
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        keyboard.add(types.KeyboardButton('📤 Загрузить новый ключ'))
        keyboard.add(types.KeyboardButton('📋 Просмотреть ключи'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(message.chat.id, result_text, parse_mode='HTML', reply_markup=keyboard)
        
    except Exception as e:
        logger.error(f"Ошибка в process_key_name: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка сохранения: {str(e)}", reply_markup=admin_menu_keyboard())
        
@bot.message_handler(func=lambda message: message.text == '🗑 Удалить ключ' and is_admin(message.from_user.id))
def delete_key(message):
    """Удаление ключа из системы"""
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        # Загружаем ключи
        if not os.path.exists('keys_storage.json'):
            bot.send_message(message.chat.id, "❌ В хранилище нет ключей.", reply_markup=admin_menu_keyboard())
            return
        
        with open('keys_storage.json', 'r', encoding='utf-8') as f:
            keys_storage = json.load(f)
            all_keys = keys_storage.get('keys', [])
        
        if not all_keys:
            bot.send_message(message.chat.id, "❌ В хранилище нет ключей.", reply_markup=admin_menu_keyboard())
            return
        
        # Фильтруем только свободные ключи (не используемые)
        free_keys = []
        used_keys_set = set()
        
        # Собираем используемые ключи
        for server in servers_db.values():
            for key_data in server.get('used_keys', {}).values():
                if key_data and key_data.get('key'):
                    used_keys_set.add(key_data.get('key'))
        
        for i, key_data in enumerate(all_keys):
            if key_data.get('key') not in used_keys_set:
                key_data['_index'] = i
                free_keys.append(key_data)
        
        if not free_keys:
            bot.send_message(message.chat.id, "❌ Нет свободных ключей для удаления (все используются).", 
                           reply_markup=admin_menu_keyboard())
            return
        
        # Создаем клавиатуру с ключами
        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        
        # Показываем первые 20 ключей
        for key_data in free_keys[:20]:
            key_name = key_data.get('name', 'Без имени')
            server = key_data.get('server', 'Unknown')
            key_preview = key_data.get('key', '')[:30] + '...'
            btn_text = f"🗑 {key_name} ({server})"
            keyboard.add(types.KeyboardButton(btn_text))
        
        if len(free_keys) > 20:
            keyboard.add(types.KeyboardButton('📄 Показать еще...'))
        keyboard.add(types.KeyboardButton('🔍 Поиск по ключу'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        # Сохраняем список свободных ключей в данных пользователя
        user_data = get_user_data(message.from_user.id)
        user_data['free_keys_list'] = free_keys
        save_user_data(message.from_user.id, user_data)
        
        text = f"🗑 <b>Выберите ключ для удаления</b>\n\n"
        text += f"📊 Свободных ключей: {len(free_keys)} / {len(all_keys)}\n"
        text += f"🔐 Используется: {len(used_keys_set)}\n\n"
        text += "<i>Внимание! Удаление ключа невозможно отменить.</i>"
        
        msg = bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_key_deletion)
        
    except Exception as e:
        logger.error(f"Ошибка в delete_key: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())

def process_key_deletion(message):
    """Обрабатывает выбор ключа для удаления"""
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        if message.text == '🔍 Поиск по ключу':
            msg = bot.send_message(
                message.chat.id,
                "🔍 <b>Введите часть ключа для поиска:</b>",
                parse_mode='HTML',
                reply_markup=types.ReplyKeyboardRemove()
            )
            bot.register_next_step_handler(msg, process_key_search_for_delete)
            return
        
        if message.text == '📄 Показать еще...':
            # Показываем следующую порцию ключей
            user_data = get_user_data(message.from_user.id)
            free_keys = user_data.get('free_keys_list', [])
            
            keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
            for key_data in free_keys[20:40]:
                key_name = key_data.get('name', 'Без имени')
                server = key_data.get('server', 'Unknown')
                btn_text = f"🗑 {key_name} ({server})"
                keyboard.add(types.KeyboardButton(btn_text))
            
            if len(free_keys) > 40:
                keyboard.add(types.KeyboardButton('📄 Показать еще...'))
            keyboard.add(types.KeyboardButton('🔍 Поиск по ключу'))
            keyboard.add(types.KeyboardButton('🔙 Назад'))
            
            msg = bot.send_message(message.chat.id, "Выберите ключ для удаления:", reply_markup=keyboard)
            bot.register_next_step_handler(msg, process_key_deletion)
            return
        
        # Ищем выбранный ключ
        btn_text = message.text.replace('🗑 ', '')
        
        user_data = get_user_data(message.from_user.id)
        free_keys = user_data.get('free_keys_list', [])
        
        selected_key = None
        for key_data in free_keys:
            key_name = key_data.get('name', 'Без имени')
            server = key_data.get('server', 'Unknown')
            if f"{key_name} ({server})" == btn_text:
                selected_key = key_data
                break
        
        if not selected_key:
            bot.send_message(message.chat.id, "❌ Ключ не найден.", reply_markup=admin_menu_keyboard())
            return
        
        # Подтверждение удаления
        key_preview = selected_key.get('key', '')[:50] + '...'
        
        confirm_text = (
            f"⚠️ <b>Подтвердите удаление ключа</b>\n\n"
            f"📝 <b>Название:</b> {selected_key.get('name', 'Без имени')}\n"
            f"🖥 <b>Сервер:</b> {selected_key.get('server', 'Unknown')}\n"
            f"🔑 <b>Ключ:</b> <code>{key_preview}</code>\n\n"
            f"<b>Это действие нельзя отменить!</b>"
        )
        
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        keyboard.add(types.KeyboardButton('✅ Да, удалить'))
        keyboard.add(types.KeyboardButton('❌ Нет, отменить'))
        
        user_data['selected_key_to_delete'] = selected_key
        save_user_data(message.from_user.id, user_data)
        
        msg = bot.send_message(message.chat.id, confirm_text, parse_mode='HTML', reply_markup=keyboard)
        bot.register_next_step_handler(msg, confirm_key_deletion)
        
    except Exception as e:
        logger.error(f"Ошибка в process_key_deletion: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())

def process_key_search_for_delete(message):
    """Поиск ключа по тексту для удаления"""
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        search_text = message.text.strip().lower()
        
        # Загружаем ключи
        with open('keys_storage.json', 'r', encoding='utf-8') as f:
            keys_storage = json.load(f)
            all_keys = keys_storage.get('keys', [])
        
        # Собираем используемые ключи
        used_keys_set = set()
        for server in servers_db.values():
            for key_data in server.get('used_keys', {}).values():
                if key_data and key_data.get('key'):
                    used_keys_set.add(key_data.get('key'))
        
        # Ищем ключи
        found_keys = []
        for i, key_data in enumerate(all_keys):
            if key_data.get('key') not in used_keys_set:
                key_string = key_data.get('key', '').lower()
                key_name = key_data.get('name', '').lower()
                server = key_data.get('server', '').lower()
                
                if search_text in key_string or search_text in key_name or search_text in server:
                    key_data['_index'] = i
                    found_keys.append(key_data)
        
        if not found_keys:
            bot.send_message(message.chat.id, f"❌ Ключей по запросу '{message.text}' не найдено.", 
                           reply_markup=admin_menu_keyboard())
            return
        
        # Создаем клавиатуру с найденными ключами
        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        
        for key_data in found_keys[:15]:
            key_name = key_data.get('name', 'Без имени')
            server = key_data.get('server', 'Unknown')
            btn_text = f"🗑 {key_name} ({server})"
            keyboard.add(types.KeyboardButton(btn_text))
        
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        user_data = get_user_data(message.from_user.id)
        user_data['free_keys_list'] = found_keys
        save_user_data(message.from_user.id, user_data)
        
        text = f"🔍 <b>Найдено ключей:</b> {len(found_keys)}\n\nВыберите ключ для удаления:"
        
        msg = bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_key_deletion)
        
    except Exception as e:
        logger.error(f"Ошибка в process_key_search_for_delete: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())

def confirm_key_deletion(message):
    """Подтверждение удаления ключа"""
    try:
        if message.text == '❌ Нет, отменить':
            bot.send_message(message.chat.id, "✅ Удаление отменено.", reply_markup=admin_menu_keyboard())
            return
        
        if message.text != '✅ Да, удалить':
            bot.send_message(message.chat.id, "❌ Неверный ответ. Удаление отменено.", reply_markup=admin_menu_keyboard())
            return
        
        user_data = get_user_data(message.from_user.id)
        selected_key = user_data.get('selected_key_to_delete')
        
        if not selected_key:
            bot.send_message(message.chat.id, "❌ Ключ не выбран.", reply_markup=admin_menu_keyboard())
            return
        
        key_index = selected_key.get('_index')
        key_string = selected_key.get('key', '')
        
        # Удаляем из хранилища
        with open('keys_storage.json', 'r', encoding='utf-8') as f:
            keys_storage = json.load(f)
        
        if key_index is not None and key_index < len(keys_storage['keys']):
            deleted_key = keys_storage['keys'].pop(key_index)
            
            with open('keys_storage.json', 'w', encoding='utf-8') as f:
                json.dump(keys_storage, f, ensure_ascii=False, indent=2)
            
            # Удаляем из available_keys сервера
            server_name = deleted_key.get('server', '')
            for server_key, server_data in servers_db.items():
                if server_data['name'] == server_name:
                    if 'available_keys' in server_data:
                        server_data['available_keys'] = [
                            k for k in server_data['available_keys'] 
                            if k.get('key') != key_string
                        ]
                    break
            
            save_data_to_file()
            
            # Очищаем временные данные
            if 'free_keys_list' in user_data:
                del user_data['free_keys_list']
            if 'selected_key_to_delete' in user_data:
                del user_data['selected_key_to_delete']
            save_user_data(message.from_user.id, user_data)
            
            bot.send_message(
                message.chat.id,
                f"✅ <b>Ключ успешно удален!</b>\n\n"
                f"📝 Название: {deleted_key.get('name', 'Без имени')}\n"
                f"🖥 Сервер: {deleted_key.get('server', 'Unknown')}",
                parse_mode='HTML',
                reply_markup=admin_menu_keyboard()
            )
        else:
            bot.send_message(message.chat.id, "❌ Ошибка: ключ не найден в хранилище.", 
                           reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка в confirm_key_deletion: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())

# Добавляем обработчик для покупки нового ключа
@bot.message_handler(func=lambda message: message.text == '➕ Купить новый ключ')
def buy_new_key(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        bot.send_message(user_id, "Выберите сервер для нового ключа:", reply_markup=servers_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в buy_new_key: {e}")

@bot.message_handler(func=lambda message: message.text.startswith('🔄 Продлить '))
def handle_extend_subscription(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        # Извлекаем сервер из кнопки
        server_name = message.text.replace('🔄 Продлить ', '')
        
        # Проверяем наличие ключей (для продления ключи не нужны, но проверяем что сервер существует)
        server_key = None
        for key, server_data in servers_db.items():
            if server_data['name'] == server_name:
                server_key = key
                break
        
        if not server_key:
            bot.send_message(user_id, "❌ Сервер не найден!", reply_markup=main_menu_keyboard())
            return
        
        # Сохраняем выбранный сервер для продления
        user_data = get_user_data(user_id)
        user_data['extend_server'] = server_name
        save_user_data(user_id, user_data)
        
        # Предлагаем выбрать срок продления
        bot.send_message(
            user_id,
            f"🔄 <b>Продление подписки</b>\n\n"
            f"Вы выбрали продление для сервера: {server_name}\n\n"
            f"Выберите срок продления:",
            parse_mode='HTML',
            reply_markup=duration_menu_keyboard()
        )
    except Exception as e:
        logger.error(f"Ошибка в handle_extend_subscription: {e}")
        bot.send_message(user_id, "Произошла ошибка при обработке запроса.")
        
@bot.message_handler(func=lambda message: message.text == '📤 Добавить конфиги массово' and is_admin(message.from_user.id))
def bulk_upload_configs(message):
    """Массовая загрузка конфигурационных файлов"""
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        # Создаем клавиатуру с выбором сервера
        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        for server in servers_db.values():
            keyboard.add(types.KeyboardButton(f"📁 Массовая загрузка {server['name']}"))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "🖥 <b>Массовая загрузка конфигов</b>\n\n"
            "Выберите сервер для загрузки конфигурационных файлов.\n\n"
            "💡 <b>Как использовать:</b>\n"
            "1. Выберите сервер\n"  
            "2. Отправьте один или несколько .conf файлов\n"
            "3. Файлы автоматически добавятся на выбранный сервер",
            parse_mode='HTML',
            reply_markup=keyboard
        )
        bot.register_next_step_handler(msg, process_bulk_server_selection)
    except Exception as e:
        logger.error(f"Ошибка в bulk_upload_configs: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при запуске массовой загрузки")

def process_bulk_server_selection(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        # Извлекаем название сервера из кнопки
        server_name = message.text.replace('📁 Массовая загрузка ', '')
        
        # Проверяем существование сервера
        server_key = next((k for k, v in servers_db.items() if v['name'] == server_name), None)
        if not server_key:
            bot.send_message(message.chat.id, "❌ Сервер не найден", reply_markup=admin_menu_keyboard())
            return
        
        # Сохраняем выбранный сервер во временных данных
        user_data = get_user_data(message.from_user.id)
        user_data['bulk_upload_server'] = server_name
        user_data['bulk_upload_files'] = []  # Список для отслеживания загруженных файлов
        save_user_data(message.from_user.id, user_data)
        
        # Показываем текущую статистику сервера
        server_data = servers_db[server_key]
        stats_text = (
            f"📤 <b>Массовая загрузка для {server_name}</b>\n\n"
            f"📊 <b>Текущая статистика сервера:</b>\n"
            f"🆓 Доступно конфигов: {len(server_data['available_configs'])}\n"
            f"👥 Используется: {len(server_data['used_configs'])}\n\n"
            f"📨 <b>Теперь отправьте .conf файлы</b>\n\n"
            f"💡 <b>Совет:</b> Можно отправить несколько файлов за раз\n"
            f"📝 После загрузки файлов нажмите /done для завершения"
        )
        
        msg = bot.send_message(
            message.chat.id,
            stats_text,
            parse_mode='HTML',
            reply_markup=types.ReplyKeyboardRemove()
        )
        
        # Регистрируем обработчик для документов
        bot.register_next_step_handler(msg, process_bulk_configs_upload)
        
    except Exception as e:
        logger.error(f"Ошибка в process_bulk_server_selection: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка выбора сервера", reply_markup=admin_menu_keyboard())

def process_bulk_configs_upload(message):
    try:
        user_id = message.from_user.id
        user_data = get_user_data(user_id)
        server_name = user_data.get('bulk_upload_server')
        
        if not server_name:
            bot.send_message(message.chat.id, "❌ Ошибка: сервер не выбран", reply_markup=admin_menu_keyboard())
            return
        
        # Проверяем команду завершения
        if message.text and message.text == '/done':
            send_bulk_upload_summary(message.chat.id, server_name, user_data.get('bulk_upload_files', []))
            # Очищаем временные данные
            if 'bulk_upload_server' in user_data:
                del user_data['bulk_upload_server']
            if 'bulk_upload_files' in user_data:
                del user_data['bulk_upload_files']
            save_user_data(user_id, user_data)
            return
        
        if message.content_type != 'document':
            bot.send_message(
                message.chat.id, 
                "❌ Пожалуйста, отправьте .conf файлы или введите /done для завершения",
                reply_markup=admin_menu_keyboard()
            )
            return
        
        # Обрабатываем файл
        success = process_single_config_file(message.document, server_name, user_id)
        
        if success:
            # Добавляем файл в список загруженных
            if 'bulk_upload_files' not in user_data:
                user_data['bulk_upload_files'] = []
            user_data['bulk_upload_files'].append(message.document.file_name)
            save_user_data(user_id, user_data)
            
            # Отправляем подтверждение
            bot.send_message(
                message.chat.id,
                f"✅ Файл '{message.document.file_name}' успешно добавлен!\n"
                f"Отправьте следующий файл или введите /done для завершения",
                reply_markup=types.ReplyKeyboardRemove()
            )
            
            # Регистрируем следующий шаг
            bot.register_next_step_handler(message, process_bulk_configs_upload)
        else:
            bot.send_message(
                message.chat.id,
                f"❌ Ошибка при добавлении файла '{message.document.file_name}'\n"
                f"Убедитесь что это .conf файл и попробуйте снова",
                reply_markup=types.ReplyKeyboardRemove()
            )
            bot.register_next_step_handler(message, process_bulk_configs_upload)
        
    except Exception as e:
        logger.error(f"Ошибка в process_bulk_configs_upload: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при загрузке файлов", reply_markup=admin_menu_keyboard())
        
def send_bulk_upload_summary(chat_id, server_name, uploaded_files):
    """Отправляет сводку о массовой загрузке"""
    try:
        server_key = next(k for k, v in servers_db.items() if v['name'] == server_name)
        server_data = servers_db[server_key]
        
        text = f"📊 <b>Сводка массовой загрузки для {server_name}</b>\n\n"
        text += f"✅ Успешно загружено: {len(uploaded_files)} файлов\n\n"
        
        if uploaded_files:
            text += "<b>Загруженные файлы:</b>\n"
            for i, file in enumerate(uploaded_files[:15], 1):
                text += f"{i}. {file}\n"
            
            if len(uploaded_files) > 15:
                text += f"... и еще {len(uploaded_files) - 15} файлов\n"
        
        text += f"\n📈 <b>Обновленная статистика сервера:</b>\n"
        text += f"🆓 Доступно конфигов: {len(server_data['available_configs'])}\n"
        text += f"👥 Используется: {len(server_data['used_configs'])}\n"
        text += f"📁 Всего конфигов: {len(server_data['available_configs']) + len(server_data['used_configs'])}\n"
        
        # Добавляем предупреждение если файлы не соответствуют серверу
        wrong_files = []
        for file in uploaded_files:
            file_lower = file.lower()
            if 'grm' in file_lower and 'Netherlands' in server_name:
                wrong_files.append(file)
            elif 'ndr' in file_lower and 'Germany' in server_name:
                wrong_files.append(file)
        
        if wrong_files:
            text += f"\n⚠️ <b>Внимание! Файлы не соответствуют серверу:</b>\n"
            for f in wrong_files[:5]:
                text += f"• {f}\n"
            text += f"\n💡 Рекомендуется загружать конфиги на соответствующий сервер."
        
        bot.send_message(chat_id, text, parse_mode='HTML', reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка отправки сводки: {e}")
        bot.send_message(chat_id, f"✅ Загружено {len(uploaded_files)} файлов", reply_markup=admin_menu_keyboard())
        
@bot.message_handler(func=lambda message: message.text == '🔄 Перенести конфиги' and is_admin(message.from_user.id))
def move_configs_between_servers(message):
    """Переносит конфиги с Германии на Нидерланды и наоборот"""
    try:
        # Получаем все конфиги
        all_configs = [f for f in os.listdir() if f.endswith('.conf')]
        
        # Группируем по серверам
        germany_configs = []
        netherlands_configs = []
        
        for config in all_configs:
            config_lower = config.lower()
            if 'grm' in config_lower:
                germany_configs.append(config)
            elif 'ndr' in config_lower:
                netherlands_configs.append(config)
        
        text = (
            f"🔄 <b>Перенос конфигов между серверами</b>\n\n"
            f"🇩🇪 Германия: {len(germany_configs)} конфигов\n"
            f"🇳🇱 Нидерланды: {len(netherlands_configs)} конфигов\n\n"
            f"Выберите действие:"
        )
        
        markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        markup.add(
            types.KeyboardButton("🇩🇪 → 🇳🇱 Перенести Германия в Нидерланды"),
            types.KeyboardButton("🇳🇱 → 🇩🇪 Перенести Нидерланды в Германию")
        )
        markup.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=markup)
        bot.register_next_step_handler(msg, process_move_configs)
        
    except Exception as e:
        logger.error(f"Ошибка в move_configs_between_servers: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}")

def process_move_configs(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        if 'Германия в Нидерланды' in message.text:
            # Переносим конфиги с Германии на Нидерланды
            for server_key, server_data in servers_db.items():
                if 'Germany' in server_data['name']:
                    germany_available = server_data['available_configs'][:]
                    # Добавляем конфиги Германии в Нидерланды
                    for nether_key, nether_data in servers_db.items():
                        if 'Netherlands' in nether_data['name']:
                            nether_data['available_configs'].extend(germany_available)
                            server_data['available_configs'] = []
                            break
                    break
            
            save_data_to_file()
            bot.send_message(message.chat.id, "✅ Конфиги перенесены с Германии на Нидерланды!", reply_markup=admin_menu_keyboard())
            
        elif 'Нидерланды в Германию' in message.text:
            # Переносим конфиги с Нидерландов на Германию
            for server_key, server_data in servers_db.items():
                if 'Netherlands' in server_data['name']:
                    netherlands_available = server_data['available_configs'][:]
                    # Добавляем конфиги Нидерландов в Германию
                    for germany_key, germany_data in servers_db.items():
                        if 'Germany' in germany_data['name']:
                            germany_data['available_configs'].extend(netherlands_available)
                            server_data['available_configs'] = []
                            break
                    break
            
            save_data_to_file()
            bot.send_message(message.chat.id, "✅ Конфиги перенесены с Нидерландов на Германию!", reply_markup=admin_menu_keyboard())
            
    except Exception as e:
        logger.error(f"Ошибка в process_move_configs: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}")
        
@bot.message_handler(commands=['done'])
def handle_done_command(message):
    """Обрабатывает команду завершения массовой загрузки"""
    if not is_admin(message.from_user.id):
        return
    
    user_data = get_user_data(message.from_user.id)
    server_name = user_data.get('bulk_upload_server')
    
    if server_name:
        uploaded_files = user_data.get('bulk_upload_files', [])
        send_bulk_upload_summary(message.chat.id, server_name, uploaded_files)
        
        # Очищаем временные данные
        if 'bulk_upload_server' in user_data:
            del user_data['bulk_upload_server']
        if 'bulk_upload_files' in user_data:
            del user_data['bulk_upload_files']
        save_user_data(message.from_user.id, user_data)
    else:
        bot.send_message(message.chat.id, "❌ Нет активной массовой загрузки", reply_markup=admin_menu_keyboard())

def process_single_config_file(document, server_name, user_id):
    """Обрабатывает один конфигурационный файл и автоматически определяет сервер"""
    try:
        # Проверяем что файл имеет правильное расширение
        file_name = document.file_name
        if not (file_name.endswith('.conf') or file_name.endswith('.json')):
            return False
        
        # Определяем реальный сервер по имени файла
        config_lower = file_name.lower()
        actual_server = None
        
        if any(x in config_lower for x in ['grm', 'germany', 'de_', '_de', 'ger']):
            actual_server = '🇩🇪 Germany'
        elif any(x in config_lower for x in ['ndr', 'netherlands', 'nl_', '_nl', 'ned', 'holland']):
            actual_server = '🇳🇱 Netherlands'
        
        # Если файл не соответствует выбранному серверу - предупреждаем
        if actual_server and actual_server != server_name:
            logger.warning(f"Файл {file_name} предназначен для {actual_server}, но добавляется на {server_name}")
            # Можно вернуть False или продолжить - решать вам
        
        file_info = bot.get_file(document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        # Сохраняем файл с оригинальным именем
        config_filename = file_name
        
        with open(config_filename, 'wb') as new_file:
            new_file.write(downloaded_file)
        
        # Добавляем конфиг в доступные для сервера
        server_key = next(k for k, v in servers_db.items() if v['name'] == server_name)
        if config_filename not in servers_db[server_key]['available_configs']:
            servers_db[server_key]['available_configs'].append(config_filename)
        
        save_data_to_file()
        
        logger.info(f"Админ {user_id} добавил конфиг {config_filename} для сервера {server_name}")
        return True
        
    except Exception as e:
        logger.error(f"Ошибка обработки файла {document.file_name}: {e}")
        return False

# Добавляем обработчик для группы медиа (несколько файлов)
@bot.message_handler(content_types=['document'], func=lambda message: hasattr(message, 'media_group_id'))
def handle_media_group(message):
    """Обрабатывает группу файлов (несколько файлов отправленных вместе)"""
    try:
        # Эта функция будет вызываться для каждого файла в группе
        user_id = message.from_user.id
        if not is_admin(user_id):
            return
        
        user_data = get_user_data(user_id)
        server_name = user_data.get('bulk_upload_server')
        
        if server_name and message.document and message.document.file_name.endswith('.conf'):
            success = process_single_config_file(message.document, server_name, user_id)
            if success:
                logger.info(f"Успешно добавлен конфиг из группы: {message.document.file_name}")
    
    except Exception as e:
        logger.error(f"Ошибка в handle_media_group: {e}")
        
def send_bulk_upload_summary(chat_id, server_name, uploaded_files, total_files):
    """Отправляет сводку о массовой загрузке"""
    try:
        text = f"📊 <b>Сводка массовой загрузки для {server_name}</b>\n\n"
        text += f"✅ Успешно загружено: {len(uploaded_files)} файлов\n"
        text += f"📁 Всего обработано: {total_files} файлов\n\n"
        
        if uploaded_files:
            text += "<b>Загруженные файлы:</b>\n"
            for file in uploaded_files[:10]:  # Показываем первые 10 файлов
                text += f"• {file}\n"
            
            if len(uploaded_files) > 10:
                text += f"... и еще {len(uploaded_files) - 10} файлов\n"
        
        # Обновляем статистику сервера
        server_key = next(k for k, v in servers_db.items() if v['name'] == server_name)
        available_count = len(servers_db[server_key]['available_configs'])
        
        text += f"\n📈 <b>Текущая статистика сервера:</b>\n"
        text += f"🆓 Доступно конфигов: {available_count}\n"
        text += f"👥 Используется: {len(servers_db[server_key]['used_configs'])}\n"
        
        bot.send_message(chat_id, text, parse_mode='HTML', reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка отправки сводки: {e}")

@bot.message_handler(func=lambda message: message.text == '🇩🇪 Germany')
def select_server(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        server_name = '🇩🇪 Germany'
        
        user_data = get_user_data(user_id)
        user_data['selected_server'] = server_name
        save_user_data(user_id, user_data)
        
        bot.send_message(user_id, f"Выбрали сервер: {server_name}\nВыберите срок:", reply_markup=duration_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в select_server: {e}")

@bot.message_handler(func=lambda message: any(duration in message.text for duration in ['1 месяц', '3 месяца', '6 месяцев']))
def select_duration(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        duration = message.text.split(' - ')[0]
        price = message.text.split(' - ')[1]
        
        user_data = get_user_data(user_id)
        
        # Проверяем, есть ли выбранный сервер
        if 'selected_server' not in user_data and 'extend_server' not in user_data:
            bot.send_message(user_id, "Пожалуйста, сначала выберите сервер.")
            return
        
        # Определяем, продление это или новая покупка
        is_extension = 'extend_server' in user_data
        
        if is_extension:
            server_name = user_data['extend_server']
            # Удаляем флаг продления, чтобы не мешал
            del user_data['extend_server']
        else:
            server_name = user_data['selected_server']
        
        user_data['selected_duration'] = duration
        user_data['selected_price'] = price
        user_data['selected_server'] = server_name
        user_data['is_extension'] = is_extension  # Сохраняем флаг продления
        save_user_data(user_id, user_data)
        
        # Создаем платеж в ЮKassa
        create_yookassa_payment_for_user(user_id, duration, price, server_name, is_extension)
        
    except Exception as e:
        logger.error(f"Ошибка в select_duration: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при создании платежа")
        
        
        
def create_yookassa_payment_for_user(user_id, duration, price, server_name, is_extension=False):
    """Создает платеж в ЮKassa"""
    try:
        user_data = get_user_data(user_id)
        
        # Генерируем ID платежа
        payment_id = generate_payment_id()
        
        # Получаем сумму (убираем символ валюты)
        amount = price.replace('₽', '').strip()
        
        # Сохраняем платеж в базе
        payment_data = {
            'user_id': user_id,
            'username': user_data.get('username', ''),
            'server': server_name,
            'duration': duration,
            'amount': price,
            'status': 'pending',
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'yookassa_payment_id': None,
            'is_extension': is_extension
        }
        save_payment(payment_id, payment_data)
        
        # Создаем платеж в ЮKassa
        payment_type = "продление" if is_extension else "покупка"
        
        # Создаем платеж (для тестового магазина доступны карты и ЮMoney)
        yookassa_result = create_yookassa_payment(
            amount=amount,
            description=f"{payment_type} VPN подписки {duration} на сервере {server_name}",
            payment_id=payment_id,
            user_id=user_id,
            username=user_data.get('username', '')
        )
        
        if yookassa_result and yookassa_result.get('confirmation_url'):
            # Сохраняем ID платежа ЮKassa
            payment_data['yookassa_payment_id'] = yookassa_result['payment_id']
            save_payment(payment_id, payment_data)
            
            # Создаем инлайн-клавиатуру с кнопкой оплаты
            markup = types.InlineKeyboardMarkup(row_width=1)
            markup.add(
                types.InlineKeyboardButton(
                    "💳 Перейти к оплате", 
                    url=yookassa_result['confirmation_url']
                )
            )
            markup.add(
                types.InlineKeyboardButton(
                    "✅ Проверить оплату", 
                    callback_data=f"check_payment_{payment_id}"
                )
            )
            
            # Отправляем сообщение с информацией
            order_text = (
                f"🛒 <b>Оплата подписки</b>\n\n"
                f"🖥 Сервер: {server_name}\n"
                f"⏳ Срок: {duration}\n"
                f"💰 Сумма: {price}\n"
                f"🆔 ID заказа: {payment_id}\n\n"
                f"<b>Доступные способы оплаты:</b>\n"
                f"• 💳 Банковские карты (Visa, Mastercard, МИР)\n"
                f"• 🟡 ЮMoney\n\n"
                f"Нажмите кнопку для оплаты:"
            )
            
            bot.send_message(
                user_id,
                order_text,
                parse_mode='HTML',
                reply_markup=markup
            )
            
        else:
            bot.send_message(
                user_id,
                "❌ Ошибка создания платежа. Пожалуйста, попробуйте позже.",
                reply_markup=main_menu_keyboard()
            )
            
    except Exception as e:
        logger.error(f"Ошибка в create_yookassa_payment_for_user: {e}")
        bot.send_message(user_id, "❌ Произошла ошибка при создании платежа")
        
def show_payment_methods_menu(user_id, payment_id, amount, description):
    """Показывает меню выбора способа оплаты"""
    try:
        markup = types.InlineKeyboardMarkup(row_width=2)
        markup.add(
            types.InlineKeyboardButton("💳 Банковская карта", callback_data=f"pay_method_bank_card_{payment_id}"),
            types.InlineKeyboardButton("🟡 ЮMoney", callback_data=f"pay_method_yoo_money_{payment_id}")
        )
        markup.add(
            types.InlineKeyboardButton("🏦 Сбербанк Онлайн", callback_data=f"pay_method_sberbank_{payment_id}"),
            types.InlineKeyboardButton("📱 QIWI Wallet", callback_data=f"pay_method_qiwi_{payment_id}")
        )
        markup.add(
            types.InlineKeyboardButton("💻 WebMoney", callback_data=f"pay_method_webmoney_{payment_id}"),
            types.InlineKeyboardButton("🔵 Альфа-Клик", callback_data=f"pay_method_alfabank_{payment_id}")
        )
        markup.add(
            types.InlineKeyboardButton("🟢 Тинькофф", callback_data=f"pay_method_tinkoff_bank_{payment_id}"),
            types.InlineKeyboardButton("📞 Баланс телефона", callback_data=f"pay_method_mobile_balance_{payment_id}")
        )
        
        bot.send_message(
            user_id,
            f"💳 <b>Выберите способ оплаты</b>\n\n"
            f"Сумма: {amount} ₽\n"
            f"{description}\n\n"
            f"Выберите удобный способ оплаты:",
            parse_mode='HTML',
            reply_markup=markup
        )
    except Exception as e:
        logger.error(f"Ошибка в show_payment_methods_menu: {e}")
        
@bot.callback_query_handler(func=lambda call: call.data.startswith('pay_method_'))
def handle_payment_method_selection(call):
    """Обрабатывает выбор способа оплаты"""
    try:
        # Парсим callback data
        parts = call.data.split('_')
        method = parts[2]  # bank_card, yoo_money, etc.
        payment_id = parts[3]
        
        payment = get_payment(payment_id)
        if not payment:
            bot.answer_callback_query(call.id, "❌ Платеж не найден!")
            return
        
        user_id = call.from_user.id
        amount = payment['amount'].replace('₽', '').strip()
        description = f"Оплата подписки #{payment_id}"
        
        # Создаем платеж с выбранным способом
        yookassa_result = create_yookassa_payment(
            amount=amount,
            description=description,
            payment_id=payment_id,
            user_id=user_id,
            username=payment.get('username', ''),
            payment_method=method
        )
        
        if yookassa_result and yookassa_result.get('confirmation_url'):
            # Сохраняем ID платежа ЮKassa
            payment['yookassa_payment_id'] = yookassa_result['payment_id']
            save_payment(payment_id, payment)
            
            # Создаем кнопку оплаты
            markup = types.InlineKeyboardMarkup()
            markup.add(
                types.InlineKeyboardButton(
                    "💳 Перейти к оплате", 
                    url=yookassa_result['confirmation_url']
                )
            )
            markup.add(
                types.InlineKeyboardButton(
                    "✅ Проверить оплату", 
                    callback_data=f"check_payment_{payment_id}"
                )
            )
            
            # Редактируем сообщение
            bot.edit_message_text(
                f"🛒 <b>Оплата подписки</b>\n\n"
                f"💰 Сумма: {payment['amount']}\n"
                f"💳 Способ оплаты: {get_method_name(method)}\n"
                f"🆔 ID заказа: {payment_id}\n\n"
                f"Нажмите кнопку для оплаты:",
                call.message.chat.id,
                call.message.message_id,
                parse_mode='HTML',
                reply_markup=markup
            )
        else:
            bot.answer_callback_query(call.id, "❌ Ошибка создания платежа")
            
    except Exception as e:
        logger.error(f"Ошибка в handle_payment_method_selection: {e}")
        bot.answer_callback_query(call.id, "❌ Ошибка")

def get_method_name(method_code):
    """Возвращает название способа оплаты по коду"""
    methods = {
        "bank_card": "Банковская карта",
        "yoo_money": "ЮMoney",
        "sberbank": "Сбербанк Онлайн",
        "qiwi": "QIWI Wallet",
        "webmoney": "WebMoney",
        "alfabank": "Альфа-Клик",
        "tinkoff_bank": "Тинькофф",
        "mobile_balance": "Баланс телефона"
    }
    return methods.get(method_code, "Банковская карта")

@bot.message_handler(func=lambda message: any(method['bank'] in message.text for method in payment_methods.values()))
def select_payment_method(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        bank_name = message.text
        user_data = get_user_data(user_id)
        server_name = user_data['selected_server']
        duration = user_data['selected_duration']
        price = user_data['selected_price']
        
        # Генерируем ID платежа
        payment_id = generate_payment_id()
        
        # Сохраняем платеж в базе
        payment_data = {
            'user_id': user_id,
            'username': message.from_user.username,
            'server': server_name,
            'duration': duration,
            'amount': price,
            'bank': bank_name,
            'status': 'pending',
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'yookassa_payment_id': None
        }
        save_payment(payment_id, payment_data)
        
        # Создаем платеж в ЮKassa
        amount = price.replace('₽', '').strip()
        yookassa_result = create_yookassa_payment(
            amount=amount,
            description=f"VPN подписка {duration} на сервере {server_name}",
            payment_id=payment_id,
            user_id=user_id,
            username=message.from_user.username
        )
        
        if yookassa_result and yookassa_result.get('confirmation_url'):
            # Сохраняем ID платежа ЮKassa
            payment_data['yookassa_payment_id'] = yookassa_result['payment_id']
            save_payment(payment_id, payment_data)
            
            # Создаем инлайн-клавиатуру с кнопкой оплаты
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton(
                "💳 Перейти к оплате", 
                url=yookassa_result['confirmation_url']
            ))
            markup.add(types.InlineKeyboardButton(
                "✅ Проверить оплату", 
                callback_data=f"check_payment_{payment_id}"
            ))
            
            # Клавиатура для отмены
            reply_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            reply_markup.add(types.KeyboardButton('🔙 Назад'))
            
            bot.send_message(
                user_id,
                f"💳 <b>Счет на оплату</b>\n\n"
                f"🖥 Сервер: {server_name}\n"
                f"⏳ Срок: {duration}\n"
                f"💰 Сумма: {price}\n"
                f"🆔 ID платежа: {payment_id}\n\n"
                f"Нажмите на кнопку ниже, чтобы перейти к оплате.\n"
                f"После оплаты нажмите «Проверить оплату».",
                parse_mode='HTML',
                reply_markup=reply_markup
            )
            
            bot.send_message(
                user_id,
                "💳 <b>Оплата через ЮKassa</b>\n\n"
                "Нажмите кнопку для перехода к оплате:",
                parse_mode='HTML',
                reply_markup=markup
            )
        else:
            bot.send_message(
                user_id,
                "❌ Ошибка создания платежа. Пожалуйста, попробуйте позже.",
                reply_markup=main_menu_keyboard()
            )
            
    except Exception as e:
        logger.error(f"Ошибка в select_payment_method: {e}")
        bot.send_message(user_id, "❌ Ошибка при создании платежа")
        
@bot.callback_query_handler(func=lambda call: call.data.startswith('check_payment_'))
def check_payment_callback(call):
    """Проверяет статус платежа по callback"""
    try:
        payment_id = call.data.replace('check_payment_', '')
        
        payment = get_payment(payment_id)
        if not payment:
            bot.answer_callback_query(call.id, "❌ Платеж не найден!")
            return
        
        # Если платеж уже подтвержден
        if payment.get('status') == 'approved':
            bot.answer_callback_query(call.id, "✅ Платеж уже подтвержден!")
            # Удаляем сообщение с кнопками
            try:
                bot.delete_message(call.message.chat.id, call.message.message_id)
            except:
                pass
            return
        
        # Проверяем статус в ЮKassa
        if payment.get('yookassa_payment_id'):
            yookassa_status = check_payment_status(payment['yookassa_payment_id'])
            
            if yookassa_status and yookassa_status.get('paid'):
                # Платеж оплачен - подтверждаем
                bot.answer_callback_query(call.id, "✅ Платеж подтвержден!")
                
                # Обновляем статус платежа
                payment['status'] = 'approved'
                payment['approved_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                payment['approved_by'] = 'system'
                save_payment(payment_id, payment)
                
                # Удаляем сообщение с кнопками
                try:
                    bot.delete_message(call.message.chat.id, call.message.message_id)
                except:
                    pass
                
                # Выдаем конфиг пользователю
                success = issue_vpn_config(payment_id, call.message.chat.id)
                
                if not success:
                    bot.send_message(
                        call.message.chat.id,
                        "❌ Ошибка при выдаче VPN ключа. Обратитесь к администратору.",
                        reply_markup=main_menu_keyboard()
                    )
            else:
                bot.answer_callback_query(
                    call.id, 
                    "❌ Платеж еще не оплачен. Оплатите счет и нажмите сюда снова.",
                    show_alert=True
                )
        else:
            bot.answer_callback_query(call.id, "❌ Не удалось проверить статус платежа")
            
    except Exception as e:
        logger.error(f"Ошибка в check_payment_callback: {e}")
        bot.answer_callback_query(call.id, "❌ Ошибка проверки платежа")
        
def issue_vpn_config(payment_id, chat_id):
    """Выдает VPN ключ после успешной оплаты"""
    try:
        payment = get_payment(payment_id)
        if not payment:
            logger.error(f"Платеж {payment_id} не найден")
            return False
        
        user_id = payment['user_id']
        server_name = payment['server']
        duration = payment['duration']
        is_extension = payment.get('is_extension', False)
        
        logger.info(f"Выдача ключа: user_id={user_id}, server={server_name}, duration={duration}, is_extension={is_extension}")
        
        user_data = get_user_data(user_id)
        
        key_data = None
        old_subscription = None
        
        # Если это продление - находим старую подписку
        if is_extension:
            for sub in user_data.get('subscriptions', []):
                if sub['server'] == server_name:
                    old_subscription = sub
                    key_data = sub.get('key_data')
                    break
            
            if not key_data:
                is_extension = False
        
        if not is_extension:
            # Новая подписка - получаем новый ключ
            key_data = get_random_config(server_name, user_id)
            if not key_data:
                bot.send_message(chat_id, f"❌ Нет доступных ключей для сервера {server_name}!")
                return False
        
        # Рассчитываем новую дату окончания
        days = SUBSCRIPTION_PLANS.get(duration, {}).get('days', 30)
        
        if is_extension and old_subscription:
            current_expiry = datetime.strptime(old_subscription['expiry_date'], "%Y-%m-%d %H:%M:%S")
            if current_expiry > datetime.now():
                new_expiry = current_expiry + timedelta(days=days)
            else:
                new_expiry = datetime.now() + timedelta(days=days)
            expiry_date = new_expiry.strftime("%Y-%m-%d %H:%M:%S")
            
            old_subscription['expiry_date'] = expiry_date
            old_subscription['payment_id'] = payment_id
            old_subscription['duration'] = duration
            
            if 'last_warnings' in old_subscription:
                del old_subscription['last_warnings']
            if 'expiry_notification_sent' in old_subscription:
                del old_subscription['expiry_notification_sent']
        else:
            expiry_date = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
            
            if 'subscriptions' not in user_data:
                user_data['subscriptions'] = []
            
            user_data['subscriptions'].append({
                'server': server_name,
                'key_data': key_data,
                'purchase_date': payment.get('approved_at', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                'expiry_date': expiry_date,
                'payment_id': payment_id,
                'duration': duration,
                'type': 'paid'
            })
        
        user_data['username'] = payment.get('username')
        save_user_data(user_id, user_data)
        
        # Отправляем результат
        if not is_extension:
            send_vpn_config_to_user(user_id, key_data, server_name, expiry_date, duration)
        else:
            send_extension_notification(user_id, server_name, expiry_date, duration)
        
        # Уведомляем админа
        payment_type = "ПРОДЛЕНИЕ" if is_extension else "НОВАЯ ПОДПИСКА"
        key_preview = key_data.get('key', '')[:30] + '...' if key_data else 'существующий ключ'
        
        bot.send_message(
            ADMIN_ID,
            f"✅ <b>{payment_type} через ЮKassa</b>\n\n"
            f"👤 Пользователь: @{payment.get('username')} (ID: {user_id})\n"
            f"🖥 Сервер: {server_name}\n"
            f"⏳ Срок: {duration}\n"
            f"💰 Сумма: {payment['amount']}\n"
            f"📅 Действует до: {expiry_date}\n"
            f"🔑 Ключ: <code>{key_preview}</code>",
            parse_mode='HTML'
        )
        
        # ==================== НАЧИСЛЕНИЕ РЕФЕРАЛЬНОГО БОНУСА 25% ====================
        try:
            referred_user_id = str(user_id)
            
            # Извлекаем сумму платежа
            amount_str = payment.get('amount', '0₽')
            import re
            digits = re.findall(r'\d+', str(amount_str))
            amount_num = int(digits[0]) if digits else 0
            
            logger.info(f"🔗 РЕФЕРАЛ: Проверка бонуса. user_id={referred_user_id}, amount={amount_num}₽")
            
            if amount_num > 0:
                # Проверяем структуры referral_db
                if 'relations' not in referral_db:
                    referral_db['relations'] = {}
                
                if 'allowed_users' not in referral_db:
                    referral_db['allowed_users'] = {}
                
                # Проверяем, есть ли реферер у этого пользователя
                if referred_user_id in referral_db['relations']:
                    referrer_id = referral_db['relations'][referred_user_id]
                    logger.info(f"🔗 РЕФЕРАЛ: Найден реферер {referrer_id} для пользователя {referred_user_id}")
                    
                    # Проверяем, что реферер в списке разрешенных
                    if referrer_id in referral_db['allowed_users']:
                        # Рассчитываем бонус 25%
                        bonus = int(amount_num * 0.25)
                        
                        if bonus > 0:
                            # Начисляем бонус
                            if 'balance' not in referral_db['allowed_users'][referrer_id]:
                                referral_db['allowed_users'][referrer_id]['balance'] = 0
                            
                            old_balance = referral_db['allowed_users'][referrer_id]['balance']
                            referral_db['allowed_users'][referrer_id]['balance'] += bonus
                            new_balance = referral_db['allowed_users'][referrer_id]['balance']
                            
                            # Сохраняем изменения
                            save_referral_data()
                            logger.info(f"🔗 РЕФЕРАЛ: Баланс реферера {referrer_id}: {old_balance}₽ -> {new_balance}₽ (+{bonus}₽)")
                            
                            # Получаем информацию о реферале для красивого сообщения
                            referred_username = payment.get('username', 'Пользователь')
                            
                            # Красивое уведомление рефереру о начислении бонуса
                            notification_text = (
                                f"🎉 <b>Реферальный бонус начислен!</b>\n\n"
                                f"💰 <b>+{bonus}₽</b> добавлено на ваш баланс\n"
                                f"👤 Ваш друг @{referred_username} оплатил подписку «{duration}»\n"
                                f"🌍 Сервер: {server_name}\n\n"
                                f"💳 <b>Текущий баланс:</b> {new_balance}₽\n\n"
                                f"Продолжайте приглашать друзей и зарабатывайте 25% от каждой их покупки! 🚀"
                            )
                            
                            try:
                                bot.send_message(referrer_id, notification_text, parse_mode='HTML')
                                logger.info(f"🔗 РЕФЕРАЛ: Уведомление о бонусе отправлено рефереру {referrer_id}")
                            except Exception as notify_err:
                                logger.error(f"🔗 РЕФЕРАЛ: Ошибка отправки уведомления: {notify_err}")
                    else:
                        logger.warning(f"🔗 РЕФЕРАЛ: Реферер {referrer_id} не найден в allowed_users (отключен от программы)")
                else:
                    logger.info(f"🔗 РЕФЕРАЛ: Пользователь {referred_user_id} не имеет реферера")
            else:
                logger.warning(f"🔗 РЕФЕРАЛ: Сумма платежа = 0, бонус не начисляется")
                
        except Exception as bonus_err:
            logger.error(f"🔗 РЕФЕРАЛ: Ошибка при начислении бонуса: {bonus_err}")
            import traceback
            logger.error(f"🔗 РЕФЕРАЛ: Traceback: {traceback.format_exc()}")
        # ==================== КОНЕЦ НАЧИСЛЕНИЯ БОНУСА ====================
        
        return True
        
    except Exception as e:
        logger.error(f"Ошибка в issue_vpn_config: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return False
        
def send_extension_notification(user_id, server_name, expiry_date, duration):
    """Отправляет уведомление о продлении подписки"""
    try:
        expiry_date_obj = datetime.strptime(expiry_date, "%Y-%m-%d %H:%M:%S")
        formatted_date = expiry_date_obj.strftime('%d.%m.%Y')
        
        text = f"""🔄 <b>Подписка успешно продлена!</b>

📋 <b>Информация о подписке:</b>
🖥 Сервер: {server_name}
⏳ Срок: {duration}
📅 Новая дата окончания: {formatted_date}

✅ Ваш VPN ключ остался прежним и продолжает работать.
Для просмотра ключа нажмите «🔑 Мои ключи»"""

        bot.send_message(
            user_id,
            text,
            parse_mode='HTML',
            reply_markup=main_menu_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка отправки уведомления о продлении: {e}")
        
def send_vpn_config_to_user(user_id, key_data, server_name, expiry_date, duration):
    """Отправляет VPN ключ пользователю"""
    try:
        vpn_key = key_data.get('key', '')
        
        # URL видео инструкции
        VIDEO_URL = "https://t.me/karachay_aj"
        
        # Форматируем дату
        expiry_date_obj = datetime.strptime(expiry_date, "%Y-%m-%d %H:%M:%S")
        formatted_date = expiry_date_obj.strftime('%d.%m.%Y')
        
        # 1. Отправляем информацию о подписке
        info_text = f"""✅ <b>Оплата успешна! VPN активирован!</b>

📋 <b>Информация о подписке:</b>
🖥 Сервер: {server_name}
⏳ Срок: {duration}
📅 Действует до: {formatted_date}"""

        bot.send_message(user_id, info_text, parse_mode='HTML')
        
        # 2. Отправляем ключ отдельным сообщением (чистый текст)
        bot.send_message(user_id, vpn_key)
        
        # 3. Генерируем и отправляем QR-код
        qr_filename = f"qr_{user_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.png"
        if generate_qr_code(vpn_key, qr_filename):
            with open(qr_filename, 'rb') as qr_file:
                bot.send_photo(
                    user_id,
                    qr_file,
                    caption="📱 <b>ИЛИ отсканируйте QR-код</b>",
                    parse_mode='HTML'
                )
            os.remove(qr_filename)
        
        # 4. Кнопки для скачивания приложений
        app_markup = types.InlineKeyboardMarkup(row_width=2)
        app_markup.add(
            types.InlineKeyboardButton("📱 Happ Android", url="https://play.google.com/store/apps/details?id=com.happproxy"),
            types.InlineKeyboardButton("📱 Happ iOS", url="https://apps.apple.com/us/app/happ-proxy-utility/id6504287215?l=ru"),
            types.InlineKeyboardButton("📱 V2raytun Android", url="https://play.google.com/store/apps/details?id=com.v2raytun.android"),
            types.InlineKeyboardButton("📱 V2raytun iOS", url="https://apps.apple.com/us/app/v2raytun/id6476628951?l=ru")
        )
        
        bot.send_message(
            user_id,
            "📥 <b>Скачать приложение:</b>",
            parse_mode='HTML',
            reply_markup=app_markup
        )
        
        # 5. Кнопка с видео инструкцией
        video_markup = types.InlineKeyboardMarkup()
        video_markup.add(types.InlineKeyboardButton(
            "🎬 Смотреть видеоинструкцию",
            url=VIDEO_URL
        ))
        
        bot.send_message(
            user_id,
            "📹 <b>Посмотрите видеоинструкцию по настройке:</b>",
            parse_mode='HTML',
            reply_markup=video_markup
        )
        
        # 6. Инструкция текстом
        bot.send_message(
            user_id,
            "📘 Скопируйте ключ из сообщения выше, откройте приложение и импортируйте из буфера обмена.",
            reply_markup=main_menu_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка отправки ключа пользователю {user_id}: {e}")
        bot.send_message(user_id, "❌ Ошибка при отправке VPN ключа.")
        
@bot.message_handler(func=lambda message: message.text == '📤 Добавить ключи массово' and is_admin(message.from_user.id))
def bulk_upload_keys(message):
    """Массовая загрузка VPN ключей"""
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        for server in servers_db.values():
            keyboard.add(types.KeyboardButton(f"📁 Загрузить ключи для {server['name']}"))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "🖥 <b>Массовая загрузка VPN ключей</b>\n\n"
            "Выберите сервер для загрузки ключей.\n\n"
            "💡 <b>Форматы ключей:</b>\n"
            "• vless://uuid@server:port?params#name\n"
            "• trojan://password@server:port#name\n"
            "• vmess://base64_encoded\n"
            "• ss://base64_encoded\n\n"
            "📝 <b>Как загрузить:</b>\n"
            "1. Выберите сервер\n"
            "2. Отправьте текстовое сообщение с ключами\n"
            "3. Каждый ключ с новой строки\n"
            "4. Или отправьте JSON файл с ключами",
            parse_mode='HTML',
            reply_markup=keyboard
        )
        bot.register_next_step_handler(msg, process_bulk_server_selection_keys)
    except Exception as e:
        logger.error(f"Ошибка в bulk_upload_keys: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при запуске массовой загрузки")

def process_bulk_server_selection_keys(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        server_name = message.text.replace('📁 Загрузить ключи для ', '')
        
        server_key = next((k for k, v in servers_db.items() if v['name'] == server_name), None)
        if not server_key:
            bot.send_message(message.chat.id, "❌ Сервер не найден", reply_markup=admin_menu_keyboard())
            return
        
        user_data = get_user_data(message.from_user.id)
        user_data['bulk_upload_server'] = server_name
        user_data['bulk_upload_keys'] = []
        save_user_data(message.from_user.id, user_data)
        
        server_data = servers_db[server_key]
        stats_text = (
            f"📤 <b>Загрузка ключей для {server_name}</b>\n\n"
            f"📊 <b>Текущая статистика:</b>\n"
            f"🆓 Доступно ключей: {len(server_data.get('available_keys', []))}\n"
            f"👥 Используется: {len(server_data.get('used_keys', {}))}\n\n"
            f"📨 <b>Отправьте ключи:</b>\n"
            f"• Текстом (каждый ключ с новой строки)\n"
            f"• Или JSON файлом\n\n"
            f"После отправки нажмите /done для завершения"
        )
        
        msg = bot.send_message(
            message.chat.id,
            stats_text,
            parse_mode='HTML',
            reply_markup=types.ReplyKeyboardRemove()
        )
        
        bot.register_next_step_handler(msg, process_bulk_keys_upload)
        
    except Exception as e:
        logger.error(f"Ошибка в process_bulk_server_selection_keys: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка выбора сервера", reply_markup=admin_menu_keyboard())

def process_bulk_keys_upload(message):
    try:
        user_id = message.from_user.id
        user_data = get_user_data(user_id)
        server_name = user_data.get('bulk_upload_server')
        
        if not server_name:
            bot.send_message(message.chat.id, "❌ Ошибка: сервер не выбран", reply_markup=admin_menu_keyboard())
            return
        
        # Проверяем команду завершения
        if message.text and message.text == '/done':
            send_bulk_keys_summary(message.chat.id, server_name, user_data.get('bulk_upload_keys', []))
            if 'bulk_upload_server' in user_data:
                del user_data['bulk_upload_server']
            if 'bulk_upload_keys' in user_data:
                del user_data['bulk_upload_keys']
            save_user_data(user_id, user_data)
            return
        
        new_keys = []
        
        if message.content_type == 'document' and message.document.file_name.endswith('.json'):
            # Обработка JSON файла
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            keys_data = json.loads(downloaded_file.decode('utf-8'))
            
            if isinstance(keys_data, list):
                for key_item in keys_data:
                    if isinstance(key_item, str):
                        new_keys.append({'key': key_item, 'name': f'VPN-{server_name}'})
                    elif isinstance(key_item, dict):
                        new_keys.append(key_item)
            elif isinstance(keys_data, dict) and 'keys' in keys_data:
                for key_item in keys_data['keys']:
                    if isinstance(key_item, str):
                        new_keys.append({'key': key_item, 'name': f'VPN-{server_name}'})
                    elif isinstance(key_item, dict):
                        new_keys.append(key_item)
        
        elif message.content_type == 'text':
            # Обработка текстовых ключей
            lines = message.text.strip().split('\n')
            for line in lines:
                line = line.strip()
                if line and (line.startswith('vless://') or line.startswith('vmess://') or 
                            line.startswith('trojan://') or line.startswith('ss://')):
                    new_keys.append({'key': line, 'name': f'VPN-{server_name}'})
        
        if new_keys:
            # Добавляем ключи в хранилище
            if not os.path.exists('keys_storage.json'):
                keys_storage = {'keys': []}
            else:
                with open('keys_storage.json', 'r', encoding='utf-8') as f:
                    keys_storage = json.load(f)
            
            for key_data in new_keys:
                key_data['server'] = server_name
                key_data['added_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                keys_storage['keys'].append(key_data)
            
            with open('keys_storage.json', 'w', encoding='utf-8') as f:
                json.dump(keys_storage, f, ensure_ascii=False, indent=2)
            
            # Добавляем в доступные ключи сервера
            server_key = next(k for k, v in servers_db.items() if v['name'] == server_name)
            if 'available_keys' not in servers_db[server_key]:
                servers_db[server_key]['available_keys'] = []
            servers_db[server_key]['available_keys'].extend(new_keys)
            save_data_to_file()
            
            user_data['bulk_upload_keys'].extend([k.get('key', '')[:50] + '...' for k in new_keys])
            save_user_data(user_id, user_data)
            
            bot.send_message(
                message.chat.id,
                f"✅ Добавлено {len(new_keys)} ключей!\n"
                f"Отправьте еще или нажмите /done для завершения",
                reply_markup=types.ReplyKeyboardRemove()
            )
        else:
            bot.send_message(
                message.chat.id,
                "❌ Не найдено валидных VPN ключей.\n"
                "Отправьте ключи в формате vless://, vmess://, trojan:// или ss://",
                reply_markup=types.ReplyKeyboardRemove()
            )
        
        bot.register_next_step_handler(message, process_bulk_keys_upload)
        
    except Exception as e:
        logger.error(f"Ошибка в process_bulk_keys_upload: {e}")
        import traceback
        logger.error(traceback.format_exc())
        bot.send_message(message.chat.id, "❌ Ошибка при загрузке ключей", reply_markup=admin_menu_keyboard())

def send_bulk_keys_summary(chat_id, server_name, uploaded_keys):
    """Отправляет сводку о массовой загрузке ключей"""
    try:
        server_key = next(k for k, v in servers_db.items() if v['name'] == server_name)
        server_data = servers_db[server_key]
        
        text = f"📊 <b>Сводка загрузки для {server_name}</b>\n\n"
        text += f"✅ Загружено ключей: {len(uploaded_keys)}\n\n"
        
        if uploaded_keys:
            text += "<b>Примеры загруженных ключей:</b>\n"
            for i, key in enumerate(uploaded_keys[:5], 1):
                text += f"{i}. <code>{key}</code>\n"
            
            if len(uploaded_keys) > 5:
                text += f"... и еще {len(uploaded_keys) - 5}\n"
        
        text += f"\n📈 <b>Обновленная статистика:</b>\n"
        text += f"🆓 Доступно ключей: {len(server_data.get('available_keys', []))}\n"
        text += f"👥 Используется: {len(server_data.get('used_keys', {}))}\n"
        
        bot.send_message(chat_id, text, parse_mode='HTML', reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка отправки сводки: {e}")
        bot.send_message(chat_id, f"✅ Загружено {len(uploaded_keys)} ключей", reply_markup=admin_menu_keyboard())
        
def approve_payment_from_yookassa(payment_id, chat_id):
    """Выдает конфиг после успешной оплаты через ЮKassa"""
    try:
        payment = get_payment(payment_id)
        if not payment:
            return False
        
        server_name = payment['server']
        user_id = payment['user_id']
        user_data = get_user_data(user_id)
        
        # Получаем конфиг
        config_file = get_random_config(server_name, user_id)
        if not config_file:
            bot.send_message(chat_id, f"❌ Нет доступных конфигов для сервера {server_name}!")
            return False
        
        # Рассчитываем дату окончания
        duration = payment['duration']
        days = SUBSCRIPTION_PLANS.get(duration, {}).get('days', 30)
        expiry_date = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
        
        # Сохраняем подписку
        if 'subscriptions' not in user_data:
            user_data['subscriptions'] = []
        
        user_data['subscriptions'].append({
            'server': server_name,
            'config_file': config_file,
            'purchase_date': payment['approved_at'],
            'expiry_date': expiry_date,
            'payment_id': payment_id,
            'duration': duration
        })
        
        user_data['username'] = payment.get('username')
        save_user_data(user_id, user_data)
        
        # Отправляем пользователю конфиг
        send_config_to_user(user_id, config_file, server_name, expiry_date, duration)
        
        # Уведомляем админа
        bot.send_message(
            ADMIN_ID,
            f"✅ Платеж #{payment_id} подтвержден (ЮKassa)!\n"
            f"👤 Пользователь: @{payment.get('username')}\n"
            f"🖥 Сервер: {server_name}\n"
            f"💰 Сумма: {payment['amount']}\n"
            f"📅 Действует до: {expiry_date}"
        )
        
        return True
        
    except Exception as e:
        logger.error(f"Ошибка в approve_payment_from_yookassa: {e}")
        return False
        
def send_config_to_user(user_id, config_file, server_name, expiry_date, duration):
    """Отправляет конфиг пользователю"""
    try:
        # 1. Информация о подписке
        bot.send_message(
            user_id,
            f"✅ <b>Оплата подтверждена!</b>\n\n"
            f"📋 <b>Информация о подписке:</b>\n"
            f"🖥 Сервер: {server_name}\n"
            f"⏳ Срок: {duration}\n"
            f"📅 Действует до: {expiry_date}\n\n"
            f"<b>Настройка VPN:</b>",
            parse_mode='HTML'
        )
        
        # 2. Отправляем конфиг
        with open(config_file, 'rb') as f:
            bot.send_document(
                user_id,
                f,
                caption="🔑 <b>Ваш VPN ключ</b>\n\n"
                       "1. Сохраните файл на устройство\n"
                       "2. Откройте happ или V2raytun\n"
                       "3. Нажмите «+» → «Импорт из буфера обмена»\n"
                       "4. Выберите этот файл\n"
                       "5. Нажмите «Сохранить»",
                parse_mode='HTML',
                visible_file_name=os.path.basename(config_file)
            )
        
        # 3. QR-код
        with open(config_file, 'r') as f:
            config_content = f.read()
        
        qr_filename = f"{config_file}_qr.png"
        if generate_qr_code(config_content, qr_filename):
            with open(qr_filename, 'rb') as qr_file:
                bot.send_photo(
                    user_id,
                    qr_file,
                    caption="📱 <b>ИЛИ отсканируйте QR-код</b>\n\n"
                           "1. Откройте Happ или V2raytun\n"
                           "2. Нажмите «+» → «Сканировать QR-код»\n"
                           "3. Наведите камеру на QR-код\n"
                           "4. Нажмите «Сохранить»",
                    parse_mode='HTML'
                )
        
        # 4. Видеоинструкция
        inline_markup = types.InlineKeyboardMarkup()
        inline_markup.add(types.InlineKeyboardButton(
            "📺 Смотреть видеоинструкцию",
            url="https://t.me/karachay_aj"
        ))
        
        bot.send_message(
            user_id,
            "📹 Для подробной настройки посмотрите видеоинструкцию:",
            reply_markup=inline_markup
        )
        
        # 5. Клавиатура
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(
            types.KeyboardButton('🔑 Мои ключи'),
            types.KeyboardButton('🛒 Купить VPN | Продлить')
        )
        markup.add(types.KeyboardButton('🛟 Поддержка'))
        
        bot.send_message(
            user_id,
            "Выберите следующее действие:",
            reply_markup=markup
        )
        
    except Exception as e:
        logger.error(f"Ошибка отправки конфига пользователю {user_id}: {e}")
        
@bot.message_handler(func=lambda message: message.text == '🎁 Управление промокодами' and is_admin(message.from_user.id))
def manage_promo_codes(message):
    try:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('➕ Добавить промокод'))
        markup.add(types.KeyboardButton('➖ Удалить промокод'))
        markup.add(types.KeyboardButton('📋 Список промокодов'))
        markup.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(
            message.chat.id,
            "Управление промокодами:",
            reply_markup=markup
        )
    except Exception as e:
        logger.error(f"Ошибка в manage_promo_codes: {e}")

@bot.message_handler(func=lambda message: message.text == '➕ Добавить промокод' and is_admin(message.from_user.id))
def add_promo_code(message):
    try:
        # Создаем клавиатуру с доступными серверами
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        for server in servers_db.values():
            keyboard.add(types.KeyboardButton(f"🎁 {server['name']}"))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "Выберите сервер для промокода:",
            reply_markup=keyboard
        )
        bot.register_next_step_handler(msg, process_promo_server_selection)
    except Exception as e:
        logger.error(f"Ошибка в add_promo_code: {e}")
        
def process_promo_server_selection(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        server_name = message.text.replace('🎁 ', '')
        
        msg = bot.send_message(
            message.chat.id,
            f"Введите данные промокода для {server_name} в формате:\nКод: GERMANY21\nДней: 21",
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(msg, lambda m: process_add_promo(m, server_name))
    except Exception as e:
        logger.error(f"Ошибка в process_promo_server_selection: {e}")

def process_add_promo(message, server_name):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        lines = [line.strip() for line in message.text.split('\n') if line.strip()]
        promo_data = {}
        
        for line in lines:
            if line.startswith('Код:'):
                promo_data['code'] = line.split('Код:')[1].strip()
            elif line.startswith('Дней:'):
                promo_data['days'] = int(line.split('Дней:')[1].strip())
        
        # Устанавливаем выбранный сервер
        promo_data['server'] = server_name
        
        if not all(key in promo_data for key in ['code', 'days']):
            raise ValueError("Неполные данные")
        
        PROMO_CODES[promo_data['code']] = {
            'server': promo_data['server'],
            'days': promo_data['days'],
            'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'created_by': message.from_user.id
        }
        
        with open('promo_codes.json', 'w') as f:
            json.dump(PROMO_CODES, f)
            
        bot.send_message(
            message.chat.id,
            f"✅ Промокод {promo_data['code']} добавлен!\n"
            f"Сервер: {promo_data['server']}\n"
            f"Дней: {promo_data['days']}",
            reply_markup=admin_menu_keyboard()
        )
    except Exception as e:
        logger.error(f"Ошибка добавления промокода: {e}")
        bot.send_message(message.chat.id, "Ошибка формата данных. Используйте формат:\nКод: SERVER21\nДней: 21")

@bot.message_handler(func=lambda message: message.text == '➖ Удалить промокод' and is_admin(message.from_user.id))
def delete_promo_code(message):
    try:
        if not PROMO_CODES:
            bot.send_message(message.chat.id, "Нет промокодов для удаления.")
            return
            
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        for code in PROMO_CODES.keys():
            markup.add(types.KeyboardButton(code))
        markup.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "Выберите промокод для удаления:",
            reply_markup=markup
        )
        bot.register_next_step_handler(msg, process_delete_promo)
    except Exception as e:
        logger.error(f"Ошибка в delete_promo_code: {e}")

def process_delete_promo(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        code = message.text.strip()
        if code in PROMO_CODES:
            del PROMO_CODES[code]
            with open('promo_codes.json', 'w') as f:
                json.dump(PROMO_CODES, f)
            bot.send_message(
                message.chat.id,
                f"Промокод {code} успешно удален!",
                reply_markup=admin_menu_keyboard()
            )
        else:
            bot.send_message(message.chat.id, "Промокод не найден.")
    except Exception as e:
        logger.error(f"Ошибка в process_delete_promo: {e}")

# Добавляем обработчики для новых кнопок
@bot.message_handler(func=lambda message: message.text == '💾 Создать резервную копию' and is_admin(message.from_user.id))
def backup_data(message):
    try:
        backup_file = create_backup()
        if backup_file:
            with open(backup_file, 'rb') as f:
                bot.send_document(
                    message.chat.id,
                    f,
                    caption="✅ Резервная копия создана успешно!",
                    visible_file_name=os.path.basename(backup_file)
                )
        else:
            bot.send_message(message.chat.id, "❌ Ошибка создания резервной копии")
    except Exception as e:
        logger.error(f"Ошибка в backup_data: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка создания резервной копии")

@bot.message_handler(func=lambda message: message.text == '📥 Восстановить из копии' and is_admin(message.from_user.id))
def restore_data(message):
    try:
        msg = bot.send_message(
            message.chat.id,
            "📤 Отправьте файл резервной копии (ZIP архив):",
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(msg, process_restore_file)
    except Exception as e:
        logger.error(f"Ошибка в restore_data: {e}")
        
def process_restore_file(message):
    try:
        if message.content_type != 'document':
            bot.send_message(message.chat.id, "❌ Пожалуйста, отправьте ZIP файл")
            return
        
        if not message.document.file_name.endswith('.zip'):
            bot.send_message(message.chat.id, "❌ Файл должен быть в формате ZIP")
            return
        
        # Скачиваем файл
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        # Сохраняем временный файл
        temp_file = 'temp_backup.zip'
        with open(temp_file, 'wb') as f:
            f.write(downloaded_file)
        
        # Восстанавливаем из резервной копии
        success = restore_from_backup(temp_file)
        
        # Удаляем временный файл
        os.remove(temp_file)
        
        if success:
            bot.send_message(
                message.chat.id,
                "✅ Данные успешно восстановлены из резервной копии!\n"
                "Все существующие данные были объединены с резервной копией.",
                reply_markup=admin_menu_keyboard()
            )
        else:
            bot.send_message(
                message.chat.id,
                "❌ Ошибка восстановления данных",
                reply_markup=admin_menu_keyboard()
            )
            
    except Exception as e:
        logger.error(f"Ошибка в process_restore_file: {e}")
        bot.send_message(
            message.chat.id,
            "❌ Ошибка обработки файла",
            reply_markup=admin_menu_keyboard()
        )

@bot.message_handler(func=lambda message: message.text == '📋 Список промокодов' and is_admin(message.from_user.id))
def list_promo_codes(message):
    try:
        if not PROMO_CODES:
            bot.send_message(message.chat.id, "Нет активных промокодов.")
            return
            
        text = "📋 Список промокодов:\n\n"
        for code, data in PROMO_CODES.items():
            text += f"🔹 Код: <code>{code}</code>\n"
            text += f"🌍 Сервер: {data['server']}\n"
            text += f"⏳ Дней: {data['days']}\n"
            text += f"📅 Создан: {data['created_at']}\n\n"
        
        bot.send_message(
            message.chat.id,
            text,
            parse_mode='HTML',
            reply_markup=admin_menu_keyboard()
        )
    except Exception as e:
        logger.error(f"Ошибка в list_promo_codes: {e}")
        
@bot.message_handler(func=lambda message: message.text == '📲 Установить приложение')
def install_app(message):
    markup = types.InlineKeyboardMarkup()
    markup.row(
            types.InlineKeyboardButton("📱 Happ Android", url="https://play.google.com/store/apps/details?id=com.happproxy"),
            types.InlineKeyboardButton("📱 Happ iOS", url="https://apps.apple.com/us/app/happ-proxy-utility/id6504287215?l=ru"),
            types.InlineKeyboardButton("📱 V2raytun Android", url="https://play.google.com/store/apps/details?id=com.v2raytun.android"),
            types.InlineKeyboardButton("📱 V2raytun iOS", url="https://apps.apple.com/us/app/v2raytun/id6476628951?l=ru")
    )
    
    bot.send_message(
        message.chat.id,
        "📲 Скачайте Happ или V2raytun для вашей платформы:",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: message.text == 'ℹ️ История покупок')
def purchase_history(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        user_data = get_user_data(user_id)
        if 'subscriptions' not in user_data or not user_data['subscriptions']:
            bot.send_message(user_id, "У вас пока нет завершенных покупок.")
            return
        
        text = "📋 История ваших покупок:\n\n"
        for idx, sub in enumerate(reversed(user_data['subscriptions']), 1):
            expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
            is_expired = expiry_date < datetime.now()
            
            text += f"{idx}. Сервер: {sub['server']}\n"
            text += f"Файл: {sub['config_file']}\n"
            text += f"Срок: {sub.get('duration', 'N/A')}\n"
            text += f"Дата покупки: {sub['purchase_date']}\n"
            text += f"Действует до: {sub['expiry_date']}"
            
            if is_expired:
                text += " (⚠️ Истек)\n\n"
            else:
                text += " (✅ Активен)\n\n"
            
            # Отправляем файл конфига
            if os.path.exists(sub['config_file']):
                with open(sub['config_file'], 'rb') as f:
                    bot.send_document(
                        user_id,
                        f,
                        caption=f"Конфигурация для покупки #{idx}",
                        visible_file_name=os.path.basename(sub['config_file'])
                    )
            else:
                bot.send_message(user_id, f"Файл конфигурации {sub['config_file']} не найден.")
        
        bot.send_message(user_id, text)
    except Exception as e:
        logger.error(f"Ошибка в purchase_history: {e}")

@bot.message_handler(func=lambda message: message.text == '🛟 Поддержка')
def support(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        bot.send_message(message.chat.id, "По всем вопросам обращайтесь к @thetemirbolatov")
    except Exception as e:
        logger.error(f"Ошибка в support: {e}")

@bot.message_handler(func=lambda message: message.text == '⭐⭐⭐ Отзывы')
def show_reviews(message):
    try:
        markup = types.InlineKeyboardMarkup()
        btn_reviews = types.InlineKeyboardButton("📢 Посмотреть отзывы", url="https://t.me/karachay_aj/28")
        markup.add(btn_reviews)
        bot.send_message(
            message.chat.id,
            "Отзывы реальных людей у нас на канале:",
            reply_markup=markup
        )
    except Exception as e:
        logger.error(f"Ошибка в show_reviews: {e}")
        
# Админ-обработчики
@bot.message_handler(func=lambda message: message.text == '📊 Статистика' and is_admin(message.from_user.id))
def stats(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        total_users = len(users_db)
        total_payments = len(payments_db)
        approved_payments = len([p for p in payments_db.values() if p.get('status') == 'approved'])
        pending_payments = len([p for p in payments_db.values() if p.get('status') == 'pending'])
        
        # Правильный расчет дохода
        revenue = 0
        for p in payments_db.values():
            if p.get('status') == 'approved':
                amount_str = p.get('amount', '0₽')
                # Убираем символ валюты и преобразуем в число
                try:
                    amount_num = int(''.join(filter(str.isdigit, str(amount_str))))
                    revenue += amount_num
                except:
                    pass
        
        # Подсчет пользователей с бесплатной подпиской, которые потом купили
        converted_users = 0
        
        # Подсчет продлений - ИСПРАВЛЕННАЯ ЛОГИКА
        extended_count = 0
        users_with_extensions = set()
        key_extensions = {}  # изменено с config_extensions на key_extensions
        
        for user_id, user_data in users_db.items():
            if 'subscriptions' in user_data:
                # Считаем количество платежей для каждого ключа
                key_payments = {}
                for sub in user_data['subscriptions']:
                    if 'payment_id' in sub:  # Только платные подписки
                        # Используем ключ вместо config_file
                        key_data = sub.get('key_data', {})
                        key_value = key_data.get('key', '') if isinstance(key_data, dict) else ''
                        if key_value:
                            # Используем первые 30 символов ключа как идентификатор
                            key_id = key_value[:30]
                            if key_id not in key_payments:
                                key_payments[key_id] = 0
                            key_payments[key_id] += 1
                
                # Если у ключа больше 1 платежа - это продление
                for key_id, payment_count in key_payments.items():
                    if payment_count > 1:
                        extensions = payment_count - 1
                        extended_count += extensions
                        users_with_extensions.add(user_id)
                        
                        if key_id not in key_extensions:
                            key_extensions[key_id] = 0
                        key_extensions[key_id] += extensions
        
        # Формируем детальную информацию о продлениях
        extensions_text = ""
        if extended_count > 0:
            extensions_text = f"\n🔁 <b>Продления:</b> {extended_count}\n"
            extensions_text += f"👥 <b>Пользователей с продлениями:</b> {len(users_with_extensions)}\n\n"
            
            # Топ продлеваемых ключей
            sorted_keys = sorted(key_extensions.items(), key=lambda x: x[1], reverse=True)[:5]
            extensions_text += "<b>Топ продлеваемых ключей:</b>\n"
            
            for key_id, extensions_count in sorted_keys:
                short_key = key_id
                if len(short_key) > 20:
                    short_key = short_key[:17] + "..."
                extensions_text += f"• {short_key}: {extensions_count} продл.\n"
        else:
            extensions_text = f"\n🔁 <b>Продления:</b> 0\n👥 <b>Пользователей с продлениями:</b> 0\n"
        
        bot.send_message(message.chat.id, f"""
📊 <b>Статистика:</b>
👥 <b>Пользователей:</b> {total_users}
💳 <b>Платежей:</b> {total_payments}
✅ <b>Подтверждено:</b> {approved_payments}
⏳ <b>Ожидает:</b> {pending_payments}
💰 <b>Доход:</b> {revenue}₽
🔄 <b>Конверсия (промо → покупка):</b> {converted_users} пользователей
{extensions_text}
""", parse_mode='HTML', reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка в stats: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        bot.send_message(message.chat.id, f"❌ Ошибка при получении статистики: {str(e)}", reply_markup=admin_menu_keyboard())
        
@bot.message_handler(func=lambda message: message.text == '🔁 Статистика продлений' and is_admin(message.from_user.id))
def extensions_stats(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        extended_keys = {}
        user_extensions = {}
        
        # Собираем статистику по продлениям
        for user_id, user_data in users_db.items():
            if 'subscriptions' in user_data:
                key_payments = {}
                
                for sub in user_data['subscriptions']:
                    if 'payment_id' in sub:  # Только платные подписки
                        key_data = sub.get('key_data', {})
                        key_value = key_data.get('key', '') if isinstance(key_data, dict) else ''
                        if key_value:
                            key_id = key_value[:30]
                            if key_id not in key_payments:
                                key_payments[key_id] = 0
                            key_payments[key_id] += 1
                
                # Считаем продления для каждого ключа
                for key_id, payment_count in key_payments.items():
                    if payment_count > 1:  # Если больше 1 платежа - есть продления
                        extensions_count = payment_count - 1
                        
                        if key_id not in extended_keys:
                            extended_keys[key_id] = 0
                        extended_keys[key_id] += extensions_count
                        
                        if user_id not in user_extensions:
                            user_extensions[user_id] = {
                                'username': user_data.get('username', 'N/A'),
                                'total_extensions': 0,
                                'keys': {}
                            }
                        user_extensions[user_id]['total_extensions'] += extensions_count
                        user_extensions[user_id]['keys'][key_id] = extensions_count
        
        if not extended_keys:
            bot.send_message(message.chat.id, "❌ Нет данных о продлениях.", reply_markup=admin_menu_keyboard())
            return
        
        # Формируем подробный отчет
        text = "📈 <b>Детальная статистика продлений</b>\n\n"
        
        # Общая информация
        total_extensions = sum(extended_keys.values())
        total_users_with_extensions = len(user_extensions)
        
        text += f"🔁 <b>Всего продлений:</b> {total_extensions}\n"
        text += f"👥 <b>Пользователей с продлениями:</b> {total_users_with_extensions}\n"
        text += f"🔑 <b>Продлеваемых ключей:</b> {len(extended_keys)}\n\n"
        
        # Топ ключей по продлениям
        text += "🏆 <b>Топ ключей по продлениям:</b>\n"
        sorted_keys = sorted(extended_keys.items(), key=lambda x: x[1], reverse=True)[:10]
        
        for i, (key_id, extensions_count) in enumerate(sorted_keys, 1):
            short_key = key_id
            if len(short_key) > 25:
                short_key = short_key[:22] + "..."
            text += f"{i}. <code>{short_key}</code>: {extensions_count} продл.\n"
        
        text += "\n🏆 <b>Топ пользователей по продлениям:</b>\n"
        sorted_users = sorted(user_extensions.items(), key=lambda x: x[1]['total_extensions'], reverse=True)[:5]
        
        for i, (user_id, data) in enumerate(sorted_users, 1):
            text += f"{i}. @{data['username']}: {data['total_extensions']} продл.\n"
            for key_id, ext_count in data['keys'].items():
                short_key = key_id
                if len(short_key) > 15:
                    short_key = short_key[:12] + "..."
                text += f"   • <code>{short_key}</code>: {ext_count}\n"
        
        bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка в extensions_stats: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())

@bot.message_handler(func=lambda message: message.text == '📝 Список серверов' and is_admin(message.from_user.id))
def manage_servers(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        text = "🖥 <b>Серверы VPN:</b>\n\n"
        
        # Пересчитываем статистику для каждого сервера
        for server_key, server_data in servers_db.items():
            server_name = server_data['name']
            
            # Получаем актуальные данные из базы
            used_keys_count = len(server_data.get('used_keys', {}))
            available_keys_count = len(server_data.get('available_keys', []))
            total_keys_count = used_keys_count + available_keys_count
            
            text += f"<b>{server_name}</b>\n"
            text += f"📍 Локация: {server_data['location']}\n"
            text += f"🖥 IP: {server_data.get('ip', 'N/A')}\n"
            text += f"🔐 Протокол: {server_data.get('protocol', 'VLESS/V2Ray')}\n"
            text += f"⚡ Нагрузка: {server_data.get('load', 'Low')}\n"
            text += f"📊 Всего ключей: {total_keys_count}\n"
            text += f"📈 Доступно ключей: {available_keys_count}\n"
            text += f"👥 Используется: {used_keys_count}\n"
            
            # Показываем несколько примеров доступных ключей (обрезанных)
            if server_data.get('available_keys'):
                text += f"📁 Примеры доступных ключей:\n"
                for i, key_data in enumerate(server_data['available_keys'][:2], 1):
                    key_preview = key_data.get('key', '')[:40] + '...' if len(key_data.get('key', '')) > 40 else key_data.get('key', '')
                    text += f"   {i}. <code>{key_preview}</code>\n"
                if len(server_data['available_keys']) > 2:
                    text += f"   ... и еще {len(server_data['available_keys']) - 2}\n"
            
            if server_data.get('used_keys'):
                text += f"🔐 Примеры используемых ключей:\n"
                used_keys_list = list(server_data['used_keys'].values())
                for i, key_data in enumerate(used_keys_list[:2], 1):
                    key_preview = key_data.get('key', '')[:40] + '...' if len(key_data.get('key', '')) > 40 else key_data.get('key', '')
                    text += f"   {i}. <code>{key_preview}</code>\n"
                if len(used_keys_list) > 2:
                    text += f"   ... и еще {len(used_keys_list) - 2}\n"
                    
            text += "\n"
        
        # Добавляем общую статистику
        total_used = sum(len(server.get('used_keys', {})) for server in servers_db.values())
        total_available = sum(len(server.get('available_keys', [])) for server in servers_db.values())
        total_keys = total_used + total_available
        
        text += f"📈 <b>Общая статистика:</b>\n"
        text += f"📁 Всего ключей в системе: {total_keys}\n"
        text += f"🔐 Всего используется: {total_used}\n"
        text += f"📦 Всего доступно: {total_available}\n"
        
        # Предупреждение если мало ключей
        if total_available < 5:
            text += f"\n⚠️ <b>ВНИМАНИЕ!</b> Осталось мало доступных ключей! Добавьте новые ключи."
        
        bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка в manage_servers: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())
        
        
@bot.message_handler(func=lambda message: message.text == '🔄 Синхронизировать ключи' and is_admin(message.from_user.id))
def sync_keys(message):
    """Синхронизирует ключи между хранилищем и базой данных"""
    try:
        # Загружаем ключи из хранилища
        all_keys = []
        if os.path.exists('keys_storage.json'):
            with open('keys_storage.json', 'r', encoding='utf-8') as f:
                keys_storage = json.load(f)
                all_keys = keys_storage.get('keys', [])
        
        logger.info(f"=== СИНХРОНИЗАЦИЯ КЛЮЧЕЙ ===")
        logger.info(f"Найдено ключей в хранилище: {len(all_keys)}")
        
        # Группируем ключи по серверам
        germany_keys = []
        netherlands_keys = []
        unknown_keys = []
        
        for key_data in all_keys:
            key_string = key_data.get('key', '').lower()
            server_tag = key_data.get('server', '').lower()
            
            is_germany = False
            is_netherlands = False
            
            germany_patterns = ['grm', 'germany', 'de', 'ger', 'frankfurt', '🇩🇪']
            netherlands_patterns = ['ndr', 'netherlands', 'nl', 'ned', 'amsterdam', 'holland', '🇳🇱']
            
            for pattern in germany_patterns:
                if pattern in server_tag or pattern in key_string:
                    is_germany = True
                    break
            
            for pattern in netherlands_patterns:
                if pattern in server_tag or pattern in key_string:
                    is_netherlands = True
                    break
            
            if is_germany and not is_netherlands:
                germany_keys.append(key_data)
            elif is_netherlands and not is_germany:
                netherlands_keys.append(key_data)
            else:
                unknown_keys.append(key_data)
        
        # Обновляем данные для серверов
        for server_key, server_data in servers_db.items():
            server_name = server_data['name']
            
            if 'Germany' in server_name or '🇩🇪' in server_name:
                used_keys_list = list(server_data.get('used_keys', {}).values())
                available_keys = [k for k in germany_keys if k not in used_keys_list]
                server_data['available_keys'] = available_keys
                
            elif 'Netherlands' in server_name or '🇳🇱' in server_name:
                used_keys_list = list(server_data.get('used_keys', {}).values())
                available_keys = [k for k in netherlands_keys if k not in used_keys_list]
                server_data['available_keys'] = available_keys
        
        save_data_to_file()
        
        # Подсчет итогов
        total_available = sum(len(s.get('available_keys', [])) for s in servers_db.values())
        total_used = sum(len(s.get('used_keys', {})) for s in servers_db.values())
        
        result_text = (
            f"✅ <b>Ключи синхронизированы!</b>\n\n"
            f"📁 <b>Статистика:</b>\n"
            f"🇩🇪 Германия: {len(germany_keys)} ключей\n"
            f"🇳🇱 Нидерланды: {len(netherlands_keys)} ключей\n"
            f"❓ Не определено: {len(unknown_keys)} ключей\n\n"
            f"📊 <b>В базе данных:</b>\n"
            f"📦 Всего доступно: {total_available}\n"
            f"🔐 Всего используется: {total_used}"
        )
        
        if unknown_keys:
            result_text += f"\n\n⚠️ <b>Внимание!</b> {len(unknown_keys)} ключей не определены. Проверьте теги серверов."
        
        bot.send_message(
            message.chat.id,
            result_text,
            parse_mode='HTML',
            reply_markup=admin_menu_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка в sync_keys: {e}")
        import traceback
        logger.error(traceback.format_exc())
        bot.send_message(message.chat.id, f"❌ Ошибка синхронизации: {str(e)}")
        
@bot.message_handler(func=lambda message: message.text == '🔄 Перезагрузить конфиги' and is_admin(message.from_user.id))
def reload_configs(message):
    """Принудительно перезагружает все конфиги из файловой системы"""
    try:
        # Получаем все конфиги
        all_configs = [f for f in os.listdir() if f.endswith('.conf') or f.endswith('.json')]
        
        logger.info(f"Перезагрузка конфигов. Найдено файлов: {len(all_configs)}")
        logger.info(f"Список файлов: {all_configs}")
        
        # Сбрасываем конфиги для всех серверов
        for server_key, server_data in servers_db.items():
            server_name = server_data['name']
            
            # Расширенные паттерны
            patterns = []
            if 'Germany' in server_name or 'DE' in server_name or '🇩🇪' in server_name:
                patterns = ['Grm', 'germany', 'Germany', 'DE', 'de', 'ger', 'Ger']
            elif 'Netherlands' in server_name or 'NL' in server_name or '🇳🇱' in server_name:
                patterns = ['Ndr', 'netherlands', 'Netherlands', 'NL', 'nl', 'ned', 'Ned', 'holland', 'Holland']
            else:
                patterns = [server_name.replace(' ', '_')]
            
            # Ищем конфиги для этого сервера
            server_configs = []
            for config in all_configs:
                config_lower = config.lower()
                for pattern in patterns:
                    if pattern.lower() in config_lower:
                        server_configs.append(config)
                        break
            
            logger.info(f"Сервер {server_name}: найдено конфигов по паттернам {patterns}: {server_configs}")
            
            # Разделяем на используемые и доступные
            used_configs = {}
            available_configs = []
            
            for config in server_configs:
                # Проверяем, используется ли конфиг
                found_user = None
                for uid, data in users_db.items():
                    if 'subscriptions' in data:
                        for sub in data['subscriptions']:
                            if sub.get('config_file') == config:
                                found_user = uid
                                break
                    if found_user:
                        break
                
                if found_user and os.path.exists(config):
                    used_configs[str(found_user)] = config
                    logger.info(f"Конфиг {config} используется пользователем {found_user}")
                elif os.path.exists(config):
                    available_configs.append(config)
                    logger.info(f"Конфиг {config} доступен для выдачи")
                else:
                    logger.info(f"Конфиг {config} не существует, пропускаем")
            
            server_data['available_configs'] = available_configs
            server_data['used_configs'] = used_configs
            
            logger.info(f"Сервер {server_name}: доступных {len(available_configs)}, используемых {len(used_configs)}")
        
        save_data_to_file()
        
        # Подсчет итогов
        total_available = sum(len(server['available_configs']) for server in servers_db.values())
        total_used = sum(len(server['used_configs']) for server in servers_db.values())
        
        bot.send_message(
            message.chat.id,
            f"✅ Конфиги перезагружены!\n"
            f"📁 Найдено конфигов в системе: {len(all_configs)}\n"
            f"🖥 Серверов обновлено: {len(servers_db)}\n"
            f"📦 Всего доступно: {total_available}\n"
            f"🔐 Всего используется: {total_used}",
            reply_markup=admin_menu_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка в reload_configs: {e}")
        import traceback
        logger.error(traceback.format_exc())
        bot.send_message(message.chat.id, f"❌ Ошибка перезагрузки конфигов: {str(e)}")

@bot.message_handler(func=lambda message: message.text == '🧾 Проверить платежи' and is_admin(message.from_user.id))
def check_payments(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        pending_payments = [pid for pid, p in payments_db.items() if p.get('status') == 'pending']
        
        if not pending_payments:
            bot.send_message(message.chat.id, "Нет платежей для проверки.", reply_markup=admin_menu_keyboard())
            return
        
        text = "📋 Платежи для проверки:\n\n"
        for pid in pending_payments[-5:]:  # Показываем последние 5 платежей
            p = payments_db[pid]
            text += f"#{pid}\nПользователь: @{p.get('username', 'N/A')}\nСумма: {p.get('amount', 'N/A')}\nБанк: {p.get('bank', 'N/A')}\n\n"
        
        # Добавляем кнопку "Назад" если больше нет платежей
        if len(pending_payments) <= 1:
            keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
            keyboard.add(
                types.KeyboardButton(f'✅ Подтвердить {pending_payments[-1]}'),
                types.KeyboardButton(f'❌ Отклонить {pending_payments[-1]}'),
                types.KeyboardButton('🔙 Назад')
            )
        else:
            keyboard = payment_verification_keyboard(pending_payments[-1])
        
        bot.send_message(message.chat.id, text, reply_markup=keyboard)
    except Exception as e:
        logger.error(f"Ошибка в check_payments: {e}")

@bot.message_handler(func=lambda message: message.text == '✅ Я оплатил')
def payment_done(message):
    """Удален, так как оплата теперь через ЮKassa"""
    bot.send_message(
        message.chat.id,
        "ℹ️ Оплата теперь проходит через платежную систему.\n"
        "Пожалуйста, нажмите кнопку «Перейти к оплате» в предыдущем сообщении.",
        reply_markup=main_menu_keyboard()
    )
        
# Модифицируем функцию approve_payment
@bot.message_handler(func=lambda message: message.text.startswith('✅ Подтвердить') and is_admin(message.from_user.id))
def approve_payment(message):
    try:
        payment_id = message.text.split()[-1]
        
        if payment_id not in payments_db:
            bot.send_message(message.chat.id, "Платеж не найден.")
            return
        
        payment = payments_db[payment_id]
        payment['status'] = 'approved'
        payment['approved_by'] = message.from_user.id
        payment['approved_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_payment(payment_id, payment)
        
        server_name = payment['server'].split(' (')[0]
        user_id = payment['user_id']
        user_data = get_user_data(user_id)
        
        # Проверяем, это продление или новая подписка
        is_extension = 'selected_config' in user_data
        
        if is_extension:
            # Продление существующей подписки
            config_file = user_data['selected_config']
            # Находим подписку для обновления
            for sub in user_data.get('subscriptions', []):
                if sub['config_file'] == config_file:
                    # Сбрасываем ВСЕ флаги уведомлений при продлении
                    if 'last_warnings' in sub:
                        del sub['last_warnings']
                    if 'expiry_notification_sent' in sub:
                        del sub['expiry_notification_sent']
                    
                    # Рассчитываем новую дату окончания
                    duration = payment['duration']
                    days = SUBSCRIPTION_PLANS.get(duration, {}).get('days', 30)
                    current_expiry = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    if current_expiry > datetime.now():  # Если подписка еще активна, добавляем к текущей дате
                        new_expiry = current_expiry + timedelta(days=days)
                    else:  # Если подписка истекла, начинаем с текущей даты
                        new_expiry = datetime.now() + timedelta(days=days)
                    
                    sub['expiry_date'] = new_expiry.strftime("%Y-%m-%d %H:%M:%S")
                    sub['payment_id'] = payment_id
                    expiry_date = new_expiry.strftime("%Y-%m-%d %H:%M:%S")
                    break
        else:
            # Новая подписка
            config_file = get_random_config(server_name, user_id)
            if not config_file:
                bot.send_message(message.chat.id, f"Нет доступных конфигов для сервера {server_name}!")
                return
            
            # Рассчитываем дату окончания подписки
            duration = payment['duration']
            days = SUBSCRIPTION_PLANS.get(duration, {}).get('days', 30)
            expiry_date = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
            
            # Сохраняем информацию о подписке
            if 'subscriptions' not in user_data:
                user_data['subscriptions'] = []
                
            user_data['subscriptions'].append({
                'server': server_name,
                'config_file': config_file,
                'purchase_date': payment['approved_at'],
                'expiry_date': expiry_date,
                'payment_id': payment_id
            })
        
        user_data['username'] = payment.get('username')
        save_user_data(user_id, user_data)
        
        # 1. Отправляем пользователю информацию о подписке
        bot.send_message(
            user_id,
            f"""✅ Платеж #{payment_id} подтвержден!

📋 Информация о подписке:
🖥 Сервер: {server_name}
⏳ Срок: {payment['duration']}
📅 Действует до: {expiry_date if not is_extension else new_expiry.strftime("%Y-%m-%d %H:%M:%S")}

"""
        )
        
        # 2. Отправляем сам файл конфигурации с инструкцией (БЕЗ КНОПКИ)
        config_text = """1. Сохраните файл (ключ) на телефон

2. Откройте приложение Happ или V2raytun и нажмите "+"

3. Выберите "Создать из буфера обмена" и выберите сохраненный сервер

4. Введите имя подключения "Aj VPN" и "Сохранить"""
        
        with open(config_file, 'rb') as f:
            bot.send_document(
                user_id, 
                (os.path.basename(config_file), f),
                caption=config_text
            )
        
        # 3. Отправляем QR-код с инструкцией (БЕЗ КНОПКИ)
        # Читаем содержимое конфига для QR-кода
        with open(config_file, 'r') as f:
            config_content = f.read()
        
        # Генерируем QR-код
        qr_filename = f"{config_file}_qr.png"
        if generate_qr_code(config_content, qr_filename):
            # Отправляем QR-код с инструкцией
            qr_text = """ИЛИ

1. Отсканируйте QR-код (нажмите "Создать из QR-кода")

2. Введите имя подключения "Aj VPN"

3. Нажмите "Сохранить"""
            
            with open(qr_filename, 'rb') as qr_file:
                bot.send_photo(
                    user_id,
                    qr_file,
                    caption=qr_text
                )
        
        # 4. Отправляем финальное сообщение с инлайн кнопкой видеоинструкции
        final_markup = types.InlineKeyboardMarkup()
        final_markup.add(types.InlineKeyboardButton("📺 Смотреть видеоинструкцию", url="https://t.me/karachay_aj"))
        
        bot.send_message(
            user_id,
            "📹 Для подробной настройки посмотрите видеоинструкцию:",
            reply_markup=final_markup
        )
        
        # 5. Добавляем кнопку для установки приложения
        app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        app_markup.add(types.KeyboardButton('📲 Установить приложение'))
        app_markup.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(
            user_id,
            "Выберите следующее действие:",
            reply_markup=app_markup
        )
        
        bot.send_message(message.chat.id, f"Платеж #{payment_id} подтвержден. Пользователь получил конфиг: {config_file}")
    except Exception as e:
        logger.error(f"Ошибка в approve_payment: {e}")
        bot.send_message(
            ADMIN_ID,
            f"✅ Подписка продлена! Пользователь @{payment.get('username')} "
            f"получил +{days} дней. Конфиг: {config_file}"
        )
        
    except Exception as e:
        logger.error(f"Ошибка в approve_payment: {e}")

@bot.message_handler(func=lambda message: message.text.startswith('❌ Отклонить') and is_admin(message.from_user.id))
def reject_payment(message):
    try:
        payment_id = message.text.split()[-1]
        
        if payment_id not in payments_db:
            bot.send_message(message.chat.id, "Платеж не найден.")
            return
        
        payment = payments_db[payment_id]
        payment['status'] = 'rejected'
        payment['rejected_by'] = message.from_user.id
        payment['rejected_at'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_payment(payment_id, payment)
        
        try:
            bot.send_message(payment['user_id'], f"""
❌ Платеж #{payment_id} отклонен!

Проверьте чек и попробуйте снова.
""")
        except Exception as e:
            logger.error(f"Ошибка уведомления пользователя: {e}")
        
        bot.send_message(message.chat.id, f"Платеж #{payment_id} отклонен.")
    except Exception as e:
        logger.error(f"Ошибка в reject_payment: {e}")

@bot.message_handler(func=lambda message: message.text == '📢 Рассылка' and is_admin(message.from_user.id))
def broadcast_menu(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        msg = bot.send_message(
            message.chat.id, 
            "📢 <b>Рассылка</b>\n\n"
            "Отправьте сообщение для рассылки.\n\n"
            "<b>Доступные форматы:</b>\n"
            "• Текст\n"
            "• Фото + текст\n"
            "• Видео + текст\n"
            "• Документ + текст\n\n"
            "Если хотите добавить инлайн-кнопки, после отправки контента напишите:\n"
            "<code>/buttons количество_кнопок</code>\n"
            "Пример: <code>/buttons 2</code>\n\n"
            "Если кнопки не нужны, просто отправьте сообщение.\n\n"
            "Для отмены рассылки отправьте <code>/start</code>",
            parse_mode='HTML',
            reply_markup=types.ReplyKeyboardRemove()
        )
        # Сохраняем состояние рассылки для админа
        user_data = get_user_data(message.from_user.id)
        user_data['broadcast_mode'] = True
        user_data['broadcast_content'] = None
        user_data['broadcast_caption'] = None
        user_data['broadcast_content_type'] = None
        save_user_data(message.from_user.id, user_data)
        
        bot.register_next_step_handler(msg, process_broadcast_content)
    except Exception as e:
        logger.error(f"Ошибка в broadcast_menu: {e}")
        
def process_broadcast_content(message):
    """Обрабатывает контент для рассылки"""
    try:
        admin_id = message.from_user.id
        user_data = get_user_data(admin_id)
        
        # Проверяем отмену через /start
        if message.text and message.text.startswith('/start'):
            cancel_broadcast(message)
            return
        
        # Проверяем команду для добавления кнопок
        if message.text and message.text.startswith('/buttons'):
            process_buttons_command(message)
            return
        
        # Сохраняем контент
        broadcast_data = {
            'content_type': message.content_type,
            'content_id': None,
            'caption': None,
            'text': None
        }
        
        if message.content_type == 'text':
            broadcast_data['text'] = message.text
        elif message.content_type == 'photo':
            broadcast_data['content_id'] = message.photo[-1].file_id
            broadcast_data['caption'] = message.caption or ""
        elif message.content_type == 'video':
            broadcast_data['content_id'] = message.video.file_id
            broadcast_data['caption'] = message.caption or ""
        elif message.content_type == 'document':
            broadcast_data['content_id'] = message.document.file_id
            broadcast_data['caption'] = message.caption or ""
        elif message.content_type == 'audio':
            broadcast_data['content_id'] = message.audio.file_id
            broadcast_data['caption'] = message.caption or ""
        elif message.content_type == 'voice':
            broadcast_data['content_id'] = message.voice.file_id
            broadcast_data['caption'] = message.caption or ""
        else:
            bot.send_message(
                admin_id,
                "❌ Неподдерживаемый формат сообщения. Отправьте текст, фото, видео, документ, аудио или голосовое сообщение.",
                reply_markup=admin_menu_keyboard()
            )
            # Очищаем режим рассылки
            if 'broadcast_mode' in user_data:
                del user_data['broadcast_mode']
            save_user_data(admin_id, user_data)
            return
        
        # Сохраняем данные
        user_data['broadcast_content'] = broadcast_data
        save_user_data(admin_id, user_data)
        
        # Запрашиваем количество кнопок
        msg = bot.send_message(
            admin_id,
            "✅ Контент сохранен!\n\n"
            "Введите количество инлайн-кнопок (0-5) или отправьте /skip для рассылки без кнопок:\n\n"
            "Пример: <code>2</code>",
            parse_mode='HTML'
        )
        bot.register_next_step_handler(msg, process_broadcast_buttons_count)
        
    except Exception as e:
        logger.error(f"Ошибка в process_broadcast_content: {e}")
        bot.send_message(admin_id, "❌ Ошибка при сохранении контента", reply_markup=admin_menu_keyboard())
        # Очищаем режим рассылки
        user_data = get_user_data(admin_id)
        if 'broadcast_mode' in user_data:
            del user_data['broadcast_mode']
        if 'broadcast_content' in user_data:
            del user_data['broadcast_content']
        save_user_data(admin_id, user_data)
        
def process_buttons_command(message):
    """Обрабатывает команду для добавления кнопок"""
    try:
        admin_id = message.from_user.id
        user_data = get_user_data(admin_id)
        
        # Проверяем что есть сохраненный контент
        if 'broadcast_content' not in user_data:
            bot.send_message(
                admin_id,
                "❌ Сначала отправьте контент для рассылки!",
                reply_markup=admin_menu_keyboard()
            )
            if 'broadcast_mode' in user_data:
                del user_data['broadcast_mode']
            save_user_data(admin_id, user_data)
            return
        
        # Парсим количество кнопок
        parts = message.text.split()
        if len(parts) != 2 or not parts[1].isdigit():
            bot.send_message(
                admin_id,
                "❌ Неверный формат. Используйте: <code>/buttons количество</code>\nПример: <code>/buttons 2</code>",
                parse_mode='HTML'
            )
            bot.register_next_step_handler(message, process_broadcast_content)
            return
        
        buttons_count = int(parts[1])
        if buttons_count < 1 or buttons_count > 5:
            bot.send_message(
                admin_id,
                "❌ Количество кнопок должно быть от 1 до 5!",
                parse_mode='HTML'
            )
            bot.register_next_step_handler(message, process_broadcast_content)
            return
        
        user_data['broadcast_buttons_count'] = buttons_count
        user_data['broadcast_buttons'] = []
        save_user_data(admin_id, user_data)
        
        # Запрашиваем данные для первой кнопки
        msg = bot.send_message(
            admin_id,
            f"🔘 <b>Настройка кнопок ({buttons_count} шт.)</b>\n\n"
            f"Введите данные для кнопки №1 в формате:\n"
            f"<code>Название кнопки | https://t.me/...</code>\n\n"
            f"Пример: <code>Наш канал | https://t.me/Ajland777</code>",
            parse_mode='HTML'
        )
        bot.register_next_step_handler(msg, process_broadcast_button, 1)
        
    except Exception as e:
        logger.error(f"Ошибка в process_buttons_command: {e}")
        bot.send_message(admin_id, "❌ Ошибка при настройке кнопок", reply_markup=admin_menu_keyboard())
        
def process_broadcast_button(message, current_button):
    """Обрабатывает ввод данных для кнопки"""
    try:
        admin_id = message.from_user.id
        user_data = get_user_data(admin_id)
        
        # Проверяем отмену
        if message.text and message.text.startswith('/start'):
            cancel_broadcast(message)
            return
        
        # Проверяем пропуск
        if message.text == '/skip':
            # Пропускаем оставшиеся кнопки
            user_data['broadcast_buttons_count'] = current_button - 1
            save_user_data(admin_id, user_data)
            confirm_and_send_broadcast(message, admin_id, user_data)
            return
        
        # Парсим кнопку
        if '|' not in message.text:
            bot.send_message(
                admin_id,
                "❌ Неверный формат. Используйте: <code>Название | Ссылка</code>",
                parse_mode='HTML'
            )
            msg = bot.send_message(
                admin_id,
                f"Попробуйте снова для кнопки №{current_button}:",
                parse_mode='HTML'
            )
            bot.register_next_step_handler(msg, process_broadcast_button, current_button)
            return
        
        button_text, button_url = message.text.split('|', 1)
        button_text = button_text.strip()
        button_url = button_url.strip()
        
        if not button_text or not button_url:
            bot.send_message(
                admin_id,
                "❌ Название и ссылка не могут быть пустыми!",
                parse_mode='HTML'
            )
            msg = bot.send_message(
                admin_id,
                f"Попробуйте снова для кнопки №{current_button}:",
                parse_mode='HTML'
            )
            bot.register_next_step_handler(msg, process_broadcast_button, current_button)
            return
        
        # Сохраняем кнопку
        user_data['broadcast_buttons'].append({
            'text': button_text,
            'url': button_url
        })
        save_user_data(admin_id, user_data)
        
        total_buttons = user_data.get('broadcast_buttons_count', 0)
        
        # Если это последняя кнопка - отправляем на подтверждение
        if current_button >= total_buttons:
            confirm_and_send_broadcast(message, admin_id, user_data)
        else:
            # Запрашиваем следующую кнопку
            msg = bot.send_message(
                admin_id,
                f"✅ Кнопка №{current_button} сохранена!\n\n"
                f"Введите данные для кнопки №{current_button + 1} (осталось {total_buttons - current_button}):",
                parse_mode='HTML'
            )
            bot.register_next_step_handler(msg, process_broadcast_button, current_button + 1)
        
    except Exception as e:
        logger.error(f"Ошибка в process_broadcast_button: {e}")
        bot.send_message(admin_id, "❌ Ошибка при настройке кнопок", reply_markup=admin_menu_keyboard())
        
def process_broadcast_buttons_count(message):
    """Обрабатывает количество кнопок"""
    try:
        admin_id = message.from_user.id
        user_data = get_user_data(admin_id)
        
        # Проверяем отмену
        if message.text and message.text.startswith('/start'):
            cancel_broadcast(message)
            return
        
        # Проверяем пропуск
        if message.text == '/skip':
            # Рассылка без кнопок
            confirm_and_send_broadcast(message, admin_id, user_data)
            return
        
        # Проверяем количество кнопок
        if not message.text.isdigit():
            bot.send_message(
                admin_id,
                "❌ Введите число от 0 до 5 или /skip для пропуска",
                parse_mode='HTML'
            )
            msg = bot.send_message(
                admin_id,
                "Введите количество кнопок (0-5):",
                parse_mode='HTML'
            )
            bot.register_next_step_handler(msg, process_broadcast_buttons_count)
            return
        
        buttons_count = int(message.text)
        if buttons_count < 0 or buttons_count > 5:
            bot.send_message(
                admin_id,
                "❌ Количество кнопок должно быть от 0 до 5!",
                parse_mode='HTML'
            )
            msg = bot.send_message(
                admin_id,
                "Введите количество кнопок (0-5):",
                parse_mode='HTML'
            )
            bot.register_next_step_handler(msg, process_broadcast_buttons_count)
            return
        
        if buttons_count == 0:
            # Рассылка без кнопок
            confirm_and_send_broadcast(message, admin_id, user_data)
        else:
            user_data['broadcast_buttons_count'] = buttons_count
            user_data['broadcast_buttons'] = []
            save_user_data(admin_id, user_data)
            
            # Запрашиваем данные для первой кнопки
            msg = bot.send_message(
                admin_id,
                f"🔘 <b>Настройка кнопок ({buttons_count} шт.)</b>\n\n"
                f"Введите данные для кнопки №1 в формате:\n"
                f"<code>Название кнопки | https://t.me/...</code>\n\n"
                f"Пример: <code>Наш канал | https://t.me/Ajland777</code>\n\n"
                f"Если хотите пропустить оставшиеся кнопки, отправьте <code>/skip</code>",
                parse_mode='HTML'
            )
            bot.register_next_step_handler(msg, process_broadcast_button, 1)
        
    except Exception as e:
        logger.error(f"Ошибка в process_broadcast_buttons_count: {e}")
        bot.send_message(admin_id, "❌ Ошибка при настройке кнопок", reply_markup=admin_menu_keyboard())

def confirm_and_send_broadcast(message, admin_id, user_data):
    """Подтверждение и отправка рассылки"""
    try:
        broadcast_data = user_data.get('broadcast_content')
        if not broadcast_data:
            bot.send_message(admin_id, "❌ Нет сохраненного контента для рассылки", reply_markup=admin_menu_keyboard())
            return
        
        buttons = user_data.get('broadcast_buttons', [])
        
        # Создаем инлайн-клавиатуру если есть кнопки
        markup = None
        if buttons:
            markup = types.InlineKeyboardMarkup(row_width=1)
            for btn in buttons:
                markup.add(types.InlineKeyboardButton(btn['text'], url=btn['url']))
        
        # Показываем превью
        preview_text = "📢 <b>Предпросмотр рассылки:</b>\n\n"
        
        try:
            if broadcast_data['content_type'] == 'text':
                preview_text += broadcast_data['text']
                bot.send_message(admin_id, preview_text, parse_mode='HTML', reply_markup=markup)
            elif broadcast_data['content_type'] == 'photo':
                bot.send_photo(
                    admin_id,
                    broadcast_data['content_id'],
                    caption=broadcast_data['caption'] or preview_text,
                    parse_mode='HTML',
                    reply_markup=markup
                )
            elif broadcast_data['content_type'] == 'video':
                bot.send_video(
                    admin_id,
                    broadcast_data['content_id'],
                    caption=broadcast_data['caption'] or preview_text,
                    parse_mode='HTML',
                    reply_markup=markup
                )
            elif broadcast_data['content_type'] == 'document':
                bot.send_document(
                    admin_id,
                    broadcast_data['content_id'],
                    caption=broadcast_data['caption'] or preview_text,
                    parse_mode='HTML',
                    reply_markup=markup
                )
            elif broadcast_data['content_type'] == 'audio':
                bot.send_audio(
                    admin_id,
                    broadcast_data['content_id'],
                    caption=broadcast_data['caption'] or preview_text,
                    parse_mode='HTML',
                    reply_markup=markup
                )
            elif broadcast_data['content_type'] == 'voice':
                bot.send_voice(
                    admin_id,
                    broadcast_data['content_id'],
                    caption=broadcast_data['caption'] or preview_text,
                    parse_mode='HTML',
                    reply_markup=markup
                )
        except Exception as e:
            logger.error(f"Ошибка предпросмотра: {e}")
            bot.send_message(admin_id, f"❌ Ошибка предпросмотра: {str(e)}")
            return
        
        # Клавиатура подтверждения
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('✅ Да, отправить рассылку'))
        keyboard.add(types.KeyboardButton('❌ Нет, отменить'))
        
        msg = bot.send_message(
            admin_id,
            "\n\n✅ Отправить эту рассылку всем пользователям?",
            reply_markup=keyboard
        )
        bot.register_next_step_handler(msg, finalize_broadcast)
        
    except Exception as e:
        logger.error(f"Ошибка в confirm_and_send_broadcast: {e}")
        bot.send_message(admin_id, "❌ Ошибка при подготовке рассылки", reply_markup=admin_menu_keyboard())

def finalize_broadcast(message):
    """Финальная отправка рассылки"""
    try:
        admin_id = message.from_user.id
        
        if message.text == '❌ Нет, отменить':
            cancel_broadcast(message)
            return
        
        if message.text != '✅ Да, отправить рассылку':
            bot.send_message(admin_id, "❌ Неверный выбор. Рассылка отменена.", reply_markup=admin_menu_keyboard())
            cancel_broadcast(message)
            return
        
        # Получаем данные рассылки
        user_data = get_user_data(admin_id)
        broadcast_data = user_data.get('broadcast_content')
        buttons = user_data.get('broadcast_buttons', [])
        
        if not broadcast_data:
            bot.send_message(admin_id, "❌ Нет данных для рассылки", reply_markup=admin_menu_keyboard())
            cancel_broadcast(message)
            return
        
        # Создаем клавиатуру
        markup = None
        if buttons:
            markup = types.InlineKeyboardMarkup(row_width=1)
            for btn in buttons:
                markup.add(types.InlineKeyboardButton(btn['text'], url=btn['url']))
        
        # Получаем список пользователей
        users = list(users_db.keys())
        success = 0
        failed = 0
        
        status_msg = bot.send_message(admin_id, "📤 Начинаю рассылку...")
        
        for i, user_id in enumerate(users):
            try:
                if broadcast_data['content_type'] == 'text':
                    bot.send_message(user_id, broadcast_data['text'], parse_mode='HTML', reply_markup=markup)
                elif broadcast_data['content_type'] == 'photo':
                    bot.send_photo(
                        user_id,
                        broadcast_data['content_id'],
                        caption=broadcast_data['caption'],
                        parse_mode='HTML',
                        reply_markup=markup
                    )
                elif broadcast_data['content_type'] == 'video':
                    bot.send_video(
                        user_id,
                        broadcast_data['content_id'],
                        caption=broadcast_data['caption'],
                        parse_mode='HTML',
                        reply_markup=markup
                    )
                elif broadcast_data['content_type'] == 'document':
                    bot.send_document(
                        user_id,
                        broadcast_data['content_id'],
                        caption=broadcast_data['caption'],
                        parse_mode='HTML',
                        reply_markup=markup
                    )
                elif broadcast_data['content_type'] == 'audio':
                    bot.send_audio(
                        user_id,
                        broadcast_data['content_id'],
                        caption=broadcast_data['caption'],
                        parse_mode='HTML',
                        reply_markup=markup
                    )
                elif broadcast_data['content_type'] == 'voice':
                    bot.send_voice(
                        user_id,
                        broadcast_data['content_id'],
                        caption=broadcast_data['caption'],
                        parse_mode='HTML',
                        reply_markup=markup
                    )
                success += 1
                
                # Обновляем статус каждые 10 пользователей
                if i % 10 == 0 and i > 0:
                    try:
                        bot.edit_message_text(
                            f"📤 Рассылка в процессе...\n✅ Отправлено: {success}\n❌ Ошибок: {failed}\nВсего: {len(users)}",
                            admin_id,
                            status_msg.message_id
                        )
                    except:
                        pass
                
                time.sleep(0.1)  # Небольшая задержка чтобы не превысить лимиты
                
            except Exception as e:
                failed += 1
                logger.error(f"Ошибка отправки пользователю {user_id}: {e}")
        
        # Финальный отчет
        result_text = f"""
📢 <b>Результаты рассылки:</b>

✅ <b>Успешно:</b> {success}
❌ <b>Не удалось:</b> {failed}
👥 <b>Всего пользователей:</b> {len(users)}
🔘 <b>С кнопками:</b> {"Да" if buttons else "Нет"} ({len(buttons)} кнопок)
"""
        
        bot.edit_message_text(
            result_text,
            admin_id,
            status_msg.message_id,
            parse_mode='HTML'
        )
        
        # Очищаем данные рассылки
        cancel_broadcast(message)
        
        # Показываем админ-меню
        bot.send_message(admin_id, "✅ Рассылка завершена!", reply_markup=admin_menu_keyboard())
        
    except Exception as e:
        logger.error(f"Ошибка в finalize_broadcast: {e}")
        bot.send_message(admin_id, f"❌ Ошибка при отправке рассылки: {str(e)}", reply_markup=admin_menu_keyboard())
        cancel_broadcast(message)
        
def cancel_broadcast(message):
    """Отменяет рассылку и очищает данные"""
    try:
        admin_id = message.from_user.id
        user_data = get_user_data(admin_id)
        
        # Очищаем все данные рассылки
        if 'broadcast_mode' in user_data:
            del user_data['broadcast_mode']
        if 'broadcast_content' in user_data:
            del user_data['broadcast_content']
        if 'broadcast_buttons_count' in user_data:
            del user_data['broadcast_buttons_count']
        if 'broadcast_buttons' in user_data:
            del user_data['broadcast_buttons']
        save_user_data(admin_id, user_data)
        
        bot.send_message(
            admin_id,
            "❌ Рассылка отменена.",
            reply_markup=admin_menu_keyboard()
        )
    except Exception as e:
        logger.error(f"Ошибка в cancel_broadcast: {e}")

def process_broadcast_message(message):
    try:
        if message.text == '/cancel':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        users = list(users_db.keys())
        success = 0
        failed = 0
        
        for user_id in users:
            try:
                if message.content_type == 'text':
                    bot.send_message(user_id, message.text)
                elif message.content_type == 'photo':
                    bot.send_photo(user_id, message.photo[-1].file_id, caption=message.caption)
                elif message.content_type == 'video':
                    bot.send_video(user_id, message.video.file_id, caption=message.caption)
                elif message.content_type == 'document':
                    bot.send_document(user_id, message.document.file_id, caption=message.caption)
                elif message.content_type == 'audio':
                    bot.send_audio(user_id, message.audio.file_id, caption=message.caption)
                elif message.content_type == 'voice':
                    bot.send_voice(user_id, message.voice.file_id, caption=message.caption)
                
                success += 1
                time.sleep(0.1)
            except Exception as e:
                failed += 1
        
        bot.send_message(message.chat.id, f"""
📢 Результаты рассылки:
✅ Успешно: {success}
❌ Не удалось: {failed}
""", reply_markup=admin_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в process_broadcast_message: {e}")

@bot.message_handler(func=lambda message: message.text == '⚙️ Настройки оплаты' and is_admin(message.from_user.id))
def payment_settings(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        text = "⚙️ Реквизиты:\n\n"
        for method in payment_methods.values():
            text += f"<b>{method['bank']}</b>\nКарта: <code>{method['card_number']}</code>\n\n"
        
        text += "Используйте команды:\n/set_payment - добавить новый способ\n/delete_payment - удалить способ \nПример названия банка писать так Сбербанк"
        
        bot.send_message(message.chat.id, text, parse_mode='HTML')
    except Exception as e:
        logger.error(f"Ошибка в payment_settings: {e}")

@bot.message_handler(func=lambda message: message.text == '🗂 Управление конфигами' and is_admin(message.from_user.id))
def config_management(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        bot.send_message(message.chat.id, "Управление конфигурациями:", reply_markup=config_management_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в config_management: {e}")

# Добавляем команду для загрузки конфигов
@bot.message_handler(func=lambda message: message.text == '📤 Загрузить новый конфиг' and is_admin(message.from_user.id))
def upload_config(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        for server in servers_db.values():
            keyboard.add(types.KeyboardButton(f"📥 Загрузить для {server['name']}"))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id, 
            "Выберите сервер для загрузки конфига:\n\n"
            "Поддерживаемые форматы: .conf, .json",
            reply_markup=keyboard
        )
        bot.register_next_step_handler(msg, process_config_upload)
    except Exception as e:
        logger.error(f"Ошибка в upload_config: {e}")

def process_config_upload(message):
    try:
        if not message.text:
            bot.send_message(message.chat.id, "❌ Не получен текст сообщения.")
            return
            
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        server_name = message.text.replace('📥 Загрузить для ', '')
        msg = bot.send_message(
            message.chat.id, 
            f"Отправьте файл конфигурации для сервера {server_name}:\n\n"
            f"Поддерживаемые форматы: .conf, .json",
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(msg, lambda m: save_config_file(m, server_name))
    except Exception as e:
        logger.error(f"Ошибка в process_config_upload: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при загрузке конфига")

def save_config_file(message, server_name):
    try:
        if message.content_type != 'document':
            bot.send_message(message.chat.id, "Пожалуйста, отправьте файл конфигурации.")
            return
        
        # Проверяем расширение файла
        file_name = message.document.file_name
        if not (file_name.endswith('.conf') or file_name.endswith('.json')):
            bot.send_message(
                message.chat.id, 
                "❌ Неподдерживаемый формат файла. Отправьте .conf или .json файл."
            )
            return
        
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        # Сохраняем файл с оригинальным именем
        config_filename = file_name
        
        with open(config_filename, 'wb') as new_file:
            new_file.write(downloaded_file)
        
        # Добавляем конфиг в доступные для сервера
        server_key = next(k for k, v in servers_db.items() if v['name'] == server_name)
        if config_filename not in servers_db[server_key]['available_configs']:
            servers_db[server_key]['available_configs'].append(config_filename)
        save_data_to_file()
        
        bot.send_message(
            message.chat.id, 
            f"✅ Конфигурация для сервера {server_name} успешно сохранена как {config_filename}!",
            reply_markup=admin_menu_keyboard()
        )
    except Exception as e:
        logger.error(f"Ошибка в save_config_file: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка сохранения файла: {str(e)}")

@bot.message_handler(func=lambda message: message.text == '🗑 Удалить конфиг' and is_admin(message.from_user.id))
def delete_config(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        configs = [f for f in os.listdir() if f.endswith('.conf')]
        if not configs:
            bot.send_message(message.chat.id, "Нет конфигураций для удаления.")
            return
        
        # Сортировка конфигов: сначала Германия, потом Нидерланды, затем по цифрам
        def sort_key(config_name):
            # Определяем приоритет сервера
            if 'Grm' in config_name or 'germany' in config_name.lower() or 'de' in config_name.lower():
                server_priority = 1  # Германия - первый
            elif 'Ndr' in config_name or 'netherlands' in config_name.lower() or 'nl' in config_name.lower():
                server_priority = 2  # Нидерланды - второй
            else:
                server_priority = 3  # Остальные
            
            # Извлекаем цифры из названия
            import re
            numbers = re.findall(r'\d+', config_name)
            number = int(numbers[0]) if numbers else 0
            
            return (server_priority, number)
        
        # Сортируем конфиги
        sorted_configs = sorted(configs, key=sort_key)
        
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        for config in sorted_configs:
            keyboard.add(types.KeyboardButton(config))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(message.chat.id, "Выберите конфиг для удаления:", reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_config_deletion)
    except Exception as e:
        logger.error(f"Ошибка в delete_config: {e}")

def process_config_deletion(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        config_name = message.text
        if os.path.exists(config_name):
            os.remove(config_name)
            
            # Удаляем из базы серверов
            for key, server in list(servers_db.items()):
                if server['config'] == config_name:
                    del servers_db[key]
            
            save_data_to_file()
            bot.send_message(message.chat.id, f"Конфигурация {config_name} успешно удалена!", reply_markup=admin_menu_keyboard())
        else:
            bot.send_message(message.chat.id, "Файл не найден.")
    except Exception as e:
        logger.error(f"Ошибка в process_config_deletion: {e}")

@bot.message_handler(commands=['set_payment'])
def set_payment_method(message):
    if not is_admin(message.from_user.id):
        return
    
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        msg = bot.send_message(message.chat.id, """
Введите данные в формате:
Банк: Название
Карта: Номер
""")
        bot.register_next_step_handler(msg, process_payment_method)
    except Exception as e:
        logger.error(f"Ошибка в set_payment_method: {e}")

def process_payment_method(message):
    try:
        if message.text == '/cancel':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        lines = message.text.split('\n')
        method_data = {}
        
        for line in lines:
            if 'Банк:' in line:
                method_data['bank'] = line.split('Банк:')[1].strip()
            elif 'Карта:' in line:
                method_data['card_number'] = line.split('Карта:')[1].strip()
        
        method_key = method_data['bank'].lower().replace(' ', '_')
        payment_methods[method_key] = method_data
        save_data_to_file()
        
        bot.send_message(message.chat.id, "✅ Реквизиты обновлены!", reply_markup=admin_menu_keyboard())
    except Exception as e:
        logger.error(f"Ошибка в process_payment_method: {e}")

@bot.message_handler(content_types=['photo'])
def handle_receipt(message):
    try:
        user_id = message.from_user.id
        
        # Ищем последний ожидающий платеж пользователя
        user_payments = [p for pid, p in payments_db.items() 
                        if str(p.get('user_id')) == str(user_id) 
                        and p.get('status') == 'pending']
        
        if not user_payments:
            bot.send_message(user_id, "У вас нет ожидающих платежей.")
            return
        
        # Берем последний платеж
        payment = user_payments[-1]
        payment_id = next(pid for pid, p in payments_db.items() if p == payment)
        
        bot.send_photo(ADMIN_ID, message.photo[-1].file_id, caption=f"""
📸 Чек #{payment_id}
Пользователь: @{message.from_user.username}
Сумма: {payment.get('amount', 'N/A')}
Банк: {payment.get('bank', 'N/A')}
""")
        
        bot.send_message(user_id, f"""
📨 Чек #{payment_id} получен!
Ожидайте подтверждения.
""")
    except Exception as e:
        logger.error(f"Ошибка в handle_receipt: {e}")

# Обработчики действий с конфигами
@bot.message_handler(func=lambda message: message.text == '📲 Установить приложение')
def install_wireguard(message):
    try:
        bot.send_message(message.chat.id, "Скачайте WireGuard с официального сайта: https://www.wireguard.com/install/")
    except Exception as e:
        logger.error(f"Ошибка в install_wireguard: {e}")

# Модифицируем функцию import_config
@bot.message_handler(func=lambda message: message.text.startswith('⚙️ Импортировать'))
def import_config(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        user_data = get_user_data(user_id)
        if 'subscriptions' not in user_data or not user_data['subscriptions']:
            bot.send_message(user_id, "У вас нет активных конфигураций.")
            return
        
        last_sub = user_data['subscriptions'][-1]
        config_file = last_sub['config_file']
        
        if not os.path.exists(config_file):
            bot.send_message(user_id, "Файл конфигурации не найден.")
            return
        
        # Проверяем не истекла ли подписка
        expiry_date = datetime.strptime(last_sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
        is_expired = expiry_date < datetime.now()
        
        if is_expired:
            bot.send_message(user_id, "⚠️ Срок действия вашей подписки истек! Для возобновления работы приобретите новый доступ.")
            return
        
        # Читаем содержимое конфига для QR-кода
        with open(config_file, 'r') as f:
            config_content = f.read()
        
        # Генерируем QR-код
        qr_filename = f"{config_file}_qr.png"
        if generate_qr_code(config_content, qr_filename):
            # Отправляем QR-код
            with open(qr_filename, 'rb') as qr_file:
                bot.send_photo(
                    user_id,
                    qr_file,
                    caption="📲 Отсканируйте этот QR-код в приложении Amnezia для быстрой настройки"
                )
        
        # Отправляем инструкцию
        instructions = """📲 <b>Инструкция по установке в Happ:</b>

1. <b>Способ 1: Сканирование QR-кода</b>
   • Откройте Happ
   • Нажмите "+" (Добавить подключение)
   • Выберите "Сканировать QR-код"
   • Наведите камеру на QR-код выше
   • Нажмите "Сохранить"

2. <b>Способ 2: Импорт из файла</b>
   • Нажмите "+" (Добавить подключение)
   • Выберите "Импорт из файла"
   • Найдите и выберите отправленный вам файл конфигурации
   • Введите имя подключения (например: "Мой VPN")
   • Нажмите "Сохранить"

3. <b>Подключение:</b>
   • Выберите созданное подключение в списке
   • Нажмите "Подключиться"
   • Разрешите запрос на создание VPN-подключения

🔹 <b>Важно:</b>
• При первом подключении может потребоваться 1-2 минуты для установки соединения
• Не удаляйте файл конфигурации - он может понадобиться для повторного импорта
• Для автоматического подключения включите "Автоподключение" в настройках Happ

📹 <b>Видеоинструкция:</b> https://t.me/karachay_aj
"""
        
        # Отправляем сам файл конфигурации
        with open(config_file, 'rb') as f:
            bot.send_document(
                user_id,
                f,
                caption=instructions,
                parse_mode='HTML',
                visible_file_name=os.path.basename(config_file))
        
        # Дополнительные кнопки для удобства
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(
            types.KeyboardButton('📲 Установить приложение'),
            types.KeyboardButton('💾 Скачать конфиг'),
            types.KeyboardButton('🔙 Назад')
        )
        
        bot.send_message(
            user_id,
            "Выберите следующее действие:",
            reply_markup=markup
        )
            
    except Exception as e:
        logger.error(f"Ошибка в import_config: {e}")
        bot.send_message(user_id, "Произошла ошибка при подготовке инструкции. Пожалуйста, попробуйте позже.")

@bot.message_handler(commands=['delete_payment'])
def delete_payment_method(message):
    if not is_admin(message.from_user.id):
        return
    
    try:
        if not payment_methods:
            bot.send_message(message.chat.id, "Нет способов оплаты для удаления.")
            return
            
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        for method in payment_methods.values():
            keyboard.add(types.KeyboardButton(f"❌ Удалить {method['bank']}"))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "Выберите способ оплаты для удаления:",
            reply_markup=keyboard
        )
        bot.register_next_step_handler(msg, process_payment_deletion)
    except Exception as e:
        logger.error(f"Ошибка в delete_payment_method: {e}")

def process_payment_deletion(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        bank_name = message.text.replace('❌ Удалить ', '')
        method_key = next(key for key, val in payment_methods.items() if val['bank'] == bank_name)
        
        if method_key:
            del payment_methods[method_key]
            save_data_to_file()
            bot.send_message(
                message.chat.id,
                f"Способ оплаты {bank_name} успешно удален!",
                reply_markup=admin_menu_keyboard()
            )
        else:
            bot.send_message(message.chat.id, "Способ оплаты не найден.")
    except Exception as e:
        logger.error(f"Ошибка в process_payment_deletion: {e}")
        
@bot.message_handler(func=lambda message: message.text == '👥 Список покупателей' and is_admin(message.from_user.id))
def customers_list(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('🔍 Просроченные подписки'))
        keyboard.add(types.KeyboardButton('🗑 Удалить ключ пользователя'))
        keyboard.add(types.KeyboardButton('🔄 Обновить список'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        active_users = []
        current_time = datetime.now()
        
        for user_id, user_data in users_db.items():
            if 'subscriptions' in user_data:
                for sub in user_data['subscriptions']:
                    expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    if expiry_date > current_time:
                        days_left = (expiry_date - current_time).days
                        key_data = sub.get('key_data', {})
                        key_preview = key_data.get('key', '')[:30] + '...' if key_data.get('key') else 'N/A'
                        
                        active_users.append({
                            'user_id': user_id,
                            'username': user_data.get('username', 'N/A'),
                            'server': sub['server'],
                            'key_preview': key_preview,
                            'expiry_date': sub['expiry_date'],
                            'days_left': days_left,
                            'purchase_date': sub.get('purchase_date', 'N/A')
                        })
        
        active_users.sort(key=lambda x: x['days_left'])
        
        if not active_users:
            text = "📋 <b>Активные подписки:</b>\n\nНет активных подписок."
            bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=keyboard)
        else:
            bot.send_message(
                message.chat.id, 
                f"📋 <b>Активные подписки:</b>\n\nВсего активных подписок: {len(active_users)}",
                parse_mode='HTML',
                reply_markup=keyboard
            )
            
            time.sleep(0.5)
            
            chunk_size = 10
            for i in range(0, len(active_users), chunk_size):
                chunk = active_users[i:i + chunk_size]
                text = f"📋 <b>Активные подписки (часть {i//chunk_size + 1}):</b>\n\n"
                
                for user in chunk:
                    text += (f"👤 <b>Пользователь:</b> @{user['username']} (ID: {user['user_id']})\n"
                            f"🖥 <b>Сервер:</b> {user['server']}\n"
                            f"🔑 <b>Ключ:</b> <code>{user['key_preview']}</code>\n"
                            f"📅 <b>Куплено:</b> {user['purchase_date']}\n"
                            f"⏳ <b>Осталось дней:</b> {user['days_left']}\n"
                            f"────────────────────\n")
                
                if len(text) > 4000:
                    lines = text.split('\n')
                    current_chunk = ""
                    for line in lines:
                        if len(current_chunk + line + '\n') > 4000:
                            bot.send_message(message.chat.id, current_chunk, parse_mode='HTML')
                            current_chunk = line + '\n'
                            time.sleep(0.3)
                        else:
                            current_chunk += line + '\n'
                    if current_chunk:
                        bot.send_message(message.chat.id, current_chunk, parse_mode='HTML')
                        time.sleep(0.3)
                else:
                    bot.send_message(message.chat.id, text, parse_mode='HTML')
                    time.sleep(0.3)
        
    except Exception as e:
        logger.error(f"Ошибка в customers_list: {e}")
        bot.send_message(message.chat.id, f"Ошибка при получении списка: {str(e)}")
        
@bot.message_handler(func=lambda message: message.text == '🔍 Просроченные подписки' and is_admin(message.from_user.id))
def expired_subscriptions(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        expired_subscriptions_list = []
        current_time = datetime.now()
        
        for user_id, user_data in users_db.items():
            if 'subscriptions' in user_data:
                username = user_data.get('username', 'N/A')
                
                for sub in user_data['subscriptions']:
                    try:
                        expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                        if expiry_date < current_time:
                            expired_days = (current_time - expiry_date).days
                            
                            # Получаем ключ
                            key_data = sub.get('key_data', {})
                            key_value = key_data.get('key', 'N/A') if isinstance(key_data, dict) else 'N/A'
                            key_preview = key_value[:30] + '...' if len(key_value) > 30 else key_value
                            
                            expired_subscriptions_list.append({
                                'user_id': user_id,
                                'username': username,
                                'server': sub.get('server', 'N/A'),
                                'key_preview': key_preview,
                                'key_full': key_value,
                                'expiry_date': sub.get('expiry_date', 'N/A'),
                                'expired_days': expired_days,
                                'purchase_date': sub.get('purchase_date', 'N/A'),
                                'promo_code': sub.get('promo_code', 'платная')
                            })
                    except:
                        pass
        
        expired_subscriptions_list.sort(key=lambda x: x['expired_days'], reverse=True)
        
        if not expired_subscriptions_list:
            bot.send_message(message.chat.id, "✅ Нет просроченных подписок.")
            return
        
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('🗑 Удалить конкретную подписку'))
        keyboard.add(types.KeyboardButton('🔄 Обновить список'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        text_parts = []
        current_text = "⚠️ <b>Просроченные подписки:</b>\n\n"
        
        for i, sub in enumerate(expired_subscriptions_list, 1):
            user_text = (f"<b>📋 #{i}</b>\n"
                        f"👤 @{sub['username']} (ID: {sub['user_id']})\n"
                        f"🖥 Сервер: {sub['server']}\n"
                        f"🔑 Ключ: <code>{sub['key_preview']}</code>\n"
                        f"📅 Истекла: {sub['expiry_date']} ({sub['expired_days']} дн. назад)\n"
                        f"🎫 Тип: {sub['promo_code']}\n"
                        f"🆔 ID: <code>{sub['user_id']}_{sub['key_preview'][:10]}</code>\n"
                        f"────────────────────\n\n")
            
            if len(current_text + user_text) > 4000:
                text_parts.append(current_text)
                current_text = user_text
            else:
                current_text += user_text
        
        if current_text:
            text_parts.append(current_text)
        
        for i, text_part in enumerate(text_parts):
            if i == 0:
                bot.send_message(message.chat.id, text_part, parse_mode='HTML', reply_markup=keyboard)
            else:
                bot.send_message(message.chat.id, text_part, parse_mode='HTML')
            
            if i < len(text_parts) - 1:
                time.sleep(1)
        
    except Exception as e:
        logger.error(f"Ошибка в expired_subscriptions: {e}")
        bot.send_message(message.chat.id, f"Ошибка: {str(e)}")
        
@bot.message_handler(func=lambda message: message.text == '🗑 Удалить конкретную подписку' and is_admin(message.from_user.id))
def delete_specific_subscription(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        msg = bot.send_message(
            message.chat.id,
            "Введите ID пользователя для удаления подписки:",
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(msg, process_subscription_deletion_step1)
    except Exception as e:
        logger.error(f"Ошибка в delete_specific_subscription: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при удалении подписки")

def process_subscription_deletion_step1(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        user_id = message.text.strip()
        
        if user_id not in users_db:
            bot.send_message(message.chat.id, f"❌ Пользователь с ID {user_id} не найден!")
            return
        
        user_data = users_db[user_id]
        username = user_data.get('username', 'N/A')
        
        if 'subscriptions' not in user_data or not user_data['subscriptions']:
            bot.send_message(message.chat.id, f"❌ У пользователя @{username} нет подписок!")
            return
        
        # Показываем подписки пользователя
        text = f"📋 <b>Подписки пользователя @{username} (ID: {user_id}):</b>\n\n"
        
        for i, sub in enumerate(user_data['subscriptions'], 1):
            key_data = sub.get('key_data', {})
            key_preview = key_data.get('key', '')[:30] + '...' if key_data.get('key') else 'N/A'
            
            text += f"<b>#{i}</b> Сервер: {sub.get('server', 'N/A')}\n"
            text += f"   🔑 Ключ: <code>{key_preview}</code>\n"
            text += f"   📅 До: {sub.get('expiry_date', 'N/A')}\n\n"
        
        text += "Введите номер подписки для удаления (1, 2, 3...):"
        
        # Сохраняем user_id для следующего шага
        user_data_temp = get_user_data(message.from_user.id)
        user_data_temp['delete_sub_user_id'] = user_id
        save_user_data(message.from_user.id, user_data_temp)
        
        msg = bot.send_message(message.chat.id, text, parse_mode='HTML')
        bot.register_next_step_handler(msg, process_subscription_deletion_step2)
        
    except Exception as e:
        logger.error(f"Ошибка в process_subscription_deletion_step1: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка", reply_markup=admin_menu_keyboard())
        
def process_subscription_deletion_step2(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        sub_num = message.text.strip()
        if not sub_num.isdigit():
            bot.send_message(message.chat.id, "❌ Введите номер подписки цифрой!")
            return
        
        sub_index = int(sub_num) - 1
        
        admin_data = get_user_data(message.from_user.id)
        user_id = admin_data.get('delete_sub_user_id')
        
        if not user_id or user_id not in users_db:
            bot.send_message(message.chat.id, "❌ Пользователь не найден!")
            return
        
        user_data = users_db[user_id]
        
        if sub_index < 0 or sub_index >= len(user_data.get('subscriptions', [])):
            bot.send_message(message.chat.id, "❌ Неверный номер подписки!")
            return
        
        # Удаляем подписку
        deleted_sub = user_data['subscriptions'].pop(sub_index)
        
        # Освобождаем ключ
        server_name = deleted_sub.get('server', '')
        key_data = deleted_sub.get('key_data', {})
        
        for server_key, server_data in servers_db.items():
            if server_data['name'] == server_name:
                if 'used_keys' in server_data and user_id in server_data['used_keys']:
                    del server_data['used_keys'][user_id]
                break
        
        save_data_to_file()
        
        # Очищаем временные данные
        if 'delete_sub_user_id' in admin_data:
            del admin_data['delete_sub_user_id']
        save_user_data(message.from_user.id, admin_data)
        
        key_preview = key_data.get('key', '')[:30] + '...' if key_data.get('key') else 'N/A'
        
        bot.send_message(
            message.chat.id,
            f"✅ <b>Подписка удалена!</b>\n\n"
            f"👤 Пользователь: @{user_data.get('username', 'N/A')}\n"
            f"🖥 Сервер: {deleted_sub.get('server', 'N/A')}\n"
            f"🔑 Ключ: <code>{key_preview}</code>\n"
            f"🔓 Ключ освобожден",
            parse_mode='HTML',
            reply_markup=admin_menu_keyboard()
        )
        
        # Уведомляем пользователя
        try:
            bot.send_message(
                user_id,
                f"⚠️ Ваша подписка на сервере {deleted_sub.get('server', 'N/A')} была отключена администратором."
            )
        except:
            pass
        
    except Exception as e:
        logger.error(f"Ошибка в process_subscription_deletion_step2: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка", reply_markup=admin_menu_keyboard())
        
@bot.message_handler(func=lambda message: message.text == '🔄 Обновить список' and is_admin(message.from_user.id))
def refresh_lists(message):
    """Обновляет списки подписок"""
    try:
        # Просто вызываем соответствующую функцию в зависимости от текущего контекста
        if 'просроченные' in message.text.lower() or '🔍' in message.text:
            expired_subscriptions(message)
        else:
            customers_list(message)
    except Exception as e:
        logger.error(f"Ошибка в refresh_lists: {e}")
        bot.send_message(message.chat.id, "Ошибка при обновлении списка.")
        
@bot.message_handler(func=lambda message: message.text == '🗑 Очистить просроченные' and is_admin(message.from_user.id))
def cleanup_expired(message):
    try:
        # Создаем клавиатуру с предупреждением
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('✅ Да, удалить ВСЕ просроченные'))
        keyboard.add(types.KeyboardButton('❌ Нет, вернуться назад'))
        
        bot.send_message(
            message.chat.id,
            "⚠️ <b>ВНИМАНИЕ!</b>\n\n"
            "Эта функция удалит ВСЕ просроченные подписки у ВСЕХ пользователей!\n"
            "Конфигурационные файлы будут полностью удалены с сервера.\n\n"
            "Вы уверены что хотите продолжить?",
            parse_mode='HTML',
            reply_markup=keyboard
        )
    except Exception as e:
        logger.error(f"Ошибка в cleanup_expired: {e}")
        bot.send_message(message.chat.id, f"Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())
        
@bot.message_handler(func=lambda message: message.text == '✅ Да, удалить ВСЕ просроченные' and is_admin(message.from_user.id))
def confirm_cleanup_expired(message):
    try:
        deleted_configs = []
        deleted_files = 0
        affected_users = set()
        
        for user_id, user_data in list(users_db.items()):
            if 'subscriptions' in user_data:
                # Создаем новый список без просроченных подписок
                active_subs = []
                for sub in user_data['subscriptions']:
                    expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    if expiry_date > datetime.now():
                        active_subs.append(sub)
                    else:
                        # УДАЛЯЕМ конфиг полностью
                        config_file = sub['config_file']
                        server_key = next(k for k, v in servers_db.items() if v['name'] == sub['server'])
                        
                        if server_key:
                            # Удаляем из используемых
                            user_id_key = str(user_id)
                            if user_id_key in servers_db[server_key]['used_configs']:
                                del servers_db[server_key]['used_configs'][user_id_key]
                            
                            # Удаляем из доступных
                            if config_file in servers_db[server_key]['available_configs']:
                                servers_db[server_key]['available_configs'].remove(config_file)
                        
                        deleted_configs.append(config_file)
                        affected_users.add(user_id)
                        
                        # Удаляем файл конфига физически
                        if os.path.exists(config_file):
                            try:
                                os.remove(config_file)
                                deleted_files += 1
                                
                                # Удаляем QR-код если есть
                                qr_file = f"{config_file}_qr.png"
                                if os.path.exists(qr_file):
                                    os.remove(qr_file)
                            except Exception as e:
                                logger.error(f"Ошибка удаления файла {config_file}: {e}")
                
                # Обновляем подписки пользователя
                if active_subs:
                    users_db[user_id]['subscriptions'] = active_subs
                else:
                    if 'subscriptions' in users_db[user_id]:
                        del users_db[user_id]['subscriptions']
        
        save_data_to_file()
        
        text = "✅ <b>Массовая очистка завершена:</b>\n\n"
        if deleted_configs:
            text += f"🗑 <b>Удалено подписок:</b> {len(deleted_configs)}\n"
            text += f"👥 <b>Затронуто пользователей:</b> {len(affected_users)}\n"
            text += f"🗑 <b>Удалено файлов конфигов:</b> {deleted_files}\n"
            text += "\n<b>Рекомендуется:</b> Использовать '🗑 Удалить конкретную подписку' для точечного удаления."
        else:
            text += "Не найдено просроченных подписок для очистки."
        
        # Возвращаем стандартную клавиатуру
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('🔍 Просроченные подписки'))
        keyboard.add(types.KeyboardButton('🗑 Удалить конкретную подписку'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=keyboard)
        
    except Exception as e:
        logger.error(f"Ошибка в confirm_cleanup_expired: {e}")
        bot.send_message(message.chat.id, f"Ошибка при очистке: {str(e)}", reply_markup=admin_menu_keyboard())
        
@bot.message_handler(func=lambda message: message.text == '❌ Нет, вернуться назад' and is_admin(message.from_user.id))
def cancel_cleanup(message):
    try:
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('🔍 Просроченные подписки'))
        keyboard.add(types.KeyboardButton('🗑 Удалить конкретную подписку'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(
            message.chat.id,
            "✅ Массовая очистка отменена.",
            reply_markup=keyboard
        )
    except Exception as e:
        logger.error(f"Ошибка в cancel_cleanup: {e}")                        
                
@bot.message_handler(func=lambda message: message.text == 'Активировать промокод')
def promo_code_handler(message):
    try:
        if not check_channel_subscription(message.from_user.id):
            return
            
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "✏️ Введите промокод:",
            reply_markup=markup
        )
        bot.register_next_step_handler(msg, process_promo_code)
    except Exception as e:
        logger.error(f"Ошибка в promo_code_handler: {e}")

# Находим функцию process_promo_code и изменяем часть отправки сообщений:

def process_promo_code(message):
    try:
        user_id = message.from_user.id
        
        # URL видео инструкции
        VIDEO_URL = "https://t.me/karachay_aj"
        
        if message.text == '🔙 Назад':
            bot.send_message(user_id, "Возвращаемся в главное меню", reply_markup=main_menu_keyboard())
            return
        
        # Проверяем, что сообщение содержит текст
        if not message.text or not hasattr(message, 'text'):
            bot.send_message(user_id, "❌ Пожалуйста, введите текстовый промокод.")
            return
        
        promo_code = message.text.strip().upper()
        
        # Проверяем, что промокод не пустой
        if not promo_code:
            bot.send_message(user_id, "❌ Промокод не может быть пустым!")
            return
            
        # Проверяем, активировал ли пользователь уже этот промокод
        user_data = get_user_data(user_id)
        if 'used_promo_codes' in user_data and promo_code in user_data['used_promo_codes']:
            bot.send_message(user_id, "❌ Вы уже активировали этот промокод ранее!")
            return
            
        if promo_code not in PROMO_CODES:
            bot.send_message(user_id, "⚠️ Неверный промокод! Вводите только заглавными буквами.")
            return
            
        promo_data = PROMO_CODES[promo_code]
        server_name = promo_data['server']
        
        # Получаем случайный ключ для выбранного сервера
        key_data = get_random_config(server_name, user_id)
        if not key_data:
            bot.send_message(user_id, "⚠️ На сервере закончились свободные ключи. Попробуйте позже.")
            return
            
        expiry_date = (datetime.now() + timedelta(days=promo_data['days'])).strftime("%Y-%m-%d %H:%M:%S")
        
        # Сохраняем информацию о подписке
        if 'subscriptions' not in user_data:
            user_data['subscriptions'] = []
            
        user_data['subscriptions'].append({
            'server': server_name,
            'key_data': key_data,
            'purchase_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'expiry_date': expiry_date,
            'type': 'promo',
            'promo_code': promo_code
        })
        
        # Добавляем промокод в список использованных
        if 'used_promo_codes' not in user_data:
            user_data['used_promo_codes'] = []
        user_data['used_promo_codes'].append(promo_code)
        
        save_user_data(user_id, user_data)
        
        # Получаем ключ
        vpn_key = key_data.get('key', '')

        # 1. Отправляем информацию об активации
        bot.send_message(
            user_id,
            f"🎉 <b>Промокод активирован!</b>\n\n"
            f"📋 <b>Информация о подписке:</b>\n"
            f"🌍 Сервер: {server_name}\n"
            f"⏳ Срок: {promo_data['days']} дней\n"
            f"📅 Активен до: {expiry_date}",
            parse_mode='HTML'
        )
        
        # 2. Отправляем чистый ключ отдельным сообщением
        bot.send_message(
            user_id,
            f"<pre><code class=\"language-text\">{vpn_key}</code></pre>\n\n"
            "<i>👆 Нажмите на ключ чтобы скопировать его в буфер обмена</i>",
            parse_mode='HTML'
        )
        
        # 3. Генерируем и отправляем QR-код
        qr_filename = f"qr_promo_{user_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.png"
        if generate_qr_code(vpn_key, qr_filename):
            with open(qr_filename, 'rb') as qr_file:
                bot.send_photo(
                    user_id,
                    qr_file,
                    caption="📱 <b>ИЛИ отсканируйте QR-код</b>",
                    parse_mode='HTML'
                )
            os.remove(qr_filename)
        
        # 4. Кнопки для скачивания приложений
        time.sleep(1)
        
        app_markup = types.InlineKeyboardMarkup(row_width=2)
        app_markup.add(
            types.InlineKeyboardButton("📱 Happ Android", url="https://play.google.com/store/apps/details?id=com.happproxy"),
            types.InlineKeyboardButton("📱 Happ iOS", url="https://apps.apple.com/us/app/happ-proxy-utility/id6504287215?l=ru"),
            types.InlineKeyboardButton("📱 V2raytun Android", url="https://play.google.com/store/apps/details?id=com.v2raytun.android"),
            types.InlineKeyboardButton("📱 V2raytun iOS", url="https://apps.apple.com/us/app/v2raytun/id6476628951?l=ru")
        )
        
        bot.send_message(
            user_id,
            "📥 <b>Скачать приложение:</b>",
            parse_mode='HTML',
            reply_markup=app_markup
        )
        
        # 5. Кнопка с видео инструкцией
        video_markup = types.InlineKeyboardMarkup()
        video_markup.add(types.InlineKeyboardButton(
            "🎬 Смотреть видеоинструкцию",
            url=VIDEO_URL
        ))
        
        bot.send_message(
            user_id,
            "📹 <b>Посмотрите видеоинструкцию по настройке:</b>",
            parse_mode='HTML',
            reply_markup=video_markup
        )
        
        # 6. Инструкция текстом
        instruction_text = """📘 <b>Как подключить:</b>

1️⃣ Установите приложение (ссылки выше)

2️⃣ Скопируйте ключ (нажмите на него выше)

3️⃣ Откройте приложение → «+» → «Импорт из буфера»

4️⃣ Нажмите «Подключиться» ✅"""

        bot.send_message(
            user_id,
            instruction_text,
            parse_mode='HTML',
            reply_markup=main_menu_keyboard()
        )
        
        # Уведомляем админа
        key_preview = vpn_key[:30] + '...' if len(vpn_key) > 30 else vpn_key
        bot.send_message(
            ADMIN_ID,
            f"🔔 <b>Промокод активирован!</b>\n\n"
            f"🎫 Код: {promo_code}\n"
            f"👤 Пользователь: @{message.from_user.username} (ID: {user_id})\n"
            f"🖥 Сервер: {server_name}\n"
            f"📅 Дней: {promo_data['days']}\n"
            f"🔑 Ключ: <code>{key_preview}</code>",
            parse_mode='HTML'
        )
        
    except Exception as e:
        logger.error(f"Ошибка в process_promo_code: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        bot.send_message(message.chat.id, "❌ Произошла ошибка при обработке промокода.")
        
        
@bot.callback_query_handler(func=lambda call: call.data.startswith('extend:'))
def handle_extend_callback(call):
    try:
        user_id = call.from_user.id
        config_file = call.data.split(':')[1]
        
        # Находим подписку пользователя по файлу конфига
        user_data = get_user_data(user_id)
        subscription = None
        
        for sub in user_data.get('subscriptions', []):
            if sub['config_file'] == config_file:
                subscription = sub
                break
        
        if not subscription:
            bot.answer_callback_query(call.id, "Подписка не найдена!")
            return
            
        # Сохраняем выбранный сервер для продления
        user_data['selected_server'] = subscription['server']
        save_user_data(user_id, user_data)
        
        # Предлагаем выбрать срок продления
        bot.send_message(
            user_id,
            f"Вы выбрали продление подписки для сервера {subscription['server']}.\nВыберите срок продления:",
            reply_markup=duration_menu_keyboard()
        )
        
        bot.answer_callback_query(call.id)
    except Exception as e:
        logger.error(f"Ошибка в handle_extend_callback: {e}")
        bot.answer_callback_query(call.id, "Произошла ошибка!")
        
@bot.message_handler(func=lambda message: message.text == '🔑 Мои ключи')
def my_keys_handler(message):
    try:
        user_id = message.from_user.id
        delete_previous_message(user_id, message.message_id - 1)
        
        user_data = get_user_data(user_id)
        if 'subscriptions' not in user_data or not user_data['subscriptions']:
            # Добавляем кнопку "Назад", чтобы пользователь не терялся
            back_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            back_markup.add(types.KeyboardButton('🔙 Назад'))
            bot.send_message(
                user_id,
                "❌ У вас пока нет активных ключей.\n\n"
                "Нажмите «Назад» чтобы вернуться в главное меню.",
                reply_markup=back_markup
            )
            return
        
        current_time = datetime.now()
        valid_subscriptions = []
        
        for sub in user_data['subscriptions']:
            try:
                expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                if expiry_date > current_time:
                    valid_subscriptions.append(sub)
            except:
                pass
        
        if not valid_subscriptions:
            back_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            back_markup.add(types.KeyboardButton('🔙 Назад'))
            bot.send_message(
                user_id,
                "❌ У вас нет активных ключей. Срок действия истек.\n\n"
                "Нажмите «Назад» чтобы вернуться в главное меню.",
                reply_markup=back_markup
            )
            return
        
        text = "🔑 <b>Ваши активные VPN ключи:</b>\n\n"
        
        for idx, sub in enumerate(reversed(valid_subscriptions), 1):
            try:
                expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                days_left = (expiry_date - current_time).days
                
                text += f"🔹 <b>Ключ #{idx}</b> - {sub['server']}\n"
                text += f"📅 Действует до: {expiry_date.strftime('%d.%m.%Y')} ({days_left} дн.)\n\n"
            except:
                pass
        
        bot.send_message(user_id, text, parse_mode='HTML')
        
        # Отправляем каждый ключ отдельным сообщением (чистый текст)
        for idx, sub in enumerate(reversed(valid_subscriptions), 1):
            key_data = sub.get('key_data', {})
            vpn_key = key_data.get('key', '') if isinstance(key_data, dict) else ''
            
            if vpn_key:
                # Отправляем информацию о ключе
                bot.send_message(
                    user_id,
                    f"🔑 <b>Ключ #{idx} ({sub['server']})</b>",
                    parse_mode='HTML'
                )
                
                # Отправляем чистый ключ
                bot.send_message(user_id, vpn_key)
                
                # QR-код
                qr_filename = f"qr_mykeys_{user_id}_{idx}.png"
                if generate_qr_code(vpn_key, qr_filename):
                    with open(qr_filename, 'rb') as qr_file:
                        bot.send_photo(
                            user_id,
                            qr_file,
                            caption=f"📱 QR-код для ключа #{idx}"
                        )
                    os.remove(qr_filename)
        
        # Кнопки приложений
        app_markup = types.InlineKeyboardMarkup(row_width=2)
        app_markup.add(
            types.InlineKeyboardButton("📱 Happ Android", url="https://play.google.com/store/apps/details?id=com.happproxy"),
            types.InlineKeyboardButton("📱 Happ iOS", url="https://apps.apple.com/us/app/happ-proxy-utility/id6504287215?l=ru"),
            types.InlineKeyboardButton("📱 V2raytun Android", url="https://play.google.com/store/apps/details?id=com.v2raytun.android"),
            types.InlineKeyboardButton("📱 V2raytun iOS", url="https://apps.apple.com/us/app/v2raytun/id6476628951?l=ru")
        )
        
        bot.send_message(
            user_id,
            "📥 <b>Скачать приложение:</b>",
            parse_mode='HTML',
            reply_markup=app_markup
        )
        
        # В конце показываем главное меню (с кнопкой "Назад" оно и так есть)
        bot.send_message(
            user_id,
            "Выберите следующее действие:",
            reply_markup=main_menu_keyboard()
        )
                
    except Exception as e:
        logger.error(f"Ошибка в my_keys_handler: {e}")
        # При ошибке тоже даём возможность вернуться
        back_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        back_markup.add(types.KeyboardButton('🔙 Назад'))
        bot.send_message(
            message.chat.id,
            "❌ Произошла ошибка при получении списка ключей.\n\n"
            "Нажмите «Назад» чтобы вернуться в главное меню.",
            reply_markup=back_markup
        )
        
@bot.callback_query_handler(func=lambda call: call.data.startswith('copytext_'))
def copy_text_callback(call):
    """Отправляет текст для копирования в буфер обмена"""
    try:
        parts = call.data.split('_')
        user_id = int(parts[1])
        
        # Проверяем что пользователь тот же
        if call.from_user.id != user_id:
            bot.answer_callback_query(call.id, "❌ Это не ваш ключ!")
            return
        
        user_data = get_user_data(user_id)
        
        # Определяем какой ключ копировать
        if len(parts) > 2:
            key_idx = parts[2]
            vpn_key = user_data.get(f'last_key_{key_idx}', '')
        else:
            vpn_key = user_data.get('last_key', '')
        
        if vpn_key:
            # Отправляем ключ в моноширинном формате для копирования
            bot.send_message(
                call.message.chat.id,
                f"<pre><code class=\"language-text\">{vpn_key}</code></pre>\n\n"
                "<i>👆 Нажмите на текст выше чтобы скопировать в буфер обмена</i>",
                parse_mode='HTML'
            )
            bot.answer_callback_query(call.id, "✅ Нажмите на ключ чтобы скопировать!")
        else:
            bot.answer_callback_query(call.id, "❌ Ключ не найден")
            
    except Exception as e:
        logger.error(f"Ошибка в copy_text_callback: {e}")
        bot.answer_callback_query(call.id, "❌ Ошибка")

@bot.callback_query_handler(func=lambda call: call.data.startswith('copy_key_'))
def copy_key_callback(call):
    """Отправляет ключ для копирования"""
    try:
        parts = call.data.split('_')
        user_id = int(parts[2])
        
        # Проверяем что пользователь тот же
        if call.from_user.id != user_id:
            bot.answer_callback_query(call.id, "❌ Это не ваш ключ!")
            return
        
        user_data = get_user_data(user_id)
        
        # Определяем какой ключ копировать
        if len(parts) > 3:
            key_idx = parts[3]
            vpn_key = user_data.get(f'temp_copy_key_{key_idx}', '')
        else:
            vpn_key = user_data.get('temp_copy_key', '')
        
        if vpn_key:
            # Отправляем ключ отдельным сообщением для легкого копирования
            bot.send_message(
                call.message.chat.id,
                f"<pre><code class=\"language-text\">{vpn_key}</code></pre>",
                parse_mode='HTML'
            )
            bot.answer_callback_query(call.id, "✅ Ключ отправлен! Нажмите на него чтобы скопировать")
        else:
            bot.answer_callback_query(call.id, "❌ Ключ не найден")
            
    except Exception as e:
        logger.error(f"Ошибка в copy_key_callback: {e}")
        bot.answer_callback_query(call.id, "❌ Ошибка")

# Модифицируем функцию download_config
@bot.message_handler(func=lambda message: message.text == '💾 Скачать конфиг')
def download_config(message):
    try:
        user_id = message.from_user.id
        user_data = get_user_data(user_id)
        
        if 'subscriptions' not in user_data or not user_data['subscriptions']:
            bot.send_message(user_id, "У вас нет активных подписок.")
            return
        
        last_sub = user_data['subscriptions'][-1]
        config_file = last_sub['config_file']
        
        if not os.path.exists(config_file):
            bot.send_message(user_id, "Файл конфигурации не найден.")
            return
        
        # Проверяем не истекла ли подписка
        expiry_date = datetime.strptime(last_sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
        is_expired = expiry_date < datetime.now()
        
        # Читаем содержимое конфига для QR-кода
        with open(config_file, 'r') as f:
            config_content = f.read()
        
        # Генерируем QR-код
        qr_filename = f"{config_file}_qr.png"
        if generate_qr_code(config_content, qr_filename):
            # Отправляем QR-код
            with open(qr_filename, 'rb') as qr_file:
                bot.send_photo(
                    user_id,
                    qr_file,
                    caption="📲 Отсканируйте этот QR-код в приложении Amnezia для быстрой настройки"
                )
        
        caption = f"Ваш конфигурационный файл для {last_sub['server']}"
        if is_expired:
            caption += "\n⚠️ Срок действия истек! Для продления приобретите новый доступ."
        
        with open(config_file, 'rb') as f:
            bot.send_document(
                user_id,
                f,
                caption=caption,
                visible_file_name=os.path.basename(config_file))
            
    except Exception as e:
        logger.error(f"Ошибка в download_config: {e}")
        bot.send_message(message.chat.id, "Произошла ошибка при отправке файла")

# Запуск бота
def run_bot():
    # Запускаем мониторинг подписок в отдельном потоке
    monitor_thread = threading.Thread(target=subscription_monitor, daemon=True)
    monitor_thread.start()
    
    logger.info("Бот запущен и готов к работе")
    
    while True:
        try:
            logger.info("Запуск polling...")
            # УБИРАЕМ restart_on_change=True или устанавливаем в False
            bot.infinity_polling(timeout=60, long_polling_timeout=60, restart_on_change=False)
        except requests.exceptions.ConnectionError:
            logger.error("Ошибка соединения. Повторная попытка через 15 секунд...")
            time.sleep(15)
        except requests.exceptions.ReadTimeout:
            logger.error("Таймаут соединения. Повторная попытка через 10 секунд...")
            time.sleep(10)
        except Exception as e:
            logger.error(f"Критическая ошибка бота: {e}")
            logger.error("Перезапуск через 30 секунд...")
            time.sleep(30)

@bot.message_handler(func=lambda message: message.text == '🗑 Удалить пользователя' and is_admin(message.from_user.id))
def delete_user_by_id(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        msg = bot.send_message(
            message.chat.id,
            "Введите ID пользователя для полного удаления:\n\n"
            "⚠️ <b>ВНИМАНИЕ!</b> Это удалит:\n"
            "• Все подписки пользователя\n"
            "• Все конфиг файлы пользователя\n"
            "• Все данные пользователя из системы\n\n"
            "Для удаления только одной подписки используйте '🗑 Удалить конкретную подписку'",
            parse_mode='HTML',
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(msg, process_user_deletion)
    except Exception as e:
        logger.error(f"Ошибка в delete_user_by_id: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при удалении пользователя")

def process_user_deletion(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        user_id_to_delete = message.text.strip()
        
        if not user_id_to_delete.isdigit():
            bot.send_message(message.chat.id, "❌ ID пользователя должен содержать только цифры!")
            return
        
        if user_id_to_delete not in users_db:
            bot.send_message(message.chat.id, "❌ Пользователь с таким ID не найден!")
            return
        
        user_data = users_db[user_id_to_delete]
        username = user_data.get('username', 'N/A')
        
        # Подтверждение удаления
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton(f'✅ Да, удалить @{username}'))
        markup.add(types.KeyboardButton('❌ Нет, отменить'))
        
        confirmation_text = (
            f"⚠️ <b>ПОДТВЕРЖДЕНИЕ УДАЛЕНИЯ</b>\n\n"
            f"Вы собираетесь полностью удалить пользователя:\n"
            f"👤 <b>Username:</b> @{username}\n"
            f"🆔 <b>ID:</b> {user_id_to_delete}\n\n"
            f"<b>Будет удалено:</b>\n"
        )
        
        if 'subscriptions' in user_data:
            subscription_count = len(user_data['subscriptions'])
            confirmation_text += f"• {subscription_count} подписок\n"
            for sub in user_data['subscriptions'][:3]:
                key_data = sub.get('key_data', {})
                key_preview = key_data.get('key', '')[:20] + '...' if key_data.get('key') else 'N/A'
                confirmation_text += f"  - {sub['server']}: {key_preview}\n"
            if subscription_count > 3:
                confirmation_text += f"  ... и еще {subscription_count - 3} подписок\n"
        
        confirmation_text += "• Все данные пользователя из системы\n\n"
        confirmation_text += "<b>Вы уверены?</b>"
        
        msg = bot.send_message(
            message.chat.id,
            confirmation_text,
            parse_mode='HTML',
            reply_markup=markup
        )
        
        bot.register_next_step_handler(msg, lambda m: confirm_user_deletion(m, user_id_to_delete, username, user_data))
        
    except Exception as e:
        logger.error(f"Ошибка в process_user_deletion: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при обработке запроса", reply_markup=admin_menu_keyboard())

def confirm_user_deletion(message, user_id_to_delete, username, user_data):
    try:
        if message.text == '❌ Нет, отменить':
            bot.send_message(message.chat.id, "✅ Удаление отменено.", reply_markup=admin_menu_keyboard())
            return
        
        if not message.text.startswith('✅ Да, удалить'):
            bot.send_message(message.chat.id, "❌ Неверный ответ. Удаление отменено.", reply_markup=admin_menu_keyboard())
            return
        
        # Освобождаем ключи пользователя
        freed_keys = 0
        if 'subscriptions' in user_data:
            for sub in user_data['subscriptions']:
                server_name = sub.get('server', '')
                key_data = sub.get('key_data', {})
                
                for server_key, server_data in servers_db.items():
                    if server_data['name'] == server_name:
                        # Удаляем из используемых ключей
                        if 'used_keys' in server_data and user_id_to_delete in server_data['used_keys']:
                            del server_data['used_keys'][user_id_to_delete]
                            freed_keys += 1
                        break
        
        # Удаляем пользователя
        del users_db[user_id_to_delete]
        
        # Удаляем связанные платежи
        payments_to_delete = []
        for payment_id, payment_data in payments_db.items():
            if str(payment_data.get('user_id')) == user_id_to_delete:
                payments_to_delete.append(payment_id)
        
        for payment_id in payments_to_delete:
            del payments_db[payment_id]
        
        save_data_to_file()
        
        # Уведомляем пользователя
        try:
            bot.send_message(
                user_id_to_delete,
                "⚠️ Ваш VPN доступ был полностью отключен администратором.\n\n"
                "Для возобновления работы необходимо приобрести новый доступ."
            )
        except:
            pass
        
        # Отчет
        report_text = (
            f"✅ <b>Пользователь успешно удален!</b>\n\n"
            f"👤 <b>Username:</b> @{username}\n"
            f"🆔 <b>ID:</b> {user_id_to_delete}\n"
            f"🔓 <b>Освобождено ключей:</b> {freed_keys}\n"
            f"🗑 <b>Удалено платежей:</b> {len(payments_to_delete)}"
        )
        
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('👤 Список пользователей'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(message.chat.id, report_text, parse_mode='HTML', reply_markup=keyboard)
        
    except Exception as e:
        logger.error(f"Ошибка в confirm_user_deletion: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при удалении пользователя", reply_markup=admin_menu_keyboard())

@bot.message_handler(func=lambda message: message.text == '🗑 Удалить конфиг пользователя' and is_admin(message.from_user.id))
def delete_user_config(message):
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        # Получаем список пользователей с активными подписками
        active_users = []
        for user_id, user_data in users_db.items():
            if 'subscriptions' in user_data:
                for sub in user_data['subscriptions']:
                    expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                    if expiry_date > datetime.now():
                        active_users.append({
                            'user_id': user_id,
                            'username': user_data.get('username', 'N/A'),
                            'server': sub['server'],
                            'config': sub['config_file']
                        })
        
        if not active_users:
            bot.send_message(message.chat.id, "Нет активных пользователей для удаления.")
            return
        
        # Создаем клавиатуру с пользователями
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        for user in active_users:
            btn_text = f"👤 {user['username']} ({user['server']})"
            keyboard.add(types.KeyboardButton(btn_text))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(
            message.chat.id,
            "Выберите пользователя для удаления конфига:",
            reply_markup=keyboard
        )
        bot.register_next_step_handler(msg, process_user_config_deletion)
    except Exception as e:
        logger.error(f"Ошибка в delete_user_config: {e}")

def process_user_config_deletion(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        # Парсим выбор пользователя
        username = message.text.split(' (')[0].replace('👤 ', '')
        server = message.text.split(' (')[1].replace(')', '')
        
        # Находим пользователя
        found = False
        for user_id, user_data in users_db.items():
            if user_data.get('username') == username and 'subscriptions' in user_data:
                for sub in user_data['subscriptions']:
                    if sub['server'] == server:
                        config_file = sub['config_file']
                        
                        # Освобождаем конфиг на сервере (удаляем полностью)
                        for server_key, server_data in servers_db.items():
                            if server_data['name'] == sub['server']:
                                # Удаляем из используемых
                                if str(user_id) in server_data['used_configs']:
                                    del server_data['used_configs'][str(user_id)]
                                
                                # Удаляем из доступных
                                if config_file in server_data['available_configs']:
                                    server_data['available_configs'].remove(config_file)
                                break
                        
                        # Удаляем файл конфига физически
                        if os.path.exists(config_file):
                            try:
                                os.remove(config_file)
                                
                                # Удаляем QR-код если есть
                                qr_file = f"{config_file}_qr.png"
                                if os.path.exists(qr_file):
                                    os.remove(qr_file)
                            except Exception as e:
                                logger.error(f"Ошибка удаления файла {config_file}: {e}")
                        
                        # Удаляем подписку
                        user_data['subscriptions'].remove(sub)
                        save_data_to_file()
                        
                        # Уведомляем пользователя
                        try:
                            bot.send_message(
                                user_id,
                                f"⚠️ Ваш VPN конфиг для сервера {sub['server']} был удален администратором.\n\n"
                                "Конфигурационный файл удален из системы полностью.\n"
                                "Для возобновления работы необходимо приобрести новый доступ."
                            )
                        except Exception as e:
                            logger.error(f"Не удалось уведомить пользователя: {e}")
                        
                        bot.send_message(
                            message.chat.id,
                            f"✅ Конфиг пользователя @{username} для сервера {sub['server']} успешно удален!\n"
                            f"🗑 Файл конфига удален из системы.",
                            reply_markup=admin_menu_keyboard()
                        )
                        found = True
                        break
                if found:
                    break
        
        if not found:
            bot.send_message(message.chat.id, "❌ Пользователь не найден.", reply_markup=admin_menu_keyboard())
            
    except Exception as e:
        logger.error(f"Ошибка в process_user_config_deletion: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка при удалении конфига", reply_markup=admin_menu_keyboard())
        
@bot.callback_query_handler(func=lambda call: call.data == "download_app_menu")
def handle_download_app_callback(call):
    """Обработчик нажатия на кнопку скачивания приложения"""
    try:
        user_id = call.from_user.id
        
        # Создаем inline-клавиатуру со ссылками для скачивания
        markup = types.InlineKeyboardMarkup()
        markup.row(
            types.InlineKeyboardButton("📱 Happ Android", url="https://play.google.com/store/apps/details?id=com.happproxy"),
            types.InlineKeyboardButton("📱 Happ iOS", url="https://apps.apple.com/us/app/happ-proxy-utility/id6504287215?l=ru"),
            types.InlineKeyboardButton("📱 V2raytun Android", url="https://play.google.com/store/apps/details?id=com.v2raytun.android"),
            types.InlineKeyboardButton("📱 V2raytun iOS", url="https://apps.apple.com/us/app/v2raytun/id6476628951?l=ru")
        )
        
        bot.send_message(
            user_id,
            "📥 Выберите вашу платформу для скачивания Happ или V2raytun:",
            reply_markup=markup
        )
        
        bot.answer_callback_query(call.id)
        
    except Exception as e:
        logger.error(f"Ошибка в handle_download_app_callback: {e}")
        bot.answer_callback_query(call.id, "❌ Ошибка при открытии меню скачивания")
# Добавляем команду для отключения конфига
@bot.message_handler(func=lambda message: message.text.startswith('🚫 Отключить конфиг') and is_admin(message.from_user.id))
def disable_config(message):
    try:
        config_file = message.text.replace('🚫 Отключить конфиг ', '')
        
        # Находим сервер и пользователя для этого конфига
        for server in servers_db.values():
            if config_file in server['used_configs'].values():
                user_id = next(uid for uid, cfg in server['used_configs'].items() if cfg == config_file)
                
                # Возвращаем конфиг в доступные
                server['available_configs'].append(config_file)
                del server['used_configs'][user_id]
                save_data_to_file()
                
                # Уведомляем пользователя
                try:
                    bot.send_message(user_id, f"""
⚠️ Ваш VPN конфиг {config_file} был отключен.
Для возобновления работы необходимо приобрести новый доступ.
""")
                except Exception as e:
                    logger.error(f"Не удалось уведомить пользователя {user_id}: {e}")
                
                bot.send_message(message.chat.id, f"Конфиг {config_file} успешно отключен.", reply_markup=admin_menu_keyboard())
                return
        
        bot.send_message(message.chat.id, "Конфиг не найден среди используемых.")
    except Exception as e:
        logger.error(f"Ошибка в disable_config: {e}")

# В функции process_config_deletion добавляем проверку на используемые конфиги
def process_config_deletion(message):
    try:
        if message.text == '🔙 Назад':
            bot.send_message(message.chat.id, "Отменено.", reply_markup=admin_menu_keyboard())
            return
        
        config_name = message.text
        if os.path.exists(config_name):
            # Проверяем, не используется ли конфиг
            for server in servers_db.values():
                if config_name in server['used_configs'].values():
                    bot.send_message(message.chat.id, "Этот конфиг используется и не может быть удален!")
                    return
                
                if config_name in server['available_configs']:
                    server['available_configs'].remove(config_name)
            
            os.remove(config_name)
            save_data_to_file()
            bot.send_message(message.chat.id, f"Конфигурация {config_name} успешно удалена!", reply_markup=admin_menu_keyboard())
        else:
            bot.send_message(message.chat.id, "Файл не найден.")
    except Exception as e:
        logger.error(f"Ошибка в process_config_deletion: {e}")
        
@bot.message_handler(func=lambda message: message.text == '🎁 Реферальная система' and is_admin(message.from_user.id))
def admin_referral_panel(message):
    """Админ-панель управления рефералами"""
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        keyboard.add(
            types.KeyboardButton('➕ Добавить в рефералы'),
            types.KeyboardButton('➖ Удалить из рефералов'),
            types.KeyboardButton('📋 Список участников'),
            types.KeyboardButton('💰 Заявки на вывод'),
            types.KeyboardButton('🔙 Назад')
        )
        
        # Считаем статистику
        total_balance = sum(u.get('balance', 0) for u in referral_db['allowed_users'].values())
        pending_requests = len([r for r in referral_db['withdraw_requests'] if r['status'] == 'pending'])
        
        text = f"""🎁 <b>Реферальная система (Админ)</b>

👥 <b>Участников:</b> {len(referral_db['allowed_users'])}
💰 <b>Общий баланс к выплате:</b> {total_balance}₽
⏳ <b>Заявок на вывод:</b> {pending_requests}

Выберите действие:"""
        
        bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=keyboard)
        
    except Exception as e:
        logger.error(f"Ошибка в admin_referral_panel: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка", reply_markup=admin_menu_keyboard())

@bot.message_handler(func=lambda message: message.text == '➕ Добавить в рефералы' and is_admin(message.from_user.id))
def add_user_to_referral(message):
    try:
        msg = bot.send_message(
            message.chat.id,
            "👤 Введите <b>Telegram ID</b> пользователя, которому разрешить участвовать в реферальной системе:",
            parse_mode='HTML',
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(msg, process_add_referral_user)
    except Exception as e:
        logger.error(f"Ошибка в add_user_to_referral: {e}")

def process_add_referral_user(message):
    try:
        if message.text == '🔙 Назад':
            admin_panel_handler(message)
            return
            
        user_id = message.text.strip()
        if not user_id.isdigit():
            bot.send_message(message.chat.id, "❌ ID должен состоять из цифр.", reply_markup=admin_menu_keyboard())
            return

        if user_id in referral_db['allowed_users']:
            bot.send_message(message.chat.id, f"❌ Пользователь {user_id} уже участвует в программе.", reply_markup=admin_menu_keyboard())
            return
            
        # Генерируем уникальный реферальный код
        import hashlib
        ref_code = hashlib.md5(f"{user_id}{time.time()}".encode()).hexdigest()[:8]
        
        referral_db['allowed_users'][user_id] = {
            'balance': 0,
            'ref_code': ref_code,
            'added_by': str(message.from_user.id),
            'added_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        referral_db['referrals'][ref_code] = user_id
        save_referral_data()
        
        # Уведомляем пользователя
        try:
            bot.send_message(
                user_id,
                f"🎉 <b>Вам открыт доступ к реферальной системе!</b>\n\n"
                f"🔗 Ваша реферальная ссылка:\n<code>https://t.me/{(bot.get_me().username)}?start={ref_code}</code>\n\n"
                f"💵 Вы будете получать <b>25%</b> от каждой оплаты приглашенных пользователей.\n"
                f"💰 Баланс и вывод доступны в главном меню по кнопке «👥 Рефералы».",
                parse_mode='HTML'
            )
        except:
            pass
            
        bot.send_message(
            message.chat.id,
            f"✅ Пользователь <code>{user_id}</code> добавлен в реферальную систему!\n"
            f"Код: <code>{ref_code}</code>",
            parse_mode='HTML',
            reply_markup=admin_menu_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка в process_add_referral_user: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка", reply_markup=admin_menu_keyboard())

@bot.message_handler(func=lambda message: message.text == '➖ Удалить из рефералов' and is_admin(message.from_user.id))
def remove_user_from_referral(message):
    try:
        if not referral_db['allowed_users']:
            bot.send_message(message.chat.id, "❌ Список участников пуст.", reply_markup=admin_menu_keyboard())
            return
            
        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        for uid, data in referral_db['allowed_users'].items():
            # Пытаемся получить username
            uname = f"ID:{uid}"
            try:
                user_info = bot.get_chat_member(uid, uid)
                if user_info.user.username:
                    uname = f"@{user_info.user.username}"
            except:
                pass
            keyboard.add(types.KeyboardButton(f"❌ {uname} | Баланс: {data['balance']}₽"))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        msg = bot.send_message(message.chat.id, "Выберите пользователя для удаления из программы:", reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_remove_referral_user)
    except Exception as e:
        logger.error(f"Ошибка в remove_user_from_referral: {e}")

def process_remove_referral_user(message):
    try:
        if message.text == '🔙 Назад':
            admin_panel_handler(message)
            return
            
        # Извлекаем ID из строки вида "❌ @username | Баланс: 100₽"
        text = message.text
        if 'ID:' in text:
            user_id = text.split('ID:')[1].split(' ')[0].strip()
        else:
            # Пытаемся найти по username (менее надежно)
            bot.send_message(message.chat.id, "❌ Не удалось определить ID.", reply_markup=admin_menu_keyboard())
            return
            
        if user_id in referral_db['allowed_users']:
            del referral_db['allowed_users'][user_id]
            # Удаляем также реф код
            refs_to_del = [k for k, v in referral_db['referrals'].items() if v == user_id]
            for k in refs_to_del:
                del referral_db['referrals'][k]
            save_referral_data()
            
            try:
                bot.send_message(user_id, "⚠️ Ваше участие в реферальной программе приостановлено администратором.")
            except:
                pass
                
            bot.send_message(message.chat.id, f"✅ Пользователь {user_id} удален из реферальной программы.", reply_markup=admin_menu_keyboard())
        else:
            bot.send_message(message.chat.id, "❌ Пользователь не найден.", reply_markup=admin_menu_keyboard())
            
    except Exception as e:
        logger.error(f"Ошибка в process_remove_referral_user: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка", reply_markup=admin_menu_keyboard())

@bot.message_handler(func=lambda message: message.text == '📋 Список участников' and is_admin(message.from_user.id))
def list_referral_users(message):
    try:
        if not referral_db['allowed_users']:
            bot.send_message(message.chat.id, "❌ Список участников пуст.")
            return
            
        text = "📋 <b>Участники реферальной системы:</b>\n\n"
        for uid, data in referral_db['allowed_users'].items():
            uname = f"ID:{uid}"
            try:
                user_info = bot.get_chat_member(uid, uid)
                if user_info.user.username:
                    uname = f"@{user_info.user.username}"
            except:
                pass
            text += f"👤 {uname}\n"
            text += f"   💰 Баланс: {data['balance']}₽\n"
            text += f"   🔗 Код: <code>{data['ref_code']}</code>\n\n"
            
            if len(text) > 3500:
                bot.send_message(message.chat.id, text, parse_mode='HTML')
                text = ""
                time.sleep(0.5)
                
        if text:
            bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=admin_menu_keyboard())
            
    except Exception as e:
        logger.error(f"Ошибка в list_referral_users: {e}")

@bot.message_handler(func=lambda message: message.text == '💰 Заявки на вывод' and is_admin(message.from_user.id))
def admin_withdraw_requests(message):
    """Просмотр заявок на вывод для админа"""
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        # Проверяем существование структуры
        if 'withdraw_requests' not in referral_db:
            referral_db['withdraw_requests'] = []
            save_referral_data()
        
        # Фильтруем только pending заявки
        pending = [r for r in referral_db['withdraw_requests'] if r.get('status') == 'pending']
        
        if not pending:
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            keyboard.add(types.KeyboardButton('📋 Все заявки'))  # Кнопка для просмотра всех заявок
            keyboard.add(types.KeyboardButton('🔙 Назад'))
            
            bot.send_message(
                message.chat.id, 
                "✅ Нет активных заявок на вывод.\n\n"
                "Нажмите «📋 Все заявки» чтобы посмотреть историю.",
                reply_markup=keyboard
            )
            return
        
        bot.send_message(
            message.chat.id,
            f"💰 <b>Активные заявки на вывод ({len(pending)}):</b>\n\n"
            "Выберите действие с заявкой:",
            parse_mode='HTML'
        )
        
        for req in pending:
            req_id = req.get('id', 'N/A')
            req_user_id = req.get('user_id', 'N/A')
            req_username = req.get('username', f'ID:{req_user_id}')
            req_amount = req.get('amount', 0)
            req_contact = req.get('contact', 'N/A')
            req_date = req.get('date', 'N/A')
            
            text = f"""💰 <b>Заявка #{req_id}</b>

👤 <b>Пользователь:</b> {req_username}
🆔 <b>ID:</b> <code>{req_user_id}</code>
💵 <b>Сумма:</b> {req_amount}₽
📞 <b>Контакт:</b> <code>{req_contact}</code>
📅 <b>Дата:</b> {req_date}

<b>Действия:</b>"""
            
            markup = types.InlineKeyboardMarkup(row_width=2)
            markup.add(
                types.InlineKeyboardButton("✅ Выполнено", callback_data=f"wd_approve_{req_id}"),
                types.InlineKeyboardButton("❌ Отклонить", callback_data=f"wd_reject_{req_id}")
            )
            
            bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=markup)
            time.sleep(0.3)  # Небольшая задержка между сообщениями
        
        # Клавиатура для возврата
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('📋 Все заявки'))
        keyboard.add(types.KeyboardButton('🎁 Реферальная система'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(
            message.chat.id,
            "Выберите действие или вернитесь назад.",
            reply_markup=keyboard
        )
        
    except Exception as e:
        logger.error(f"Ошибка в admin_withdraw_requests: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())
        
@bot.message_handler(func=lambda message: message.text == '📋 Все заявки' and is_admin(message.from_user.id))
def admin_all_withdraw_requests(message):
    """Просмотр ВСЕХ заявок на вывод (история)"""
    try:
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        if 'withdraw_requests' not in referral_db or not referral_db['withdraw_requests']:
            bot.send_message(message.chat.id, "📭 Нет заявок на вывод.", reply_markup=admin_menu_keyboard())
            return
        
        all_requests = referral_db['withdraw_requests']
        
        # Сортируем по дате (новые сверху)
        all_requests.sort(key=lambda x: x.get('date', ''), reverse=True)
        
        text = f"📋 <b>История заявок на вывод ({len(all_requests)}):</b>\n\n"
        
        for req in all_requests[:20]:  # Показываем последние 20
            req_id = req.get('id', 'N/A')
            req_username = req.get('username', 'N/A')
            req_amount = req.get('amount', 0)
            req_status = req.get('status', 'pending')
            
            status_emoji = {
                'pending': '⏳',
                'completed': '✅',
                'rejected': '❌'
            }.get(req_status, '❓')
            
            text += f"{status_emoji} <b>#{req_id}</b> - {req_username} - {req_amount}₽ - {req_status}\n"
        
        if len(all_requests) > 20:
            text += f"\n... и еще {len(all_requests) - 20} заявок"
        
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        keyboard.add(types.KeyboardButton('💰 Заявки на вывод'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=keyboard)
        
    except Exception as e:
        logger.error(f"Ошибка в admin_all_withdraw_requests: {e}")
        bot.send_message(message.chat.id, f"❌ Ошибка: {str(e)}", reply_markup=admin_menu_keyboard())
        
@bot.message_handler(func=lambda message: message.text == '📋 Моя статистика')
def referral_statistics(message):
    """Показывает подробную статистику реферальной системы для пользователя"""
    try:
        user_id = str(message.from_user.id)
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        # Проверяем, участвует ли пользователь в программе
        if user_id not in referral_db.get('allowed_users', {}):
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            keyboard.add(types.KeyboardButton('📝 Подать заявку на участие'))
            keyboard.add(types.KeyboardButton('🔙 Назад'))
            
            bot.send_message(
                message.chat.id,
                "❌ У вас нет доступа к реферальной статистике.\n"
                "Подайте заявку на участие в программе.",
                reply_markup=keyboard
            )
            return
        
        user_data = referral_db['allowed_users'][user_id]
        
        # Подсчет рефералов
        total_referrals = 0
        active_referrals = 0
        referrals_list = []
        
        if 'relations' in referral_db:
            for referred_id, referrer_id in referral_db['relations'].items():
                if referrer_id == user_id:
                    total_referrals += 1
                    
                    # Получаем информацию о реферале
                    referred_user_data = get_user_data(referred_id)
                    referred_username = referred_user_data.get('username', f'ID:{referred_id}')
                    
                    # Проверяем, есть ли у реферала активные подписки
                    has_active_sub = False
                    total_spent = 0
                    
                    if 'subscriptions' in referred_user_data:
                        current_time = datetime.now()
                        for sub in referred_user_data['subscriptions']:
                            # Проверяем активность подписки
                            try:
                                expiry_date = datetime.strptime(sub['expiry_date'], "%Y-%m-%d %H:%M:%S")
                                if expiry_date > current_time:
                                    has_active_sub = True
                            except:
                                pass
                            
                            # Считаем общую сумму потраченную рефералом (только платные подписки)
                            if sub.get('type') == 'paid' and 'payment_id' in sub:
                                # Ищем платеж в базе
                                payment = get_payment(sub['payment_id'])
                                if payment:
                                    amount_str = payment.get('amount', '0₽')
                                    import re
                                    digits = re.findall(r'\d+', str(amount_str))
                                    amount_num = int(digits[0]) if digits else 0
                                    total_spent += amount_num
                    
                    if has_active_sub:
                        active_referrals += 1
                    
                    referrals_list.append({
                        'id': referred_id,
                        'username': referred_username,
                        'has_active_sub': has_active_sub,
                        'total_spent': total_spent
                    })
        
        # Считаем заработанные бонусы
        total_bonus_earned = 0
        if 'relations' in referral_db:
            for referred_id, referrer_id in referral_db['relations'].items():
                if referrer_id == user_id:
                    referred_user_data = get_user_data(referred_id)
                    if 'subscriptions' in referred_user_data:
                        for sub in referred_user_data['subscriptions']:
                            if sub.get('type') == 'paid' and 'payment_id' in sub:
                                payment = get_payment(sub['payment_id'])
                                if payment:
                                    amount_str = payment.get('amount', '0₽')
                                    import re
                                    digits = re.findall(r'\d+', str(amount_str))
                                    amount_num = int(digits[0]) if digits else 0
                                    total_bonus_earned += int(amount_num * 0.25)
        
        current_balance = user_data.get('balance', 0)
        withdrawn_amount = total_bonus_earned - current_balance
        
        # Статистика выводов
        withdrawals = []
        if 'withdraw_requests' in referral_db:
            for req in referral_db['withdraw_requests']:
                if req.get('user_id') == user_id:
                    withdrawals.append(req)
        
        # Формируем текст статистики
        text = f"""📊 <b>МОЯ РЕФЕРАЛЬНАЯ СТАТИСТИКА</b>

👥 <b>Всего приглашено:</b> {total_referrals} чел.
✅ <b>Активных рефералов:</b> {active_referrals} чел.
💤 <b>Неактивных:</b> {total_referrals - active_referrals} чел.

💰 <b>Финансы:</b>
• Заработано всего: {total_bonus_earned}₽
• Текущий баланс: {current_balance}₽
• Выведено: {withdrawn_amount}₽

"""
        
        # Добавляем список рефералов
        if referrals_list:
            text += "<b>📋 Список рефералов:</b>\n"
            for i, ref in enumerate(referrals_list[:10], 1):
                status = "🟢" if ref['has_active_sub'] else "⚪"
                text += f"{i}. {status} @{ref['username']} | Потратил: {ref['total_spent']}₽\n"
            
            if len(referrals_list) > 10:
                text += f"... и еще {len(referrals_list) - 10} рефералов\n"
        else:
            text += "<b>📋 У вас пока нет приглашенных пользователей</b>\n"
        
        # Добавляем историю выводов
        if withdrawals:
            text += "\n<b>💸 История выводов:</b>\n"
            for w in sorted(withdrawals, key=lambda x: x.get('date', ''), reverse=True)[:5]:
                status_emoji = {
                    'pending': '⏳',
                    'completed': '✅',
                    'rejected': '❌'
                }.get(w.get('status', 'pending'), '❓')
                
                text += f"{status_emoji} {w.get('amount', 0)}₽ - {w.get('date', 'N/A')} - {w.get('status', 'pending')}\n"
        
        text += f"""
💡 <b>Как заработать больше:</b>
1. Отправляйте реферальную ссылку друзьям
2. Вы получаете 25% от каждой их покупки
3. Чем больше активных рефералов - тем больше доход!

🔗 <b>Ваша ссылка:</b>
<code>https://t.me/{(bot.get_me().username)}?start={user_data['ref_code']}</code>"""
        
        # Создаем клавиатуру
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        if current_balance > 0:
            keyboard.add(types.KeyboardButton('💸 Вывести средства'))
        keyboard.add(types.KeyboardButton('👥 Рефералы'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        # Если сообщение слишком длинное, разбиваем на части
        if len(text) > 4000:
            parts = []
            lines = text.split('\n')
            current_part = ""
            
            for line in lines:
                if len(current_part + line + '\n') > 3800:
                    parts.append(current_part)
                    current_part = line + '\n'
                else:
                    current_part += line + '\n'
            
            if current_part:
                parts.append(current_part)
            
            for i, part in enumerate(parts):
                if i == len(parts) - 1:
                    bot.send_message(message.chat.id, part, parse_mode='HTML', reply_markup=keyboard)
                else:
                    bot.send_message(message.chat.id, part, parse_mode='HTML')
                time.sleep(0.5)
        else:
            bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=keyboard)
        
    except Exception as e:
        logger.error(f"Ошибка в referral_statistics: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        bot.send_message(
            message.chat.id,
            "❌ Произошла ошибка при загрузке статистики.",
            reply_mup=main_menu_keyboard()
        )
        
@bot.callback_query_handler(func=lambda call: call.data.startswith('wd_'))
def handle_withdraw_action(call):
    try:
        if not is_admin(call.from_user.id):
            bot.answer_callback_query(call.id, "⛔ Доступ запрещен")
            return
            
        action, req_id = call.data.split('_')[1], call.data.split('_')[2]
        
        # Находим заявку
        req = None
        req_index = -1
        for i, r in enumerate(referral_db['withdraw_requests']):
            if r['id'] == req_id:
                req = r
                req_index = i
                break
                
        if not req:
            bot.answer_callback_query(call.id, "Заявка не найдена")
            return
            
        user_id = req['user_id']
        
        if action == 'approve':
            referral_db['withdraw_requests'][req_index]['status'] = 'completed'
            save_referral_data()
            
            try:
                bot.send_message(user_id, f"✅ Ваша заявка на вывод {req['amount']}₽ выполнена! Средства отправлены на указанные реквизиты.")
            except:
                pass
                
            bot.edit_message_text(
                f"✅ Заявка #{req_id} отмечена как выполненная.",
                call.message.chat.id,
                call.message.message_id
            )
            
        elif action == 'reject':
            # Возвращаем деньги на баланс
            if user_id in referral_db['allowed_users']:
                referral_db['allowed_users'][user_id]['balance'] += req['amount']
                
            referral_db['withdraw_requests'][req_index]['status'] = 'rejected'
            save_referral_data()
            
            try:
                bot.send_message(user_id, f"❌ Ваша заявка на вывод {req['amount']}₽ отклонена администратором. Средства возвращены на баланс.")
            except:
                pass
                
            bot.edit_message_text(
                f"❌ Заявка #{req_id} отклонена.",
                call.message.chat.id,
                call.message.message_id
            )
            
        bot.answer_callback_query(call.id, "Готово")
        
    except Exception as e:
        logger.error(f"Ошибка в handle_withdraw_action: {e}")
        bot.answer_callback_query(call.id, "Ошибка")

# ==================== ПОЛЬЗОВАТЕЛЬСКАЯ ЧАСТЬ ====================

@bot.message_handler(func=lambda message: message.text == '👥 Рефералы')
def user_referral_menu(message):
    """Меню рефералов для пользователя"""
    try:
        user_id = str(message.from_user.id)
        delete_previous_message(message.chat.id, message.message_id - 1)
        
        if user_id not in referral_db.get('allowed_users', {}):
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            keyboard.add(types.KeyboardButton('📝 Подать заявку на участие'))
            keyboard.add(types.KeyboardButton('🔙 Назад'))
            
            text = """🎁 <b>Реферальная система</b>

💰 Получайте <b>25%</b> от всех оплат приглашенных друзей!

❌ У вас пока нет доступа к реферальной программе.
Нажмите кнопку ниже, чтобы подать заявку на участие."""
            bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=keyboard)
            return
            
        # Пользователь участвует
        user_data = referral_db['allowed_users'][user_id]
        ref_link = f"https://t.me/{(bot.get_me().username)}?start={user_data['ref_code']}"
        
        # Правильный подсчет рефералов - только те, кто зарегистрировался по ссылке ЭТОГО пользователя
        ref_count = 0
        if 'relations' in referral_db:
            for referred_id, referrer_id in referral_db['relations'].items():
                if referrer_id == user_id:
                    ref_count += 1
        
        # Получаем баланс
        balance = user_data.get('balance', 0)
        
        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        if balance > 0:
            keyboard.add(types.KeyboardButton('💸 Вывести средства'))
        keyboard.add(types.KeyboardButton('📋 Моя статистика'))
        keyboard.add(types.KeyboardButton('🔙 Назад'))
        
        text = f"""👥 <b>Реферальная система</b>

💰 <b>Ваш баланс:</b> {balance}₽
👤 <b>Приглашено друзей:</b> {ref_count}
💵 <b>Процент отчислений:</b> 25%

🔗 <b>Ваша реферальная ссылка:</b>
<code>{ref_link}</code>

📊 Вы получаете 25% от каждой оплаты приглашенных пользователей.
Средства можно вывести в любой момент.

💡 <b>Как это работает:</b>
1. Отправьте ссылку другу
2. Друг регистрируется в боте
3. Когда друг оплачивает подписку, вы получаете 25% на баланс
4. Накопленные средства можно вывести"""
        
        bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=keyboard)
        
    except Exception as e:
        logger.error(f"Ошибка в user_referral_menu: {e}")
        bot.send_message(message.chat.id, "❌ Ошибка", reply_markup=main_menu_keyboard())

@bot.message_handler(func=lambda message: message.text == '📝 Подать заявку на участие')
def request_referral_access(message):
    try:
        user_id = str(message.from_user.id)
        username = message.from_user.username or "N/A"
        
        # Проверяем, не отправлял ли уже заявку (опционально)
        
        # Отправляем админу
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("✅ Одобрить", callback_data=f"ref_approve_{user_id}"))
        
        admin_text = f"""📝 <b>Заявка на участие в реферальной системе</b>

👤 Пользователь: @{username} (ID: <code>{user_id}</code>)

Нажмите кнопку ниже, чтобы одобрить заявку:"""
        
        bot.send_message(ADMIN_ID, admin_text, parse_mode='HTML', reply_markup=markup)
        
        bot.send_message(
            message.chat.id,
            "✅ Ваша заявка на участие в реферальной системе отправлена администратору. Ожидайте подтверждения.",
            reply_markup=main_menu_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка в request_referral_access: {e}")

@bot.callback_query_handler(func=lambda call: call.data.startswith('ref_approve_'))
def approve_referral_request(call):
    try:
        if not is_admin(call.from_user.id):
            bot.answer_callback_query(call.id, "⛔ Доступ запрещен")
            return
            
        user_id = call.data.split('_')[2]
        
        if user_id in referral_db['allowed_users']:
            bot.answer_callback_query(call.id, "Пользователь уже участвует")
            bot.edit_message_text("✅ Заявка уже была одобрена ранее.", call.message.chat.id, call.message.message_id)
            return
            
        # Генерируем код
        import hashlib
        ref_code = hashlib.md5(f"{user_id}{time.time()}".encode()).hexdigest()[:8]
        
        referral_db['allowed_users'][user_id] = {
            'balance': 0,
            'ref_code': ref_code,
            'added_by': str(call.from_user.id),
            'added_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        referral_db['referrals'][ref_code] = user_id
        save_referral_data()
        
        try:
            bot.send_message(
                user_id,
                f"🎉 <b>Ваша заявка на участие в реферальной системе одобрена!</b>\n\n"
                f"🔗 Ваша реферальная ссылка:\n<code>https://t.me/{(bot.get_me().username)}?start={ref_code}</code>\n\n"
                f"Нажмите «👥 Рефералы» в главном меню, чтобы увидеть баланс.",
                parse_mode='HTML'
            )
        except:
            pass
            
        bot.edit_message_text(
            f"✅ Заявка пользователя {user_id} одобрена!\nКод: <code>{ref_code}</code>",
            call.message.chat.id,
            call.message.message_id,
            parse_mode='HTML'
        )
        bot.answer_callback_query(call.id, "Пользователь добавлен в реферальную систему")
        
    except Exception as e:
        logger.error(f"Ошибка в approve_referral_request: {e}")

@bot.message_handler(func=lambda message: message.text == '💸 Вывести средства')
def withdraw_funds(message):
    try:
        user_id = str(message.from_user.id)
        if user_id not in referral_db['allowed_users']:
            bot.send_message(message.chat.id, "❌ Вы не участвуете в реферальной программе.")
            return
            
        balance = referral_db['allowed_users'][user_id]['balance']
        if balance <= 0:
            bot.send_message(message.chat.id, "❌ На вашем балансе недостаточно средств для вывода.")
            return
            
        msg = bot.send_message(
            message.chat.id,
            f"💰 <b>Заявка на вывод средств</b>\n\n"
            f"Доступно к выводу: {balance}₽\n\n"
            f"Введите сумму для вывода (целое число) или отправьте /cancel для отмены:",
            parse_mode='HTML',
            reply_markup=types.ReplyKeyboardRemove()
        )
        bot.register_next_step_handler(msg, process_withdraw_amount)
        
    except Exception as e:
        logger.error(f"Ошибка в withdraw_funds: {e}")

def process_withdraw_amount(message):
    try:
        if message.text == '/cancel':
            bot.send_message(message.chat.id, "❌ Операция отменена.", reply_markup=main_menu_keyboard())
            return
            
        user_id = str(message.from_user.id)
        balance = referral_db['allowed_users'][user_id]['balance']
        
        try:
            amount = int(message.text.strip())
        except:
            bot.send_message(message.chat.id, "❌ Введите целое число.")
            return
            
        if amount <= 0:
            bot.send_message(message.chat.id, "❌ Сумма должна быть больше 0.")
            return
            
        if amount > balance:
            bot.send_message(message.chat.id, f"❌ Недостаточно средств. Доступно: {balance}₽")
            return
            
        # Сохраняем сумму в стейт
        user_state = get_user_data(message.from_user.id)
        user_state['withdraw_amount'] = amount
        save_user_data(message.from_user.id, user_state)
        
        msg = bot.send_message(
            message.chat.id,
            "📞 Введите номер телефона или карты для перевода средств (только цифры):"
        )
        bot.register_next_step_handler(msg, process_withdraw_contact)
        
    except Exception as e:
        logger.error(f"Ошибка в process_withdraw_amount: {e}")

def process_withdraw_contact(message):
    """Обрабатывает ввод контактных данных для вывода средств"""
    try:
        user_id = str(message.from_user.id)
        user_state = get_user_data(message.from_user.id)
        amount = user_state.get('withdraw_amount', 0)
        
        if not amount:
            bot.send_message(message.chat.id, "❌ Ошибка. Попробуйте снова.", reply_markup=main_menu_keyboard())
            return
            
        contact = message.text.strip()
        if not contact:
            bot.send_message(message.chat.id, "❌ Введите корректные данные.")
            return
        
        # Проверяем, что пользователь есть в allowed_users
        if user_id not in referral_db.get('allowed_users', {}):
            bot.send_message(message.chat.id, "❌ Вы не участвуете в реферальной программе.", reply_markup=main_menu_keyboard())
            return
        
        # Проверяем, что баланс достаточный
        current_balance = referral_db['allowed_users'][user_id].get('balance', 0)
        if amount > current_balance:
            bot.send_message(
                message.chat.id, 
                f"❌ Недостаточно средств. Доступно: {current_balance}₽",
                reply_markup=main_menu_keyboard()
            )
            return
        
        # Списание с баланса
        referral_db['allowed_users'][user_id]['balance'] = current_balance - amount
        new_balance = referral_db['allowed_users'][user_id]['balance']
        
        logger.info(f"💰 ВЫВОД: Списали {amount}₽ с баланса пользователя {user_id}. Новый баланс: {new_balance}₽")
        
        # Создание заявки
        import uuid
        req_id = str(uuid.uuid4())[:8]
        request_data = {
            'id': req_id,
            'user_id': user_id,
            'username': message.from_user.username or f'ID:{user_id}',
            'amount': amount,
            'contact': contact,
            'date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'status': 'pending'
        }
        
        # Инициализируем список заявок если его нет
        if 'withdraw_requests' not in referral_db:
            referral_db['withdraw_requests'] = []
        
        referral_db['withdraw_requests'].append(request_data)
        
        # Сохраняем изменения
        save_referral_data()
        logger.info(f"💰 ВЫВОД: Заявка #{req_id} создана и сохранена в referral_db.json")
        
        # ==================== УВЕДОМЛЕНИЕ АДМИНУ ====================
        try:
            # Получаем username пользователя
            user_username = message.from_user.username
            user_mention = f"@{user_username}" if user_username else f"ID: {user_id}"
            
            admin_text = f"""💰 <b>НОВАЯ ЗАЯВКА НА ВЫВОД #{req_id}</b>

👤 <b>Пользователь:</b> {user_mention}
🆔 <b>Telegram ID:</b> <code>{user_id}</code>
💵 <b>Сумма к выплате:</b> {amount}₽
📞 <b>Контакт для перевода:</b> <code>{contact}</code>
📅 <b>Дата запроса:</b> {request_data['date']}
💳 <b>Остаток на балансе:</b> {new_balance}₽

⚡ <b>Действия:</b>"""
            
            # Создаем inline-клавиатуру для быстрой обработки
            markup = types.InlineKeyboardMarkup(row_width=2)
            markup.add(
                types.InlineKeyboardButton("✅ Выполнено", callback_data=f"wd_approve_{req_id}"),
                types.InlineKeyboardButton("❌ Отклонить", callback_data=f"wd_reject_{req_id}")
            )
            
            # Отправляем админу
            bot.send_message(ADMIN_ID, admin_text, parse_mode='HTML', reply_markup=markup)
            logger.info(f"💰 ВЫВОД: Уведомление админу {ADMIN_ID} о заявке #{req_id} отправлено")
            
        except Exception as admin_err:
            logger.error(f"💰 ВЫВОД: Ошибка отправки уведомления админу: {admin_err}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
        # ==================== КОНЕЦ УВЕДОМЛЕНИЯ ====================
        
        # Уведомление пользователю
        user_text = f"""✅ <b>Заявка на вывод средств принята!</b>

💰 <b>Сумма:</b> {amount}₽
📞 <b>Контакт:</b> <code>{contact}</code>
🆔 <b>Номер заявки:</b> <code>{req_id}</code>
📅 <b>Дата:</b> {request_data['date']}

⏳ Ожидайте, администратор обработает заявку в ближайшее время.
💡 Средства поступят на указанные реквизиты.

📊 <b>Остаток на балансе:</b> {new_balance}₽"""
        
        bot.send_message(
            message.chat.id,
            user_text,
            parse_mode='HTML',
            reply_markup=main_menu_keyboard()
        )
        
        logger.info(f"💰 ВЫВОД: Пользователь {user_id} уведомлен о создании заявки #{req_id}")
        
        # Очищаем стейт
        if 'withdraw_amount' in user_state:
            del user_state['withdraw_amount']
        save_user_data(message.from_user.id, user_state)
        
    except Exception as e:
        logger.error(f"💰 ВЫВОД: Критическая ошибка в process_withdraw_contact: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        # В случае ошибки возвращаем деньги пользователю
        try:
            if user_id in referral_db.get('allowed_users', {}):
                referral_db['allowed_users'][user_id]['balance'] = current_balance
                save_referral_data()
                logger.info(f"💰 ВЫВОД: Возвращены средства {amount}₽ пользователю {user_id} из-за ошибки")
        except:
            pass
            
        bot.send_message(
            message.chat.id,
            "❌ Произошла ошибка при создании заявки. Средства возвращены на баланс. Попробуйте позже.",
            reply_markup=main_menu_keyboard()
        )

if __name__ == '__main__':
    # Синхронизируем конфиги при запуске
    logger.info("Запуск бота...")
    
    # Синхронизация конфигов
    reload_all_data()
    
    # Запускаем мониторинг подписок в отдельном потоке
    monitor_thread = threading.Thread(target=subscription_monitor, daemon=True)
    monitor_thread.start()
    
    logger.info("Бот запущен и готов к работе")
    
    # Запускаем бота
    try:
        bot.infinity_polling(timeout=60, long_polling_timeout=60)
    except Exception as e:
        logger.error(f"Ошибка при запуске бота: {e}")
        time.sleep(5)